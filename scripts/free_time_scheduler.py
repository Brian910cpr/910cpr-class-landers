#!/usr/bin/env python3
"""Read-only preview scheduler for 910CPR free-time teaching windows."""

from __future__ import annotations

import html
import json
import os
import re
import sys
import urllib.request
import argparse
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "docs" / "data" / "free_time_scheduler_config.json"
SECRETS_PATH = ROOT / "docs" / "data" / "free_time_scheduler_secrets.local.json"
PROPOSED_SESSIONS_PATH = ROOT / "docs" / "data" / "proposed_sessions.json"
PREVIEW_DIR = ROOT / "docs" / "proposed-sessions"
DEBUG_DIR = ROOT / "debug"
REPORT_MD_PATH = DEBUG_DIR / "free_time_scheduler_report.md"
REPORT_JSON_PATH = DEBUG_DIR / "free_time_scheduler_report.json"
RECHECK_MD_PATH = DEBUG_DIR / "free_time_scheduler_recheck.md"
RECHECK_JSON_PATH = DEBUG_DIR / "free_time_scheduler_recheck.json"
OFFERS_PATH = ROOT / "docs" / "data" / "customer_facing_offers.json"
CANONICAL_SCHEDULE_PATH = ROOT / "docs" / "data" / "canonical_schedule_from_class_report.json"
OFFER_REPORT_MD_PATH = DEBUG_DIR / "free_time_offer_strategy_report.md"
OFFER_REPORT_JSON_PATH = DEBUG_DIR / "free_time_offer_strategy_report.json"
APPOINTMENT_LINK_REPORT_JSON_PATH = DEBUG_DIR / "enrollware_appointment_link_report.json"
APPOINTMENT_LINK_REPORT_MD_PATH = DEBUG_DIR / "enrollware_appointment_link_report.md"
REQUESTABLE_APPOINTMENT_URL_REPORT_JSON_PATH = DEBUG_DIR / "requestable_appointment_url_report.json"
REQUESTABLE_APPOINTMENT_URL_REPORT_MD_PATH = DEBUG_DIR / "requestable_appointment_url_report.md"


@dataclass
class CalendarEvent:
    calendar: str
    summary: str
    description: str
    location: str
    start: datetime
    end: datetime
    source_location_name: str
    metadata: dict[str, Any] | None = None


@dataclass
class Location:
    name: str
    address: str
    city: str


@dataclass
class CalendarLoadResult:
    adr_events: list[CalendarEvent]
    brian_adr_blocks: list[CalendarEvent]
    dns_blocks: list[CalendarEvent]
    blocks: list[CalendarEvent]
    warnings: list[str]
    do_not_schedule_source: str
    do_not_schedule_private: bool
    do_not_schedule_loaded: bool
    enrollware_existing_sessions_found: int
    class_report_sessions_found: int
    schedule_future_sessions_found: int
    merged_existing_enrollware_sessions: int
    existing_enrollware_source_mode: str
    duplicate_existing_enrollware_sessions_deduped: int
    existing_enrollware_source_mismatches: list[str]
    hot_sync_active_entries_found: int
    hot_sync_entries_absorbed: int
    hot_sync_entries_kept_active: int
    hot_sync_entries_flagged_for_review: int
    hot_sync_entries_merged: int
    hot_sync_active_path_exists: bool
    class_report_found: bool
    schedule_future_fallback_used: bool
    enrollware_brian_blocks: list[CalendarEvent]
    enrollware_non_brian_ignored: int


def load_config() -> dict[str, Any]:
    with CONFIG_PATH.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def load_local_secrets() -> dict[str, Any]:
    if not SECRETS_PATH.exists():
        return {}
    with SECRETS_PATH.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def resolve_do_not_schedule_ics_url(config: dict[str, Any]) -> tuple[str, str, bool]:
    env_url = os.environ.get("DONOT_SCHEDULE_ICS_URL", "").strip()
    if env_url:
        return env_url, "private ICS URL from DONOT_SCHEDULE_ICS_URL", True

    secrets = load_local_secrets()
    secret_url = str(secrets.get("DONOT_SCHEDULE_ICS_URL") or secrets.get("do_not_schedule_ics_url") or "").strip()
    if secret_url:
        return secret_url, "private ICS URL from local secrets file", True

    return config["calendars"]["do_not_schedule"]["ics_url"], "placeholder ICS URL from public config", False


def normalize_match_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def course_display_name(course_key: str, config: dict[str, Any]) -> str:
    return str(config.get("course_display_names", {}).get(course_key) or "").strip()


def infer_course_key_from_title(title: Any, config: dict[str, Any]) -> str:
    text = normalize_match_text(str(title or ""))
    if not text:
        return ""
    if "bls" in text and ("heartcode" in text or "skills" in text or "olt" in text or "online" in text):
        return "heartcode-bls-skills"
    if "bls" in text and ("renewal" in text or "update" in text):
        return "bls-renewal"
    if "bls" in text and "instructor" not in text:
        return "bls"
    if "heartsaver" in text and "firstaid" in text:
        return "heartsaver-fa-cpr-aed"

    for course in config.get("course_templates", []):
        key = str(course.get("course_key") or "").strip()
        title_key = normalize_match_text(str(course.get("course_title") or ""))
        display_key = normalize_match_text(course_display_name(key, config))
        if key and ((title_key and title_key in text) or (display_key and display_key in text)):
            return key
    return ""


def parse_requested_start(value: Any, config: dict[str, Any]) -> datetime:
    try:
        requested = datetime.fromisoformat(str(value))
    except Exception as exc:
        raise ValueError("requested_start is invalid") from exc
    if requested.tzinfo is None:
        requested = requested.replace(tzinfo=ZoneInfo(config["timezone"]))
    return requested.astimezone(ZoneInfo(config["timezone"]))


def appointment_link_config(config: dict[str, Any]) -> dict[str, Any]:
    return config.get("enrollware_appointment_links", {}) if isinstance(config, dict) else {}


def calculate_appointment_day_id(instructor: str, requested_date: date | str, config: dict[str, Any] | None = None) -> int:
    config = config or load_config()
    bases = config.get("instructor_day_bases", {})
    base = bases.get(str(instructor))
    if not base:
        raise ValueError(f"instructor day base is missing for {instructor}")

    day = date.fromisoformat(str(requested_date)) if not isinstance(requested_date, date) else requested_date
    start_day = date.fromisoformat(str(base.get("appointment_start_date")))
    valid_through = date.fromisoformat(str(base.get("valid_through_date")))
    max_days = int(appointment_link_config(config).get("max_days_from_start", 365) or 365)
    if day < start_day:
        raise ValueError("requested date is before instructor appointment start date")
    if day > valid_through:
        raise ValueError("requested date is after instructor valid_through_date")
    offset = (day - start_day).days
    if offset > max_days:
        raise ValueError("requested date exceeds max_days_from_start")
    base_id = int(base["appointment_start_day_id"])
    return base_id + offset


def format_enrollware_start_time(requested_start: Any, config: dict[str, Any] | None = None) -> str:
    config = config or load_config()
    requested = parse_requested_start(requested_start, config)
    hour = requested.hour % 12 or 12
    return f"{hour}:{requested:%M} {requested:%p}"


def build_enrollware_appointment_url(
    instructor: str,
    course_key: str,
    requested_start: Any,
    config: dict[str, Any] | None = None,
    *,
    explicit_test_mode: bool = False,
) -> str:
    config = config or load_config()
    link_config = appointment_link_config(config)
    if not link_config.get("enabled") and not explicit_test_mode:
        raise ValueError("enrollware appointment links are disabled")
    requested = parse_requested_start(requested_start, config)
    course_ids = link_config.get("course_ids", {})
    if course_key not in course_ids:
        raise ValueError(f"course ID is missing for {course_key}")
    appointment_day_id = calculate_appointment_day_id(instructor, requested.date(), config)
    start_time = format_enrollware_start_time(requested, config)
    base_url = str(link_config.get("base_url") or "https://coastalcprtraining.enrollware.com/enroll").strip()
    encoded_start = start_time.replace(" ", "%20")
    return f"{base_url}?appointmentDayId={appointment_day_id}&startTime={encoded_start}&courseId={int(course_ids[course_key])}"


def instructor_matches_primary(instructor: str | None, config: dict[str, Any]) -> bool:
    instructor_key = normalize_match_text(instructor or "")
    if not instructor_key:
        return False
    terms = config.get("primary_instructor", {}).get("match_terms", [])
    return any(normalize_match_text(term) in instructor_key for term in terms)


def unfold_ics_lines(text: str) -> list[str]:
    lines: list[str] = []
    for raw_line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        if raw_line.startswith((" ", "\t")) and lines:
            lines[-1] += raw_line[1:]
        else:
            lines.append(raw_line)
    return lines


def parse_ics_datetime(value: str, params: dict[str, str], default_tz: ZoneInfo) -> datetime:
    if "T" not in value:
        return datetime.combine(date.fromisoformat(value.replace("-", "")), time.min, tzinfo=default_tz)
    if value.endswith("Z"):
        parsed = datetime.strptime(value, "%Y%m%dT%H%M%SZ").replace(tzinfo=ZoneInfo("UTC"))
    else:
        parsed = datetime.strptime(value, "%Y%m%dT%H%M%S")
        parsed = parsed.replace(tzinfo=ZoneInfo(params.get("TZID", default_tz.key)))
    return parsed.astimezone(default_tz)


def parse_property(line: str) -> tuple[str, dict[str, str], str]:
    head, _, value = line.partition(":")
    parts = head.split(";")
    params: dict[str, str] = {}
    for param in parts[1:]:
        key, _, param_value = param.partition("=")
        params[key.upper()] = param_value.strip('"')
    return parts[0].upper(), params, value


def fetch_ics_events(calendar_name: str, ics_url: str, tz: ZoneInfo) -> tuple[list[dict[str, Any]], str | None]:
    try:
        request = urllib.request.Request(ics_url, headers={"User-Agent": "910CPR-Free-Time-Scheduler/1.0"})
        with urllib.request.urlopen(request, timeout=20) as response:
            text = response.read().decode("utf-8", errors="replace")
    except Exception as exc:
        return [], f"Could not fetch {calendar_name} calendar ICS: {exc}"

    events: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for line in unfold_ics_lines(text):
        if line == "BEGIN:VEVENT":
            current = {"calendar": calendar_name}
        elif line == "END:VEVENT" and current is not None:
            events.append(current)
            current = None
        elif current is not None:
            name, params, value = parse_property(line)
            if name in {"SUMMARY", "DESCRIPTION", "LOCATION", "UID"}:
                current[name.lower()] = value.replace("\\n", "\n").replace("\\,", ",")
            elif name in {"DTSTART", "DTEND"}:
                current[name.lower()] = parse_ics_datetime(value, params, tz)
    return events, None


def normalize_event(raw: dict[str, Any], calendar_name: str, config: dict[str, Any]) -> CalendarEvent | None:
    start = raw.get("dtstart")
    end = raw.get("dtend")
    if not isinstance(start, datetime):
        return None
    if not isinstance(end, datetime):
        end = start + timedelta(hours=1)
    source_location_name = "ADR Station" if calendar_name == "adr" else "Home"
    location = raw.get("location") or source_location_name
    resolved = resolve_location(location, config, is_adr_event=(calendar_name == "adr"))
    return CalendarEvent(
        calendar=calendar_name,
        summary=raw.get("summary", "").strip(),
        description=raw.get("description", "").strip(),
        location=location.strip(),
        start=start,
        end=end,
        source_location_name=resolved.name,
    )


def is_brian_adr_event(event: CalendarEvent) -> bool:
    return "BRIAN" in f"{event.summary}\n{event.description}".upper()


def resolve_location(value: str | None, config: dict[str, Any], is_adr_event: bool = False) -> Location:
    locations = config["locations"]
    fallback_name = "ADR Station" if is_adr_event else "Home"
    text = (value or "").strip()
    if not text:
        text = fallback_name
    for name, info in locations.items():
        if text.lower() == name.lower() or text.lower() == info["address"].lower():
            return Location(name=name, address=info["address"], city=info["city"])
    if "riegelwood" in text.lower() or "john d riegel" in text.lower():
        info = locations["ADR Station"]
        return Location("ADR Station", info["address"], info["city"])
    if "shipyard" in text.lower() or "wilmington office" in text.lower():
        info = locations["Wilmington Office"]
        return Location("Wilmington Office", info["address"], info["city"])
    if "merlot" in text.lower() or "home" == text.lower():
        info = locations["Home"]
        return Location("Home", info["address"], info["city"])
    city_match = re.search(r",\s*([^,]+),\s*(NC|SC|VA)\b", text, flags=re.I)
    city = city_match.group(1).strip() if city_match else "Unknown"
    return Location(name=text, address=text, city=city)


def load_json_sessions(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if isinstance(data, dict):
        sessions = data.get("sessions", [])
        if isinstance(sessions, list):
            return [item for item in sessions if isinstance(item, dict)]
    return []


def load_class_report_sessions(path: Path, config: dict[str, Any]) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    from openpyxl import load_workbook

    tz = ZoneInfo(config["timezone"])
    now = datetime.now(tz)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    header_map = {str(value).strip().lower(): index for index, value in enumerate(headers) if value}
    sessions: list[dict[str, Any]] = []

    def cell(row: tuple[Any, ...], name: str) -> Any:
        index = header_map.get(name.lower())
        return row[index] if index is not None and index < len(row) else None

    for row in rows:
        start = parse_existing_session_time(cell(row, "Start Date / Time"), tz)
        if not start:
            continue
        end = parse_existing_session_time(cell(row, "End Date / Time"), tz)
        if end and end < now:
            continue
        if not end and start < now:
            continue
        registration_url = str(cell(row, "Registration Link") or "").strip()
        class_id_match = re.search(r"[?&]id=(\d+)", registration_url)
        class_id = class_id_match.group(1) if class_id_match else None
        sessions.append(
            {
                "source_name": "class_report_xlsx",
                "class_report_id": str(cell(row, "ID") or "").strip() or None,
                "session_id": class_id or str(cell(row, "ID") or "").strip() or None,
                "class_id": class_id or str(cell(row, "ID") or "").strip() or None,
                "course_id": str(cell(row, "ID") or "").strip() or None,
                "course_number": str(cell(row, "ID") or "").strip() or None,
                "course_name": cell(row, "Course"),
                "official_course_name": cell(row, "Course"),
                "start_at": start.isoformat(),
                "end_at": end.isoformat() if end else None,
                "location_name": cell(row, "Location"),
                "location_display": cell(row, "Location"),
                "lead_instructor_name": cell(row, "Instructor"),
                "session_status": "active",
                "enrolled_count": cell(row, "Students"),
                "registered_count": cell(row, "Students"),
                "max_students": cell(row, "Seats"),
                "available_seats": max(0, int(cell(row, "Seats") or 0) - int(cell(row, "Students") or 0)) if cell(row, "Seats") not in (None, "") and cell(row, "Students") not in (None, "") else None,
                "is_full": bool(cell(row, "Seats") not in (None, "") and cell(row, "Students") not in (None, "") and int(cell(row, "Students") or 0) >= int(cell(row, "Seats") or 0)),
                "registration_url": registration_url,
                "hours": cell(row, "Hours"),
            }
        )
    return sessions


def file_timestamp(path: Path, config: dict[str, Any]) -> datetime | None:
    if not path.exists():
        return None
    return datetime.fromtimestamp(path.stat().st_mtime, tz=ZoneInfo(config["timezone"]))


def first_present(record: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        value = record.get(key)
        if value not in (None, ""):
            return value
    return None


def parse_existing_session_time(value: Any, tz: ZoneInfo) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=tz)
        return value.astimezone(tz)
    if isinstance(value, str):
        parsed = datetime.fromisoformat(value)
        if parsed.tzinfo is None:
            return parsed.replace(tzinfo=tz)
        return parsed.astimezone(tz)
    return None


def fallback_duration_minutes(course_title: str, config: dict[str, Any]) -> int:
    text = course_title.lower()
    for template in config.get("course_templates", []):
        title = str(template.get("course_title", "")).lower()
        key = str(template.get("course_key", "")).lower()
        if key and key in text:
            return int(template.get("duration_minutes", 90))
        if "heartcode" in text and "heartcode" in title:
            return int(template.get("duration_minutes", 45))
        if "renewal" in text and "renewal" in title:
            return int(template.get("duration_minutes", 120))
        if "heartsaver" in text and "heartsaver" in title:
            return int(template.get("duration_minutes", 150))
        if "bls" in text and "provider" in title and "renewal" not in title.lower():
            return int(template.get("duration_minutes", 180))
    return 90


def session_dedupe_key(session: dict[str, Any], config: dict[str, Any]) -> str:
    session_id = str(first_present(session, ["class_id", "session_id"]) or "").strip()
    if session_id:
        return f"id:{session_id}"
    tz = ZoneInfo(config["timezone"])
    start = parse_existing_session_time(first_present(session, ["start_at", "start", "start_time"]), tz)
    course = normalize_match_text(str(first_present(session, ["official_course_name", "course_name", "title", "course"]) or ""))
    location = normalize_match_text(str(first_present(session, ["location_name", "location_display", "location"]) or ""))
    instructor = normalize_match_text(str(first_present(session, ["lead_instructor_name", "instructor", "instructor_name"]) or ""))
    return f"compound:{course}:{start.isoformat() if start else ''}:{location}:{instructor}"


def session_match_keys(session: dict[str, Any], config: dict[str, Any]) -> set[str]:
    keys = {session_dedupe_key(session, config)}
    session_id = str(first_present(session, ["class_id", "session_id", "enrollware_class_id"]) or "").strip()
    if session_id:
        keys.add(f"id:{session_id}")
    tz = ZoneInfo(config["timezone"])
    start = parse_existing_session_time(first_present(session, ["start_at", "start", "start_time"]), tz)
    course = normalize_match_text(str(first_present(session, ["official_course_name", "course_name", "course_title", "title", "course"]) or ""))
    location = normalize_match_text(str(first_present(session, ["location_name", "location_display", "location"]) or ""))
    instructor = normalize_match_text(str(first_present(session, ["lead_instructor_name", "instructor", "instructor_name"]) or ""))
    if course and start and location and instructor:
        keys.add(f"compound:{course}:{start.isoformat()}:{location}:{instructor}")
    return keys


def load_hot_sync_entries(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if isinstance(data, dict):
        for key in ("sessions", "entries", "active"):
            value = data.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
    return []


def normalize_hot_sync_entry(entry: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(entry)
    if "class_id" not in normalized:
        normalized["class_id"] = first_present(entry, ["enrollware_class_id", "session_id", "id"])
    if "session_id" not in normalized:
        normalized["session_id"] = first_present(entry, ["class_id", "enrollware_class_id", "id"])
    if "course_name" not in normalized:
        normalized["course_name"] = first_present(entry, ["course_title", "title", "course"])
    if "start_at" not in normalized:
        normalized["start_at"] = first_present(entry, ["start_time", "start"])
    if "end_at" not in normalized:
        normalized["end_at"] = first_present(entry, ["end_time", "end"])
    if "lead_instructor_name" not in normalized:
        normalized["lead_instructor_name"] = first_present(entry, ["instructor", "instructor_name"])
    normalized["source_name"] = "hot_sync_delta"
    return normalized


def entry_created_at(entry: dict[str, Any], config: dict[str, Any]) -> datetime | None:
    tz = ZoneInfo(config["timezone"])
    return parse_existing_session_time(first_present(entry, ["created_at", "changed_at", "updated_at", "synced_at"]), tz)


def absorb_hot_sync_delta(
    class_report_sessions: list[dict[str, Any]],
    class_report_timestamp: datetime | None,
    config: dict[str, Any],
) -> dict[str, Any]:
    hot_config = config.get("hot_sync_delta", {})
    if not hot_config.get("enabled", False):
        return {
            "active_path_exists": False,
            "active_entries_found": 0,
            "absorbed_entries": [],
            "kept_active_entries": [],
            "review_entries": [],
            "active_sessions_for_merge": [],
        }

    active_path = (ROOT / str(hot_config.get("active_path", "data/runtime/free_time_scheduler/hot_sync_active.json"))).resolve()
    archive_path = (ROOT / str(hot_config.get("absorbed_archive_path", "data/runtime/free_time_scheduler/hot_sync_absorbed.jsonl"))).resolve()
    active_path_exists = active_path.exists()
    raw_entries = load_hot_sync_entries(active_path)
    class_report_keys: set[str] = set()
    for session in class_report_sessions:
        class_report_keys.update(session_match_keys(session, config))

    absorbed: list[dict[str, Any]] = []
    kept_active: list[dict[str, Any]] = []
    review: list[dict[str, Any]] = []
    for raw_entry in raw_entries:
        entry = normalize_hot_sync_entry(raw_entry)
        entry_keys = session_match_keys(entry, config)
        if entry_keys & class_report_keys:
            absorbed_entry = {
                **entry,
                "hot_sync_status": "absorbed",
                "absorbed_at": datetime.now(ZoneInfo(config["timezone"])).isoformat(),
                "absorbed_reason": "Matched current Class Report by ID or course/date/time/location/instructor.",
            }
            absorbed.append(absorbed_entry)
            continue

        created_at = entry_created_at(entry, config)
        if class_report_timestamp and created_at and created_at <= class_report_timestamp:
            entry["hot_sync_status"] = "review_required"
            entry["review_reason"] = "Unmatched hot-sync entry is older than or equal to the Class Report timestamp."
            review.append(entry)
        else:
            entry["hot_sync_status"] = "active"
            kept_active.append(entry)

    if active_path_exists:
        active_path.parent.mkdir(parents=True, exist_ok=True)
        active_path.write_text(json.dumps([*kept_active, *review], indent=2), encoding="utf-8")
    if absorbed:
        archive_path.parent.mkdir(parents=True, exist_ok=True)
        with archive_path.open("a", encoding="utf-8") as handle:
            for entry in absorbed:
                handle.write(json.dumps(entry, ensure_ascii=True) + "\n")

    return {
        "active_path_exists": active_path_exists,
        "active_entries_found": len(raw_entries),
        "absorbed_entries": absorbed,
        "kept_active_entries": kept_active,
        "review_entries": review,
        "active_sessions_for_merge": [*kept_active, *review],
    }


def merge_existing_enrollware_sessions(
    class_report_sessions: list[dict[str, Any]],
    schedule_future_sessions: list[dict[str, Any]],
    hot_sync_sessions: list[dict[str, Any]],
    config: dict[str, Any],
) -> tuple[list[dict[str, Any]], int, list[str]]:
    prefer_class_report = bool(config.get("enrollware_class_report", {}).get("prefer_over_schedule_future", True))
    merged: dict[str, dict[str, Any]] = {}
    duplicates = 0
    mismatches: list[str] = []

    base_sources = [schedule_future_sessions, class_report_sessions] if prefer_class_report else [class_report_sessions, schedule_future_sessions]
    ordered_sources = [*base_sources, hot_sync_sessions]
    for source_sessions in ordered_sources:
        for session in source_sessions:
            key = session_dedupe_key(session, config)
            if key in merged:
                duplicates += 1
                existing_title = str(first_present(merged[key], ["official_course_name", "course_name", "title", "course"]) or "")
                new_title = str(first_present(session, ["official_course_name", "course_name", "title", "course"]) or "")
                if existing_title and new_title and normalize_match_text(existing_title) != normalize_match_text(new_title):
                    mismatches.append(f"{key}: course title differs between sources")
            merged[key] = session
    return list(merged.values()), duplicates, mismatches[:50]


def load_existing_enrollware_blocks(config: dict[str, Any]) -> dict[str, Any]:
    tz = ZoneInfo(config["timezone"])
    schedule_relative_path = config.get("existing_enrollware_sessions", {}).get("source_path", "docs/data/schedule_future.json")
    schedule_path = (ROOT / schedule_relative_path).resolve()
    schedule_sessions = load_json_sessions(schedule_path)

    class_report_config = config.get("enrollware_class_report", {})
    class_report_enabled = bool(class_report_config.get("enabled", False))
    prefer_class_report = bool(class_report_config.get("prefer_over_schedule_future", True))
    class_report_path = (ROOT / str(class_report_config.get("path", "data/private/Class report.xlsx"))).resolve()
    class_report_found = class_report_enabled and class_report_path.exists()
    class_report_sessions = load_class_report_sessions(class_report_path, config) if class_report_found else []
    class_report_timestamp = file_timestamp(class_report_path, config) if class_report_found else None
    hot_sync_result = absorb_hot_sync_delta(class_report_sessions, class_report_timestamp, config)
    class_report_authoritative = class_report_found and prefer_class_report
    if class_report_authoritative:
        merged_sessions, duplicates, mismatches = merge_existing_enrollware_sessions(
            class_report_sessions,
            [],
            hot_sync_result["active_sessions_for_merge"],
            config,
        )
        schedule_keys = set()
        for session in schedule_sessions:
            schedule_keys.update(session_match_keys(session, config))
        class_report_keys = set()
        for session in class_report_sessions:
            class_report_keys.update(session_match_keys(session, config))
        stale_schedule_only = [
            str(first_present(session, ["session_id", "class_id"]) or "")
            for session in schedule_sessions
            if not (session_match_keys(session, config) & class_report_keys)
        ]
        mismatches.extend([f"schedule_future session not in authoritative Class Report: {item}" for item in stale_schedule_only[:50] if item])
        source_mode = "class_report_authoritative"
    else:
        merged_sessions, duplicates, mismatches = merge_existing_enrollware_sessions(
            class_report_sessions,
            schedule_sessions,
            hot_sync_result["active_sessions_for_merge"],
            config,
        )
        source_mode = "schedule_future_fallback" if not class_report_found else "merged_comparison"

    brian_blocks: list[CalendarEvent] = []
    ignored_non_brian = 0

    for session in merged_sessions:
        instructor = str(first_present(session, ["lead_instructor_name", "instructor", "instructor_name"]) or "").strip()
        if not instructor_matches_primary(instructor, config):
            ignored_non_brian += 1
            continue

        start = parse_existing_session_time(first_present(session, ["start_at", "start", "start_time"]), tz)
        end = parse_existing_session_time(first_present(session, ["end_at", "end", "end_time"]), tz)
        if not start:
            continue
        if not end:
            hours = first_present(session, ["hours"])
            if isinstance(hours, (int, float)) and hours > 0:
                end = start + timedelta(minutes=int(float(hours) * 60))
            else:
                title_for_duration = str(first_present(session, ["official_course_name", "course_name", "title", "course"]) or "")
                end = start + timedelta(minutes=fallback_duration_minutes(title_for_duration, config))

        location_name = str(first_present(session, ["location_name", "location_display", "location"]) or "").strip()
        resolved = resolve_location(location_name, config)
        title = str(first_present(session, ["official_course_name", "course_name", "title", "course"]) or "Existing Enrollware session").strip()
        normalized_course_key = infer_course_key_from_title(title, config)
        metadata = {
            "source": "enrollware_existing_session",
            "course_title": title,
            "course_key": normalized_course_key,
            "session_id": first_present(session, ["session_id"]),
            "enrollware_class_id": first_present(session, ["class_id", "session_id"]),
            "instructor": instructor,
            "start_time": start.isoformat(),
            "end_time": end.isoformat(),
            "location_name": location_name or resolved.name,
            "location_address": resolved.address,
            "city": resolved.city,
            "state": first_present(session, ["state"]) or "NC",
            "enrolled_count": first_present(session, ["enrolled_count", "registered_count"]),
            "capacity": first_present(session, ["max_students", "capacity"]),
        }
        brian_blocks.append(
            CalendarEvent(
                calendar="enrollware_existing_session",
                summary=title,
                description="Existing Enrollware session assigned to Brian.",
                location=location_name or resolved.name,
                start=start,
                end=end,
                source_location_name=resolved.name,
                metadata=metadata,
            )
        )
    return {
        "existing_enrollware_sessions_found": len(merged_sessions),
        "class_report_sessions_found": len(class_report_sessions),
        "schedule_future_sessions_found": len(schedule_sessions),
        "merged_existing_enrollware_sessions": len(merged_sessions),
        "existing_enrollware_source_mode": source_mode,
        "duplicate_existing_enrollware_sessions_deduped": duplicates,
        "existing_enrollware_source_mismatches": mismatches,
        "hot_sync_active_entries_found": hot_sync_result["active_entries_found"],
        "hot_sync_entries_absorbed": len(hot_sync_result["absorbed_entries"]),
        "hot_sync_entries_kept_active": len(hot_sync_result["kept_active_entries"]),
        "hot_sync_entries_flagged_for_review": len(hot_sync_result["review_entries"]),
        "hot_sync_entries_merged": len(hot_sync_result["active_sessions_for_merge"]),
        "hot_sync_active_path_exists": hot_sync_result["active_path_exists"],
        "class_report_found": class_report_found,
        "schedule_future_fallback_used": not class_report_authoritative,
        "brian_blocks": brian_blocks,
        "non_brian_ignored": ignored_non_brian,
    }


def estimate_drive_minutes(from_location: Location, to_location: Location) -> int:
    if from_location.address == to_location.address or from_location.name == to_location.name:
        return 0
    pair = (from_location.name, to_location.name)
    known = {
        ("Wilmington Office", "Home"): 20,
        ("Home", "Wilmington Office"): 20,
        ("ADR Station", "Wilmington Office"): 60,
        ("Wilmington Office", "ADR Station"): 60,
        ("ADR Station", "Home"): 60,
        ("Home", "ADR Station"): 60,
    }
    if pair in known:
        return known[pair]
    if from_location.city != "Unknown" and from_location.city == to_location.city:
        return 20
    return 75


def required_buffer_minutes(drive_minutes: int, same_city: bool, relation: str, adjacent_calendar: str | None) -> int:
    if adjacent_calendar == "adr" and relation == "after_previous":
        return drive_minutes + 45
    if adjacent_calendar == "adr" and relation == "before_next":
        return drive_minutes + 90
    return drive_minutes + (15 if same_city else 30)


def candidate_dates(start_date: date, days: int) -> list[date]:
    return [start_date + timedelta(days=offset) for offset in range(days)]


def candidate_times_for_day(day: date, config: dict[str, Any]) -> list[str]:
    if day.weekday() == 6:
        return config["candidate_start_times"]["sunday"]
    if day.weekday() == 5:
        return config["candidate_start_times"]["saturday"]
    return config["candidate_start_times"]["weekday"]


def find_previous_block(blocks: list[CalendarEvent], start: datetime) -> CalendarEvent | None:
    previous = [block for block in blocks if block.end <= start]
    return max(previous, key=lambda block: block.end) if previous else None


def find_next_block(blocks: list[CalendarEvent], end: datetime) -> CalendarEvent | None:
    next_blocks = [block for block in blocks if block.start >= end]
    return min(next_blocks, key=lambda block: block.start) if next_blocks else None


def overlapping_blocks(blocks: list[CalendarEvent], start: datetime, end: datetime) -> list[CalendarEvent]:
    return [block for block in blocks if block.start < end and block.end > start]


def slugify_session(course_key: str, location_name: str, start: datetime) -> str:
    location_part = re.sub(r"[^a-z0-9]+", "-", location_name.lower()).strip("-").replace("-office", "")
    return f"{course_key}-{location_part}-{start:%Y-%m-%d-%H%M}"


def slugify_offer(course_key: str, location_name: str, start: datetime) -> str:
    location_part = re.sub(r"[^a-z0-9]+", "-", location_name.lower()).strip("-").replace("-office", "")
    return f"{course_key}-{location_part}-{start:%Y%m%d}-{start:%H%M}"


def evaluate_candidate_session(
    course: dict[str, Any],
    start: datetime,
    blocks: list[CalendarEvent],
    config: dict[str, Any],
) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    class_location = resolve_location(course["location"], config)
    end = start + timedelta(minutes=course["duration_minutes"])
    cleanup_end = end + timedelta(minutes=course["cleanup_minutes"])
    conflicts = overlapping_blocks(blocks, start, cleanup_end)
    base = {
        "course_key": course["course_key"],
        "course_title": course["course_title"],
        "start_time": start.isoformat(),
        "end_time": end.isoformat(),
        "cleanup_minutes": course["cleanup_minutes"],
        "location_name": class_location.name,
    }
    if conflicts:
        enrollware_conflict = next((block for block in conflicts if block.calendar == "enrollware_existing_session"), None)
        if enrollware_conflict:
            when = enrollware_conflict.start.strftime("%Y-%m-%d %-I:%M %p") if sys.platform != "win32" else enrollware_conflict.start.strftime("%Y-%m-%d %#I:%M %p")
            reason = (
                "Rejected because proposed session overlaps existing Brian Enrollware session: "
                f"{enrollware_conflict.summary} at {enrollware_conflict.location}, {when}."
            )
            return None, {**base, "reason": reason, "rejection_source": "enrollware_existing_session"}
        names = ", ".join(f"{block.calendar}: {block.summary or '(untitled)'}" for block in conflicts[:3])
        return None, {**base, "reason": f"Rejected because it overlaps blocking event(s): {names}."}

    previous = find_previous_block(blocks, start)
    if previous:
        previous_location = resolve_location(previous.location, config, is_adr_event=(previous.calendar == "adr"))
        drive = estimate_drive_minutes(previous_location, class_location)
        same_city = previous_location.city == class_location.city
        required = required_buffer_minutes(drive, same_city, "after_previous", previous.calendar)
        actual = int((start - previous.end).total_seconds() // 60)
        if actual < required:
            return None, {
                **base,
                "reason": (
                    f"Rejected because only {actual} minutes are available after previous block "
                    f"'{previous.summary or '(untitled)'}'; {required} minutes are required."
                ),
            }

    next_block = find_next_block(blocks, cleanup_end)
    if next_block:
        next_location = resolve_location(next_block.location, config, is_adr_event=(next_block.calendar == "adr"))
        drive = estimate_drive_minutes(class_location, next_location)
        same_city = class_location.city == next_location.city
        required = required_buffer_minutes(drive, same_city, "before_next", next_block.calendar)
        actual = int((next_block.start - cleanup_end).total_seconds() // 60)
        if actual < required:
            return None, {
                **base,
                "reason": (
                    f"Rejected because only {actual} minutes are available before next block "
                    f"'{next_block.summary or '(untitled)'}'; {required} minutes are required."
                ),
            }

    slug = slugify_session(course["course_key"], class_location.name, start)
    return {
        "session_status": "proposed",
        "course_key": course["course_key"],
        "course_title": course["course_title"],
        "start_time": start.isoformat(),
        "end_time": end.isoformat(),
        "cleanup_minutes": course["cleanup_minutes"],
        "location_name": class_location.name,
        "location_address": class_location.address,
        "capacity": course["capacity"],
        "enrollware_class_id": course.get("enrollware_class_id"),
        "enrollware_enroll_url": course.get("enrollware_enroll_url"),
        "page_slug": slug,
        "availability_explanation": "Accepted because no ADR or DoNotSchedule conflict was found and travel buffers were satisfied.",
    }, None


def generate_candidates(config: dict[str, Any], blocks: list[CalendarEvent]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    tz = ZoneInfo(config["timezone"])
    accepted: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    today = datetime.now(tz).date()
    for day in candidate_dates(today, int(config["planning_days"])):
        for start_text in candidate_times_for_day(day, config):
            hour, minute = [int(part) for part in start_text.split(":", 1)]
            start = datetime.combine(day, time(hour, minute), tzinfo=tz)
            for course in config["course_templates"]:
                if course["location"] != "Wilmington Office":
                    rejected.append({
                        "course_key": course["course_key"],
                        "course_title": course["course_title"],
                        "start_time": start.isoformat(),
                        "reason": "Rejected because Phase 1 only proposes Wilmington Office sessions.",
                    })
                    continue
                session, rejection = evaluate_candidate_session(course, start, blocks, config)
                if session:
                    accepted.append(session)
                elif rejection:
                    rejected.append(rejection)
    return accepted, rejected


def course_program_key(course: dict[str, Any]) -> str:
    return str(course.get("program_key") or course.get("course_key") or "").strip()


def block_program_key(block: CalendarEvent, config: dict[str, Any]) -> str:
    explicit = str((block.metadata or {}).get("course_key") or "").strip()
    if explicit:
        return explicit
    text = normalize_match_text(f"{block.summary} {(block.metadata or {}).get('course_title', '')}")
    inferred = infer_course_key_from_title(text, config)
    if inferred:
        return inferred
    for course in config.get("course_templates", []):
        key = course_program_key(course)
        title = normalize_match_text(str(course.get("course_title") or ""))
        if key and normalize_match_text(key) in text:
            return key
        if title and title in text:
            return key
        if key == "bls-renewal" and "renewal" in text and "bls" in text:
            return key
        if key == "heartcode-bls-skills" and "heartcode" in text and "bls" in text:
            return key
        if key == "heartsaver-fa-cpr-aed" and "heartsaver" in text:
            return key
        if key == "bls" and "bls" in text and "renewal" not in text and "heartcode" not in text:
            return key
    return ""


def blocks_on_day(blocks: list[CalendarEvent], day: date, tz: ZoneInfo) -> list[CalendarEvent]:
    start = datetime.combine(day, time.min, tzinfo=tz)
    end = start + timedelta(days=1)
    return [block for block in blocks if block.start < end and block.end > start]


def same_program_gap_status(
    course: dict[str, Any],
    start: datetime,
    existing_blocks: list[CalendarEvent],
    accepted_offers: list[dict[str, Any]],
    config: dict[str, Any],
    frequency_group: dict[str, Any],
) -> tuple[bool, dict[str, Any] | None]:
    gap_hours = float(frequency_group.get("same_program_gap_hours", 24))
    program_key = course_program_key(course)
    if not program_key:
        return True, None

    threshold = timedelta(hours=gap_hours)
    for block in existing_blocks:
        if block_program_key(block, config) != program_key:
            continue
        if abs(start - block.start) < threshold:
            metadata = block.metadata or {}
            source = "real canonical session" if block.calendar == "enrollware_existing_session" else block.calendar
            return False, {
                "suppression_reason": f"suppressed because same-program gap against {source}",
                "candidate_course_key": program_key,
                "candidate_display_name": course_display_name(program_key, config) or str(course.get("course_title") or ""),
                "candidate_start": start.isoformat(),
                "conflicting_canonical_class_id": metadata.get("enrollware_class_id") or metadata.get("session_id"),
                "conflicting_canonical_title": metadata.get("course_title") or block.summary,
                "conflicting_start": block.start.isoformat(),
                "gap_hours_used": gap_hours,
                "conflict_source": block.calendar,
            }

    for offer in accepted_offers:
        if offer.get("course_key") != program_key:
            continue
        offer_start = datetime.fromisoformat(offer["start_time"])
        if abs(start - offer_start) < threshold:
            return False, {
                "suppression_reason": f"suppressed because same-program offer already exists within {gap_hours:g} hours",
                "candidate_course_key": program_key,
                "candidate_display_name": course_display_name(program_key, config) or str(course.get("course_title") or ""),
                "candidate_start": start.isoformat(),
                "conflicting_start": offer_start.isoformat(),
                "gap_hours_used": gap_hours,
                "conflict_source": "selected_customer_facing_offer",
            }
    return True, None


def resolve_offer_frequency_group(course: dict[str, Any], config: dict[str, Any]) -> tuple[str, dict[str, Any], str]:
    strategy = config.get("offer_strategy", {})
    groups = strategy.get("offer_frequency_groups", {})
    mappings = strategy.get("course_offer_overrides", {})
    course_key = str(course.get("course_key") or "").strip()
    if course_key in mappings:
        group_name = str(mappings[course_key])
        return group_name, groups.get(group_name, {}), "explicit course_offer_overrides entry"
    default_name = str(strategy.get("default_offer_frequency_group") or "core_10_hour")
    return default_name, groups.get(default_name, {}), "default offer frequency group"


def has_same_program_claim_on_day(course: dict[str, Any], day_key: str, blocks: list[CalendarEvent], config: dict[str, Any]) -> bool:
    program_key = course_program_key(course)
    if not program_key:
        return False
    for block in blocks:
        if block.start.date().isoformat() != day_key:
            continue
        if block_program_key(block, config) == program_key:
            return True
    return False


def same_program_offer_on_day(course: dict[str, Any], day_key: str, offers: list[dict[str, Any]]) -> bool:
    program_key = course_program_key(course)
    return any(offer.get("course_key") == program_key and offer.get("date") == day_key for offer in offers)


def week_key(value: datetime) -> str:
    year, week, _weekday = value.isocalendar()
    return f"{year}-W{week:02d}"


def notice_minutes_for_group(strategy: dict[str, Any], group_name: str) -> tuple[int, str]:
    by_group = strategy.get("minimum_customer_notice_minutes_by_group", {})
    if group_name in by_group:
        return int(by_group[group_name]), "group"
    return int(strategy.get("minimum_customer_notice_minutes", 120)), "global"


def offer_type_priority(offer_type: str) -> int:
    return {
        "stack_after_existing_session": 0,
        "stack_before_next_block": 1,
        "stack_before_existing_session": 1,
        "clean_slate_anchor": 2,
        "later_seed_anchor": 3,
        "next_day_anchor": 4,
    }.get(offer_type, 99)


def dedupe_offer_seeds(seeds: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, str]] = set()
    out: list[dict[str, Any]] = []
    for seed in sorted(seeds, key=lambda item: (item["start"], offer_type_priority(item["offer_type"]))):
        key = (seed["start"].isoformat(), seed.get("offer_type", ""))
        plain_key = (seed["start"].isoformat(), "")
        if plain_key in seen:
            continue
        seen.add(key)
        seen.add(plain_key)
        out.append(seed)
    return out


def build_offer_seeds(day: date, blocks: list[CalendarEvent], course: dict[str, Any], config: dict[str, Any]) -> list[dict[str, Any]]:
    tz = ZoneInfo(config["timezone"])
    strategy = config.get("offer_strategy", {})
    day_blocks = blocks_on_day(blocks, day, tz)
    clean_slate = not day_blocks
    anchors = strategy.get("clean_slate_anchor_times", [])
    seeds: list[dict[str, Any]] = []

    for anchor in anchors:
        hour, minute = [int(part) for part in anchor.split(":", 1)]
        anchor_day = day + timedelta(days=1) if anchor == "00:00" and strategy.get("allow_overnight_anchors", True) else day
        offer_type = "clean_slate_anchor" if clean_slate else "later_seed_anchor"
        seeds.append({"start": datetime.combine(anchor_day, time(hour, minute), tzinfo=tz), "offer_type": offer_type, "context": None})

    for block in day_blocks:
        context = {
            "source": block.calendar,
            "summary": block.summary,
            "start_time": block.start.isoformat(),
            "end_time": block.end.isoformat(),
        }
        seeds.append({
            "start": block.end + timedelta(minutes=int(course.get("cleanup_minutes", 30))),
            "offer_type": "stack_after_existing_session",
            "context": context,
        })
        duration = timedelta(minutes=int(course["duration_minutes"]) + int(course.get("cleanup_minutes", 30)))
        seeds.append({
            "start": block.start - duration,
            "offer_type": "stack_before_existing_session",
            "context": context,
        })
    return dedupe_offer_seeds(seeds)


def horizon_spread_config(strategy: dict[str, Any], group_name: str) -> dict[str, Any]:
    spread = strategy.get("horizon_spread", {})
    if not isinstance(spread, dict) or not spread.get("enabled"):
        return {"enabled": False, "buckets": [], "fill_unused_slots_from_earliest_valid": True}
    by_group = spread.get("by_group", {})
    group_config = by_group.get(group_name, {}) if isinstance(by_group, dict) else {}
    default_config = spread.get("default", {}) if isinstance(spread.get("default", {}), dict) else {}
    config = {**default_config, **group_config}
    buckets = [bucket for bucket in config.get("buckets", []) if isinstance(bucket, dict)]
    return {
        "enabled": True,
        "buckets": buckets,
        "fill_unused_slots_from_earliest_valid": bool(config.get("fill_unused_slots_from_earliest_valid", True)),
    }


def max_horizon_end_offset(strategy: dict[str, Any]) -> int:
    spread = strategy.get("horizon_spread", {})
    if not isinstance(spread, dict) or not spread.get("enabled"):
        return 0
    values: list[int] = []
    configs = []
    if isinstance(spread.get("default"), dict):
        configs.append(spread["default"])
    if isinstance(spread.get("by_group"), dict):
        configs.extend(config for config in spread["by_group"].values() if isinstance(config, dict))
    for config in configs:
        for bucket in config.get("buckets", []):
            if isinstance(bucket, dict):
                try:
                    values.append(int(bucket.get("end_day_offset", 0)))
                except Exception:
                    pass
    return max(values or [0])


def bucket_for_start(start: datetime, start_day: date, spread_config: dict[str, Any]) -> dict[str, Any] | None:
    if not spread_config.get("enabled"):
        return None
    offset = (start.date() - start_day).days
    for bucket in spread_config.get("buckets", []):
        try:
            start_offset = int(bucket.get("start_day_offset", 0))
            end_offset = int(bucket.get("end_day_offset", start_offset))
        except Exception:
            continue
        if start_offset <= offset <= end_offset:
            return {
                "key": str(bucket.get("key") or f"days_{start_offset}_to_{end_offset}"),
                "start_day_offset": start_offset,
                "end_day_offset": end_offset,
                "target_count": int(bucket.get("target_count", 0)),
            }
    return None


def offer_selection_key(item: dict[str, Any]) -> tuple[int, str]:
    return (offer_type_priority(item.get("offer_type", "")), item.get("start_time", ""))


def dedupe_offer_candidates(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, str, str]] = set()
    out: list[dict[str, Any]] = []
    for item in sorted(candidates, key=offer_selection_key):
        key = (
            str(item.get("course_key") or ""),
            str(item.get("start_time") or ""),
            str(item.get("location_name") or ""),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def build_offer_record(
    session: dict[str, Any],
    offer_type: str,
    context: dict[str, Any] | None,
    same_program_checked: bool,
    same_program_passed: bool,
    notice_minutes_used: int,
    notice_window_source: str,
) -> dict[str, Any]:
    start = datetime.fromisoformat(session["start_time"])
    end = datetime.fromisoformat(session["end_time"])
    display_name = course_display_name(session["course_key"], load_config()) or session["course_title"]
    label = {
        "clean_slate_anchor": "Available by request",
        "stack_after_existing_session": "Check This Time",
        "stack_before_existing_session": "Check This Time",
        "later_seed_anchor": "Request This Time",
        "next_day_anchor": "Request This Time",
    }.get(offer_type, "Request This Time")
    return {
        "course_title_internal": session["course_title"],
        "course_display_name": display_name,
        "offer_slug": slugify_offer(session["course_key"], session["location_name"], start),
        "date": start.date().isoformat(),
        "start_time": session["start_time"],
        "end_time": session["end_time"],
        "location_name": session["location_name"],
        "location_address": session["location_address"],
        "offer_type": offer_type,
        "customer_label": label,
        "source_session_context": context,
        "explanation": f"{label}. This time passed current availability screening and requires click-time recheck before any booking action.",
        "same_program_gap_checked": same_program_checked,
        "same_program_gap_passed": same_program_passed,
        "minimum_customer_notice_minutes_used": notice_minutes_used,
        "minimum_customer_notice_window_source": notice_window_source,
        "click_time_recheck_required_before_enrollware_action": True,
    }


def select_customer_facing_offers(
    candidates: list[dict[str, Any]],
    *,
    course: dict[str, Any],
    course_limit: int,
    start_day: date,
    spread_config: dict[str, Any],
    calendar_result: CalendarLoadResult,
    config: dict[str, Any],
    frequency_group: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    candidates = dedupe_offer_candidates(candidates)
    for item in candidates:
        start = datetime.fromisoformat(item["start_time"])
        bucket = bucket_for_start(start, start_day, spread_config)
        item["horizon_bucket_key"] = bucket["key"] if bucket else "unbucketed"
        item["horizon_bucket_start_day_offset"] = bucket["start_day_offset"] if bucket else None
        item["horizon_bucket_end_day_offset"] = bucket["end_day_offset"] if bucket else None
        item["horizon_spread_selected"] = False

    selected: list[dict[str, Any]] = []
    suppressed: list[dict[str, Any]] = []
    per_day_counts: dict[str, int] = {}
    per_week_counts: dict[str, int] = {}
    selected_keys: set[tuple[str, str, str]] = set()
    max_per_day = int(config.get("offer_strategy", {}).get("max_options_per_day_per_course_page", 3))
    max_week = frequency_group.get("max_offers_per_week")
    fill_from_earliest_count = 0

    def key_for(item: dict[str, Any]) -> tuple[str, str, str]:
        return (
            str(item.get("course_key") or ""),
            str(item.get("start_time") or ""),
            str(item.get("location_name") or ""),
        )

    def try_select(item: dict[str, Any], *, fill_from_earliest: bool = False) -> bool:
        if len(selected) >= course_limit:
            return False
        key = key_for(item)
        if key in selected_keys:
            return False
        start = datetime.fromisoformat(item["start_time"])
        day_key = start.date().isoformat()
        wk = week_key(start)
        if max_week is not None and per_week_counts.get(wk, 0) >= int(max_week):
            return False
        if per_day_counts.get(day_key, 0) >= max_per_day:
            return False
        same_program_claimed_today = has_same_program_claim_on_day(course, day_key, calendar_result.blocks, config)
        if same_program_claimed_today and not bool(frequency_group.get("allow_same_day_reoffer", True)):
            return False
        if same_program_offer_on_day(course, day_key, selected) and not bool(frequency_group.get("allow_same_day_reoffer", True)):
            return False
        max_after_claim = frequency_group.get("max_offers_per_day_after_claim")
        if same_program_claimed_today and max_after_claim is not None and per_day_counts.get(day_key, 0) >= int(max_after_claim):
            return False
        gap_passed, _gap_reason = same_program_gap_status(course, start, calendar_result.blocks, selected, config, frequency_group)
        if not gap_passed:
            return False
        selected_item = dict(item)
        selected_item["horizon_spread_selected"] = True
        selected_item["horizon_spread_fill_from_earliest"] = fill_from_earliest
        selected.append(selected_item)
        selected_keys.add(key)
        per_day_counts[day_key] = per_day_counts.get(day_key, 0) + 1
        per_week_counts[wk] = per_week_counts.get(wk, 0) + 1
        return True

    bucket_debug: dict[str, dict[str, Any]] = {}
    if spread_config.get("enabled"):
        for bucket in spread_config.get("buckets", []):
            key = str(bucket.get("key") or "")
            if not key:
                continue
            bucket_candidates = [item for item in candidates if item.get("horizon_bucket_key") == key]
            target_count = int(bucket.get("target_count", 0))
            before = len(selected)
            for item in sorted(bucket_candidates, key=offer_selection_key):
                if len(selected) - before >= target_count:
                    break
                try_select(item)
            selected_count = len(selected) - before
            bucket_debug[key] = {
                "valid_candidates": len(bucket_candidates),
                "target_count": target_count,
                "selected_count": selected_count,
                "unused_bucket_slots": max(0, target_count - selected_count),
                "start_day_offset": int(bucket.get("start_day_offset", 0)),
                "end_day_offset": int(bucket.get("end_day_offset", 0)),
            }

        if spread_config.get("fill_unused_slots_from_earliest_valid", True) and len(selected) < course_limit:
            before = len(selected)
            for item in sorted(candidates, key=lambda candidate: (candidate.get("start_time", ""), offer_type_priority(candidate.get("offer_type", "")))):
                if len(selected) >= course_limit:
                    break
                try_select(item, fill_from_earliest=True)
            fill_from_earliest_count = len(selected) - before
    else:
        for item in sorted(candidates, key=lambda candidate: (candidate.get("start_time", ""), offer_type_priority(candidate.get("offer_type", "")))):
            if len(selected) >= course_limit:
                break
            try_select(item)

    for item in candidates:
        if key_for(item) not in selected_keys:
            suppressed.append({
                "course_key": item.get("course_key"),
                "date_time": item.get("start_time"),
                "suppression_reason": "not selected because course offer cap was filled after horizon spread",
                "offer_frequency_group": item.get("offer_frequency_group"),
                "horizon_bucket_key": item.get("horizon_bucket_key"),
            })

    month_counts: dict[str, int] = {}
    bucket_counts: dict[str, int] = {}
    for item in selected:
        start = datetime.fromisoformat(item["start_time"])
        month_key = start.strftime("%Y-%m")
        month_counts[month_key] = month_counts.get(month_key, 0) + 1
        bucket_key = str(item.get("horizon_bucket_key") or "unbucketed")
        bucket_counts[bucket_key] = bucket_counts.get(bucket_key, 0) + 1

    selection_debug = {
        "max_offers_allowed": course_limit,
        "horizon_spread_enabled": bool(spread_config.get("enabled")),
        "valid_candidates": len(candidates),
        "selected_offers_by_horizon_bucket": bucket_counts,
        "valid_candidates_by_horizon_bucket": {
            key: value["valid_candidates"] for key, value in bucket_debug.items()
        },
        "horizon_buckets": bucket_debug,
        "unused_bucket_slots": {
            key: value["unused_bucket_slots"] for key, value in bucket_debug.items()
        },
        "fill_from_earliest_count": fill_from_earliest_count,
        "final_selected_offers_by_month": month_counts,
    }
    return selected, suppressed, selection_debug


def generate_customer_facing_offers(course_key_filter: str = "", date_filter: str = "", limit: int | None = None) -> int:
    config = load_config()
    tz = ZoneInfo(config["timezone"])
    calendar_result = load_current_calendar_blocks(config)
    tz = ZoneInfo(config["timezone"])
    strategy = config.get("offer_strategy", {})
    start_day = date.fromisoformat(date_filter) if date_filter else datetime.now(tz).date()
    days = 1 if date_filter else max(int(config.get("planning_days", 30)), max_horizon_end_offset(strategy) + 1)
    high_demand = set(strategy.get("high_demand_course_keys", []))
    default_limit = int(strategy.get("default_max_options_per_course_page", 6))
    high_demand_limit = int(strategy.get("high_demand_max_options_per_course_page", 8))
    max_per_day = int(strategy.get("max_options_per_day_per_course_page", 3))
    global_minimum_notice_minutes = int(strategy.get("minimum_customer_notice_minutes", 120))

    grouped: list[dict[str, Any]] = []
    suppressed: list[dict[str, Any]] = []
    frequency_assignments: list[dict[str, Any]] = []
    total_offers = 0
    for course in config.get("course_templates", []):
        course_key = course.get("course_key", "")
        if course_key_filter and course_key != course_key_filter:
            continue
        group_name, frequency_group, assignment_reason = resolve_offer_frequency_group(course, config)
        notice_minutes, notice_source = notice_minutes_for_group(strategy, group_name)
        earliest_offer_start = datetime.now(tz) + timedelta(minutes=notice_minutes)
        public_offer_enabled = bool(frequency_group.get("public_offer_enabled", True))
        group_suppressed = False
        group_suppression_reasons: list[str] = []
        if not public_offer_enabled:
            group_suppressed = True
            group_suppression_reasons.append("suppressed because offer frequency group has public_offer_enabled=false")
        frequency_assignments.append({
            "course_key": course_key,
            "assigned_offer_frequency_group": group_name,
            "assignment_reason": assignment_reason,
            "same_program_gap_hours_used": frequency_group.get("same_program_gap_hours"),
            "allow_same_day_reoffer": frequency_group.get("allow_same_day_reoffer", True),
            "max_offers_per_day_after_claim": frequency_group.get("max_offers_per_day_after_claim"),
            "max_offers_per_week": frequency_group.get("max_offers_per_week"),
            "public_offer_enabled": public_offer_enabled,
            "minimum_customer_notice_minutes_used": notice_minutes,
            "minimum_customer_notice_window_source": notice_source,
            "suppressed_because_of_group_frequency_rules": group_suppressed,
            "suppression_reasons": group_suppression_reasons,
        })
        course_limit = int(limit or (high_demand_limit if course_key in high_demand else default_limit))
        candidates: list[dict[str, Any]] = []

        if not public_offer_enabled:
            grouped.append({
                "course_key": course_key,
                "course_title": course.get("course_title"),
                "page_context": f"course:{course_key}",
                "offer_frequency_group": group_name,
                "offered_options": offers,
            })
            suppressed.append({
                "course_key": course_key,
                "date_time": "",
                "suppression_reason": "suppressed because offer frequency group has public_offer_enabled=false",
                "offer_frequency_group": group_name,
            })
            continue

        for offset in range(days):
            day = start_day + timedelta(days=offset)
            seeds = build_offer_seeds(day, calendar_result.blocks, course, config)
            if offset > 0:
                for seed in seeds:
                    if seed["offer_type"] == "clean_slate_anchor":
                        seed["offer_type"] = "next_day_anchor"

            for seed in seeds:
                start = seed["start"]
                if start < earliest_offer_start:
                    reason = (
                        "suppressed because offer is inside group-specific minimum customer notice window"
                        if notice_source == "group"
                        else "suppressed because offer is inside minimum customer notice window"
                    )
                    suppressed.append({
                        "course_key": course_key,
                        "date_time": start.isoformat(),
                        "suppression_reason": reason,
                        "offer_frequency_group": group_name,
                        "minimum_customer_notice_minutes_used": notice_minutes,
                        "minimum_customer_notice_window_source": notice_source,
                    })
                    continue
                day_key = start.date().isoformat()
                same_program_claimed_today = has_same_program_claim_on_day(course, day_key, calendar_result.blocks, config)
                if same_program_claimed_today and not bool(frequency_group.get("allow_same_day_reoffer", True)):
                    suppressed.append({
                        "course_key": course_key,
                        "date_time": start.isoformat(),
                        "suppression_reason": "suppressed because offer frequency group does not allow same-day reoffer after a claimed same-program session",
                        "offer_frequency_group": group_name,
                    })
                    continue
                gap_passed, gap_reason = same_program_gap_status(course, start, calendar_result.blocks, [], config, frequency_group)
                if not gap_passed:
                    suppressed.append({
                        "course_key": course_key,
                        "date_time": start.isoformat(),
                        "suppression_reason": (gap_reason or {}).get("suppression_reason", "suppressed because same-program gap against real canonical session"),
                        "offer_frequency_group": group_name,
                        **(gap_reason or {}),
                    })
                    continue
                session, rejection = evaluate_candidate_session(course, start, calendar_result.blocks, config)
                if not session:
                    suppressed.append({
                        "course_key": course_key,
                        "date_time": start.isoformat(),
                        "suppression_reason": (rejection or {}).get("reason", "suppressed because availability re-evaluation failed"),
                        "offer_frequency_group": group_name,
                    })
                    continue
                offer = build_offer_record(session, seed["offer_type"], seed.get("context"), True, True, notice_minutes, notice_source)
                candidates.append({**session, **offer, "course_key": course_key, "offer_frequency_group": group_name})

        spread_config = horizon_spread_config(strategy, group_name)
        offers, cap_suppressed, selection_debug = select_customer_facing_offers(
            candidates,
            course=course,
            course_limit=course_limit,
            start_day=start_day,
            spread_config=spread_config,
            calendar_result=calendar_result,
            config=config,
            frequency_group=frequency_group,
        )
        suppressed.extend(cap_suppressed)

        grouped.append({
            "course_key": course_key,
            "course_title": course.get("course_title"),
            "course_display_name": course_display_name(course_key, config) or course.get("course_title"),
            "page_context": f"course:{course_key}",
            "offer_frequency_group": group_name,
            "offer_selection_debug": selection_debug,
            "offered_options": offers,
        })
        total_offers += len(offers)

    payload = {
        "generated_at": datetime.now(tz).isoformat(),
        "offer_strategy": strategy,
        "course_count": len(grouped),
        "offer_count": total_offers,
        "offer_frequency_assignments": frequency_assignments,
        "courses": grouped,
    }
    OFFERS_PATH.parent.mkdir(parents=True, exist_ok=True)
    OFFERS_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    report = {
        **payload,
        "suppressed_options": suppressed,
        "suppressed_count": len(suppressed),
        "offer_frequency_assignments": frequency_assignments,
        "calendar_counts": {
            "adr_events_found": len(calendar_result.adr_events),
            "brian_adr_blocks_found": len(calendar_result.brian_adr_blocks),
            "do_not_schedule_blocks_found": len(calendar_result.dns_blocks),
            "existing_enrollware_sessions_assigned_to_brian": len(calendar_result.enrollware_brian_blocks),
        },
        "private_do_not_schedule_url": "[REDACTED]",
    }
    OFFER_REPORT_JSON_PATH.write_text(json.dumps(report, indent=2), encoding="utf-8")
    lines = [
        "# Free-Time Offer Strategy Report",
        "",
        f"- Generated at: {payload['generated_at']}",
        f"- Courses: {len(grouped)}",
        f"- Customer-facing offers: {total_offers}",
        f"- Suppressed options: {len(suppressed)}",
        f"- Normal hours only: {strategy.get('normal_hours_only')}",
        f"- Global minimum customer notice minutes: {global_minimum_notice_minutes}",
        f"- Overnight anchors allowed: {strategy.get('allow_overnight_anchors')}",
        f"- Private DoNotSchedule URL: [REDACTED]",
    ]
    lines.extend(["", "## Offer Frequency Assignments"])
    for assignment in frequency_assignments:
        lines.append(
            "- {course_key}: {group} ({reason}); gap={gap}; same-day reoffer={same_day}; notice={notice} ({notice_source}); public={public}; suppressed={suppressed}".format(
                course_key=assignment["course_key"],
                group=assignment["assigned_offer_frequency_group"],
                reason=assignment["assignment_reason"],
                gap=assignment["same_program_gap_hours_used"],
                same_day=assignment["allow_same_day_reoffer"],
                notice=assignment["minimum_customer_notice_minutes_used"],
                notice_source=assignment["minimum_customer_notice_window_source"],
                public=assignment["public_offer_enabled"],
                suppressed=assignment["suppressed_because_of_group_frequency_rules"],
            )
        )
        for reason in assignment.get("suppression_reasons", []):
            lines.append(f"  - {reason}")
    lines.extend(["", "## Offers By Course"])
    for course in grouped:
        lines.append(f"- {course['course_key']} ({course['course_title']}): {len(course['offered_options'])}")
        debug = course.get("offer_selection_debug", {})
        if debug:
            lines.append(f"  - Max offers allowed: {debug.get('max_offers_allowed')}")
            lines.append(f"  - Horizon spread enabled: {debug.get('horizon_spread_enabled')}")
            lines.append(f"  - Valid candidates by horizon bucket: {debug.get('valid_candidates_by_horizon_bucket')}")
            lines.append(f"  - Selected offers by horizon bucket: {debug.get('selected_offers_by_horizon_bucket')}")
            lines.append(f"  - Unused bucket slots: {debug.get('unused_bucket_slots')}")
            lines.append(f"  - Fill-from-earliest count: {debug.get('fill_from_earliest_count')}")
            lines.append(f"  - Final selected offers by month: {debug.get('final_selected_offers_by_month')}")
        for offer in course["offered_options"][:8]:
            lines.append(
                f"  - {offer['start_time']} to {offer['end_time']} | {offer['offer_type']} | {offer['customer_label']} | bucket={offer.get('horizon_bucket_key')}"
            )
    lines.extend(["", "## Suppressed Options"])
    for item in suppressed[:200]:
        lines.append(f"- {item['course_key']} {item['date_time']}: {item['suppression_reason']}")
    if not suppressed:
        lines.append("- None")
    OFFER_REPORT_MD_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"Generated {total_offers} customer-facing offers across {len(grouped)} course page(s).")
    print(f"Suppressed options: {len(suppressed)}")
    print(f"Wrote {OFFERS_PATH.relative_to(ROOT)}")
    return 0


def canonical_class_report_enabled(config: dict[str, Any]) -> tuple[bool, Path]:
    class_report_config = config.get("enrollware_class_report", {})
    class_report_path = (ROOT / str(class_report_config.get("path", "data/private/Class report.xlsx"))).resolve()
    enabled = bool(class_report_config.get("enabled", False))
    prefer = bool(class_report_config.get("prefer_over_schedule_future", True))
    return enabled and prefer and class_report_path.exists(), class_report_path


def write_canonical_schedule_from_class_report(config: dict[str, Any]) -> dict[str, Any]:
    authoritative, class_report_path = canonical_class_report_enabled(config)
    sessions = load_class_report_sessions(class_report_path, config) if authoritative else []
    payload = {
        "build": {
            "source_mode": "class_report_authoritative" if authoritative else "not_authoritative",
            "class_report": str(class_report_path),
            "generated_at": datetime.now(ZoneInfo(config["timezone"])).isoformat(),
            "session_count": len(sessions),
        },
        "sessions": sessions,
    }
    CANONICAL_SCHEDULE_PATH.parent.mkdir(parents=True, exist_ok=True)
    CANONICAL_SCHEDULE_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return payload


def generate_reports(
    config: dict[str, Any],
    adr_events: list[CalendarEvent],
    brian_adr_blocks: list[CalendarEvent],
    dns_blocks: list[CalendarEvent],
    accepted: list[dict[str, Any]],
    rejected: list[dict[str, Any]],
    warnings: list[str],
    do_not_schedule_source: str,
    do_not_schedule_private: bool,
    do_not_schedule_loaded: bool,
    enrollware_existing_sessions_found: int,
    class_report_sessions_found: int,
    schedule_future_sessions_found: int,
    merged_existing_enrollware_sessions: int,
    existing_enrollware_source_mode: str,
    duplicate_existing_enrollware_sessions_deduped: int,
    existing_enrollware_source_mismatches: list[str],
    hot_sync_active_entries_found: int,
    hot_sync_entries_absorbed: int,
    hot_sync_entries_kept_active: int,
    hot_sync_entries_flagged_for_review: int,
    hot_sync_entries_merged: int,
    hot_sync_active_path_exists: bool,
    class_report_found: bool,
    schedule_future_fallback_used: bool,
    enrollware_brian_blocks: list[CalendarEvent],
    enrollware_non_brian_ignored: int,
) -> None:
    DEBUG_DIR.mkdir(parents=True, exist_ok=True)
    report = {
        "calendars_loaded": {
            "adr": config["calendars"]["adr"]["ics_url"],
            "do_not_schedule": {
                "status": "loaded" if do_not_schedule_loaded else "not_loaded",
                "source": do_not_schedule_source,
                "ics_url": "[REDACTED]" if do_not_schedule_private else config["calendars"]["do_not_schedule"]["ics_url"],
            },
        },
        "adr_events_found": len(adr_events),
        "brian_adr_blocks_found": len(brian_adr_blocks),
        "do_not_schedule_blocks_found": len(dns_blocks),
        "existing_enrollware_sessions_found": enrollware_existing_sessions_found,
        "class_report_xlsx_found": class_report_found,
        "class_report_sessions_found": class_report_sessions_found,
        "schedule_future_fallback_used": schedule_future_fallback_used,
        "schedule_future_sessions_found": schedule_future_sessions_found,
        "merged_existing_enrollware_sessions": merged_existing_enrollware_sessions,
        "existing_enrollware_source_mode": existing_enrollware_source_mode,
        "duplicate_existing_enrollware_sessions_deduped": duplicate_existing_enrollware_sessions_deduped,
        "existing_enrollware_source_mismatches": existing_enrollware_source_mismatches,
        "hot_sync_active_path_exists": hot_sync_active_path_exists,
        "hot_sync_active_entries_found": hot_sync_active_entries_found,
        "hot_sync_entries_absorbed": hot_sync_entries_absorbed,
        "hot_sync_entries_kept_active": hot_sync_entries_kept_active,
        "hot_sync_entries_flagged_for_review": hot_sync_entries_flagged_for_review,
        "hot_sync_entries_merged": hot_sync_entries_merged,
        "existing_enrollware_sessions_assigned_to_brian": len(enrollware_brian_blocks),
        "existing_enrollware_sessions_ignored_non_brian": enrollware_non_brian_ignored,
        "proposed_sessions_rejected_because_of_brian_enrollware_sessions": sum(
            1 for item in rejected if item.get("rejection_source") == "enrollware_existing_session"
        ),
        "brian_enrollware_blocks": [block.metadata for block in enrollware_brian_blocks],
        "accepted_proposed_sessions": accepted,
        "rejected_candidate_sessions": rejected,
        "calendar_fetch_warnings": warnings,
    }
    REPORT_JSON_PATH.write_text(json.dumps(report, indent=2), encoding="utf-8")

    lines = [
        "# Free-Time Scheduler Report",
        "",
        "## Calendars Loaded",
        f"- ADR Schedule ICS: {config['calendars']['adr']['ics_url']}",
        f"- DoNotSchedule calendar {'loaded successfully' if do_not_schedule_loaded else 'did not load successfully'} from {do_not_schedule_source}.",
        f"- DoNotSchedule ICS URL: {'[REDACTED]' if do_not_schedule_private else config['calendars']['do_not_schedule']['ics_url']}",
        "",
        "## Counts",
        f"- ADR events found: {len(adr_events)}",
        f"- Brian ADR blocks found: {len(brian_adr_blocks)}",
        f"- DoNotSchedule blocks found: {len(dns_blocks)}",
        f"- Class Report XLSX found: {class_report_found}",
        f"- Class Report sessions found: {class_report_sessions_found}",
        f"- schedule_future.json fallback used: {schedule_future_fallback_used}",
        f"- schedule_future sessions found: {schedule_future_sessions_found}",
        f"- Merged existing Enrollware sessions: {merged_existing_enrollware_sessions}",
        f"- Existing Enrollware source mode: {existing_enrollware_source_mode}",
        f"- Duplicate existing Enrollware sessions deduped: {duplicate_existing_enrollware_sessions_deduped}",
        f"- Hot-sync active file exists: {hot_sync_active_path_exists}",
        f"- Hot-sync active entries found: {hot_sync_active_entries_found}",
        f"- Hot-sync entries absorbed by Class Report: {hot_sync_entries_absorbed}",
        f"- Hot-sync entries kept active: {hot_sync_entries_kept_active}",
        f"- Hot-sync entries flagged for review: {hot_sync_entries_flagged_for_review}",
        f"- Hot-sync entries merged into blocking source: {hot_sync_entries_merged}",
        f"- Existing Enrollware sessions found: {enrollware_existing_sessions_found}",
        f"- Existing Enrollware sessions assigned to Brian: {len(enrollware_brian_blocks)}",
        f"- Existing Enrollware sessions ignored because instructor was not Brian: {enrollware_non_brian_ignored}",
        f"- Accepted proposed sessions: {len(accepted)}",
        f"- Rejected candidate sessions: {len(rejected)}",
        f"- Proposed sessions rejected because of Brian Enrollware sessions: {sum(1 for item in rejected if item.get('rejection_source') == 'enrollware_existing_session')}",
        f"- Existing Enrollware source mismatches: {len(existing_enrollware_source_mismatches)}",
        "",
        "## Calendar Fetch Warnings",
    ]
    lines.extend([f"- {warning}" for warning in warnings] or ["- None"])
    lines.extend(["", "## Accepted Proposed Sessions"])
    lines.extend(
        [
            f"- {item['start_time']} - {item['course_title']} at {item['location_name']} ({item['page_slug']})"
            for item in accepted
        ]
        or ["- None"]
    )
    lines.extend(["", "## Rejected Candidate Sessions"])
    lines.extend(
        [
            f"- {item.get('start_time', 'unknown')} - {item.get('course_title', item.get('course_key', 'unknown'))}: {item['reason']}"
            for item in rejected
        ]
        or ["- None"]
    )
    if existing_enrollware_source_mismatches:
        lines.extend(["", "## Existing Enrollware Source Mismatches"])
        lines.extend([f"- {item}" for item in existing_enrollware_source_mismatches])
    REPORT_MD_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def render_button(session: dict[str, Any]) -> str:
    url = session.get("enrollware_enroll_url")
    if url:
        return f'<a class="button" href="{html.escape(url)}">Book This Class</a>'
    return '<button class="button disabled" disabled>Proposed Time - Not Yet Bookable</button>'


def page_shell(title: str, body: str) -> str:
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="robots" content="noindex,nofollow">
  <title>{html.escape(title)}</title>
  <style>
    body {{ margin: 0; font-family: Arial, sans-serif; color: #1f2933; background: #f7f9fb; }}
    main {{ max-width: 860px; margin: 0 auto; padding: 32px 18px; }}
    .panel {{ background: #fff; border: 1px solid #d9e2ec; border-radius: 8px; padding: 24px; }}
    .warning {{ background: #fff7d6; border: 1px solid #e7c76f; padding: 12px 14px; border-radius: 6px; margin: 18px 0; }}
    .button {{ display: inline-block; padding: 12px 16px; border-radius: 6px; border: 1px solid #1264a3; background: #1264a3; color: #fff; text-decoration: none; font-weight: 700; }}
    .button.disabled {{ background: #e4e7eb; color: #52606d; border-color: #cbd2d9; }}
    table {{ width: 100%; border-collapse: collapse; background: #fff; }}
    th, td {{ text-align: left; border-bottom: 1px solid #d9e2ec; padding: 10px; vertical-align: top; }}
    h1 {{ margin-top: 0; }}
  </style>
</head>
<body>
<main>
{body}
</main>
</body>
</html>
"""


def format_dt(value: str) -> str:
    parsed = datetime.fromisoformat(value)
    return parsed.strftime("%A, %B %-d, %Y at %-I:%M %p") if sys.platform != "win32" else parsed.strftime("%A, %B %#d, %Y at %#I:%M %p")


def generate_preview_pages(sessions: list[dict[str, Any]]) -> None:
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)
    for session in sessions:
        warning = ""
        if session["session_status"] == "proposed":
            warning = '<div class="warning">This time is proposed but not yet published in Enrollware.</div>'
        body = f"""<article class="panel">
  <h1>{html.escape(session['course_title'])}</h1>
  {warning}
  <p><strong>Date/time:</strong> {html.escape(format_dt(session['start_time']))} to {html.escape(format_dt(session['end_time']).split(' at ')[-1])}</p>
  <p><strong>Location:</strong> {html.escape(session['location_name'])}<br>{html.escape(session['location_address'])}</p>
  <p><strong>Proposed capacity:</strong> {session['capacity']}</p>
  <p><strong>Status:</strong> {html.escape(session['session_status'])}</p>
  <p>{html.escape(session['availability_explanation'])}</p>
  {render_button(session)}
</article>"""
        (PREVIEW_DIR / f"{session['page_slug']}.html").write_text(page_shell(session["course_title"], body), encoding="utf-8")

    grouped: dict[str, list[dict[str, Any]]] = {}
    for session in sessions:
        grouped.setdefault(session["start_time"][:10], []).append(session)
    sections: list[str] = ["<h1>Proposed Sessions</h1>"]
    for day in sorted(grouped):
        sections.append(f"<h2>{html.escape(day)}</h2><table><thead><tr><th>Course</th><th>Start</th><th>End</th><th>Location</th><th>Status</th><th>Preview</th></tr></thead><tbody>")
        for session in sorted(grouped[day], key=lambda item: item["start_time"]):
            sections.append(
                "<tr>"
                f"<td>{html.escape(session['course_title'])}</td>"
                f"<td>{html.escape(datetime.fromisoformat(session['start_time']).strftime('%I:%M %p'))}</td>"
                f"<td>{html.escape(datetime.fromisoformat(session['end_time']).strftime('%I:%M %p'))}</td>"
                f"<td>{html.escape(session['location_name'])}</td>"
                f"<td>{html.escape(session['session_status'])}</td>"
                f"<td><a href=\"{html.escape(session['page_slug'])}.html\">Preview</a></td>"
                "</tr>"
            )
        sections.append("</tbody></table>")
    if not sessions:
        sections.append("<p>No proposed sessions were accepted.</p>")
    (PREVIEW_DIR / "index.html").write_text(page_shell("Proposed Sessions", "\n".join(sections)), encoding="utf-8")


def load_current_calendar_blocks(config: dict[str, Any]) -> CalendarLoadResult:
    tz = ZoneInfo(config["timezone"])
    warnings: list[str] = []
    do_not_schedule_ics_url, do_not_schedule_source, do_not_schedule_private = resolve_do_not_schedule_ics_url(config)

    adr_raw, adr_warning = fetch_ics_events("adr", config["calendars"]["adr"]["ics_url"], tz)
    dns_raw, dns_warning = fetch_ics_events("do_not_schedule", do_not_schedule_ics_url, tz)
    do_not_schedule_loaded = dns_warning is None
    if adr_warning:
        warnings.append(adr_warning)
    if dns_warning:
        warnings.append(
            dns_warning
            + " The calendar may need to be made public or accessed through the Google Calendar API later."
        )

    adr_events = [event for raw in adr_raw if (event := normalize_event(raw, "adr", config))]
    brian_adr_blocks = [event for event in adr_events if is_brian_adr_event(event)]
    dns_blocks = [event for raw in dns_raw if (event := normalize_event(raw, "do_not_schedule", config))]
    enrollware_result = load_existing_enrollware_blocks(config)
    enrollware_brian_blocks = enrollware_result["brian_blocks"]
    enrollware_non_brian_ignored = enrollware_result["non_brian_ignored"]
    blocks = sorted([*brian_adr_blocks, *dns_blocks, *enrollware_brian_blocks], key=lambda event: event.start)
    return CalendarLoadResult(
        adr_events=adr_events,
        brian_adr_blocks=brian_adr_blocks,
        dns_blocks=dns_blocks,
        blocks=blocks,
        warnings=warnings,
        do_not_schedule_source=do_not_schedule_source,
        do_not_schedule_private=do_not_schedule_private,
        do_not_schedule_loaded=do_not_schedule_loaded,
        enrollware_existing_sessions_found=enrollware_result["existing_enrollware_sessions_found"],
        class_report_sessions_found=enrollware_result["class_report_sessions_found"],
        schedule_future_sessions_found=enrollware_result["schedule_future_sessions_found"],
        merged_existing_enrollware_sessions=enrollware_result["merged_existing_enrollware_sessions"],
        existing_enrollware_source_mode=enrollware_result["existing_enrollware_source_mode"],
        duplicate_existing_enrollware_sessions_deduped=enrollware_result["duplicate_existing_enrollware_sessions_deduped"],
        existing_enrollware_source_mismatches=enrollware_result["existing_enrollware_source_mismatches"],
        hot_sync_active_entries_found=enrollware_result["hot_sync_active_entries_found"],
        hot_sync_entries_absorbed=enrollware_result["hot_sync_entries_absorbed"],
        hot_sync_entries_kept_active=enrollware_result["hot_sync_entries_kept_active"],
        hot_sync_entries_flagged_for_review=enrollware_result["hot_sync_entries_flagged_for_review"],
        hot_sync_entries_merged=enrollware_result["hot_sync_entries_merged"],
        hot_sync_active_path_exists=enrollware_result["hot_sync_active_path_exists"],
        class_report_found=enrollware_result["class_report_found"],
        schedule_future_fallback_used=enrollware_result["schedule_future_fallback_used"],
        enrollware_brian_blocks=enrollware_brian_blocks,
        enrollware_non_brian_ignored=enrollware_non_brian_ignored,
    )


def load_proposed_sessions() -> list[dict[str, Any]]:
    if not PROPOSED_SESSIONS_PATH.exists():
        return []
    with PROPOSED_SESSIONS_PATH.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def normalize_request_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def parse_scheduler_datetime(value: Any, config: dict[str, Any]) -> datetime | None:
    try:
        return parse_existing_session_time(value, ZoneInfo(config["timezone"]))
    except Exception:
        return None


def request_match_key(course_key: Any, start: Any, location_name: Any, config: dict[str, Any]) -> tuple[str, str, str]:
    parsed = parse_scheduler_datetime(start, config)
    location_key = normalize_request_key(clean_location_name(location_name))
    try:
        location_key = normalize_request_key(resolve_location(str(location_name or ""), config).name)
    except Exception:
        pass
    return (
        str(course_key or "").strip(),
        parsed.isoformat() if parsed else str(start or "").strip(),
        location_key,
    )


def clean_location_name(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"<br\s*/?>", " ", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if "::" in text:
        text = text.rsplit("::", 1)[-1].strip()
    return text or "Location TBA"


def load_customer_facing_offer_rows() -> list[dict[str, Any]]:
    if not OFFERS_PATH.exists():
        return []
    payload = json.loads(OFFERS_PATH.read_text(encoding="utf-8"))
    rows: list[dict[str, Any]] = []
    for course in payload.get("courses", []) if isinstance(payload, dict) else []:
        if not isinstance(course, dict):
            continue
        for option in course.get("offered_options", []):
            if not isinstance(option, dict):
                continue
            row = dict(option)
            row["course_key"] = str(row.get("course_key") or course.get("course_key") or "").strip()
            row["course_title"] = str(row.get("course_title") or course.get("course_title") or "").strip()
            rows.append(row)
    return rows


def find_matching_customer_offer(payload: dict[str, Any], config: dict[str, Any]) -> dict[str, Any] | None:
    course_key = str(payload.get("course_key") or "").strip()
    requested_start = payload.get("requested_start")
    requested_location = clean_location_name(payload.get("location_name") or payload.get("location") or "")
    for offer in load_customer_facing_offer_rows():
        offer_location = clean_location_name(offer.get("location_name") or offer.get("location_address") or "")
        if request_match_key(course_key, requested_start, requested_location, config) == request_match_key(
            offer.get("course_key"),
            offer.get("start_time"),
            offer_location,
            config,
        ):
            return offer
    return None


def course_from_customer_offer(offer: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    template = next(
        (item for item in config.get("course_templates", []) if item.get("course_key") == offer.get("course_key")),
        {},
    )
    start = datetime.fromisoformat(offer["start_time"])
    end = datetime.fromisoformat(offer["end_time"])
    return {
        "course_key": offer["course_key"],
        "course_title": offer.get("course_title") or template.get("course_title") or offer["course_key"],
        "duration_minutes": int((end - start).total_seconds() // 60),
        "cleanup_minutes": int(offer.get("cleanup_minutes") or template.get("cleanup_minutes") or 0),
        "location": offer.get("location_name") or template.get("location") or "Wilmington Office",
        "capacity": int(offer.get("capacity") or template.get("capacity") or 0),
        "enrollware_class_id": None,
        "enrollware_enroll_url": None,
    }


def load_canonical_schedule_sessions() -> list[dict[str, Any]]:
    return load_json_sessions(CANONICAL_SCHEDULE_PATH)


def existing_enroll_url(session: dict[str, Any]) -> str:
    return str(
        session.get("enrollware_enroll_url")
        or session.get("registration_url")
        or session.get("enrollware_url")
        or ""
    ).strip()


def find_existing_canonical_match(course_key: str, start: str, location_name: str, config: dict[str, Any]) -> dict[str, Any] | None:
    target = request_match_key(course_key, start, location_name, config)
    for session in load_canonical_schedule_sessions():
        session_course_key = canonical_session_course_key(session)
        session_location = clean_location_name(first_present(session, ["location_name", "location_display", "location"]))
        session_start = first_present(session, ["start_at", "start", "start_time"])
        if target == request_match_key(session_course_key, session_start, session_location, config):
            return session
    return None


def canonical_session_course_key(session: dict[str, Any]) -> str:
    explicit = str(session.get("course_key") or "").strip()
    if explicit:
        return explicit
    text = f"{session.get('course_name') or ''} {session.get('official_course_name') or ''}".lower()
    if "bls" in text and ("heartcode" in text or "skills" in text):
        return "heartcode-bls-skills"
    if "bls" in text and "renewal" in text:
        return "bls-renewal"
    if "bls" in text and "instructor" not in text:
        return "bls"
    if "heartsaver" in text and "first aid" in text:
        return "heartsaver-fa-cpr-aed"
    return ""


def hot_sync_active_path(config: dict[str, Any]) -> Path:
    hot_config = config.get("hot_sync_delta", {})
    return (ROOT / str(hot_config.get("active_path", "data/runtime/free_time_scheduler/hot_sync_active.json"))).resolve()


def find_hot_sync_match(course_key: str, start: str, location_name: str, config: dict[str, Any]) -> dict[str, Any] | None:
    target = request_match_key(course_key, start, location_name, config)
    for entry in load_hot_sync_entries(hot_sync_active_path(config)):
        normalized = normalize_hot_sync_entry(entry)
        entry_course_key = str(first_present(normalized, ["course_key", "program_key"]) or "").strip()
        entry_start = first_present(normalized, ["start", "start_time", "start_at"])
        entry_location = clean_location_name(first_present(normalized, ["location_name", "location_display", "location"]))
        if target == request_match_key(entry_course_key, entry_start, entry_location, config):
            return entry
    return None


def click_time_creation_enabled(config: dict[str, Any]) -> bool:
    creation = config.get("click_time_creation", {})
    return (
        creation.get("enabled") is True
        and creation.get("dry_run") is False
        and creation.get("allow_public_creation") is True
    )


def safe_user_rejection_reason(reason: str | None) -> str:
    if not reason:
        return "That time is no longer available. Please choose another option."
    return "That time is no longer available. Please choose another option."


def check_customer_facing_offer_request(payload: dict[str, Any]) -> dict[str, Any]:
    config = load_config()
    if str(payload.get("offer_source") or "") != "customer_facing_offers":
        return {
            "ok": False,
            "status": "invalid",
            "message": "That time could not be verified. Please choose another option.",
            "reason": "offer_source must be customer_facing_offers",
        }

    offer = find_matching_customer_offer(payload, config)
    if not offer:
        return {
            "ok": False,
            "status": "invalid",
            "message": "That time could not be verified. Please choose another option.",
            "reason": "requested course/time/location was not found in customer_facing_offers.json",
        }

    course_key = str(offer["course_key"])
    start = str(offer["start_time"])
    location_name = clean_location_name(offer.get("location_name") or payload.get("location_name"))

    canonical_match = find_existing_canonical_match(course_key, start, location_name, config)
    if canonical_match:
        enroll_url = existing_enroll_url(canonical_match)
        return {
            "ok": True,
            "status": "existing_canonical",
            "available": True,
            "created": False,
            "existing": True,
            "message": "A booking link already exists for this time.",
            "enrollware_enroll_url": enroll_url,
            "enrollware_class_id": first_present(canonical_match, ["session_id", "class_id", "enrollware_class_id"]),
            "no_duplicate_created": True,
        }

    hot_sync_match = find_hot_sync_match(course_key, start, location_name, config)
    if hot_sync_match:
        enroll_url = existing_enroll_url(hot_sync_match)
        return {
            "ok": True,
            "status": "existing_hot_sync",
            "available": bool(enroll_url),
            "created": False,
            "existing": True,
            "message": "This time is already being processed.",
            "enrollware_enroll_url": enroll_url,
            "enrollware_class_id": first_present(hot_sync_match, ["enrollware_class_id", "class_id", "session_id"]),
            "no_duplicate_created": True,
        }

    calendar_result = load_current_calendar_blocks(config)
    if calendar_result.warnings:
        return {
            "ok": False,
            "status": "unavailable",
            "available": False,
            "created": False,
            "message": "That time is no longer available. Please choose another option.",
            "reason": "calendar warning during click-time recheck",
        }

    course = course_from_customer_offer(offer, config)
    accepted, rejection = evaluate_candidate_session(course, datetime.fromisoformat(start), calendar_result.blocks, config)
    if not accepted:
        return {
            "ok": True,
            "status": "unavailable",
            "available": False,
            "created": False,
            "message": safe_user_rejection_reason((rejection or {}).get("reason")),
            "reason": (rejection or {}).get("reason"),
        }

    if not click_time_creation_enabled(config):
        return {
            "ok": True,
            "status": "dry_run_available",
            "available": True,
            "dry_run_available": True,
            "created": False,
            "message": "Preview only: this time passed recheck, but Enrollware creation is not wired yet.",
            "click_time_recheck_passed": True,
            "enrollware_creation_wired": False,
        }

    return {
        "ok": False,
        "status": "creation_not_implemented",
        "available": True,
        "created": False,
        "message": "Preview only: this time passed recheck, but Enrollware creation is not wired yet.",
        "click_time_recheck_passed": True,
        "enrollware_creation_wired": False,
    }


def list_proposed_sessions(limit: int) -> int:
    sessions = load_proposed_sessions()
    for session in sessions[:limit]:
        print(
            f"{session['page_slug']} | {session['course_title']} | "
            f"{session['start_time']} | {session['location_name']}"
        )
    print(f"Listed {min(limit, len(sessions))} of {len(sessions)} proposed sessions.")
    return 0


def course_from_proposed_session(session: dict[str, Any]) -> dict[str, Any]:
    start = datetime.fromisoformat(session["start_time"])
    end = datetime.fromisoformat(session["end_time"])
    return {
        "course_key": session["course_key"],
        "course_title": session["course_title"],
        "duration_minutes": int((end - start).total_seconds() // 60),
        "cleanup_minutes": session["cleanup_minutes"],
        "location": session["location_name"],
        "capacity": session["capacity"],
        "enrollware_class_id": session.get("enrollware_class_id"),
        "enrollware_enroll_url": session.get("enrollware_enroll_url"),
    }


def write_recheck_report(
    session: dict[str, Any],
    calendar_result: CalendarLoadResult,
    passed: bool,
    reason: str | None,
    evaluated_at: str,
) -> None:
    DEBUG_DIR.mkdir(parents=True, exist_ok=True)
    report = {
        "page_slug": session["page_slug"],
        "course_key": session["course_key"],
        "course_title": session["course_title"],
        "start_time": session["start_time"],
        "end_time": session["end_time"],
        "location_name": session["location_name"],
        "location_address": session["location_address"],
        "evaluated_at": evaluated_at,
        "result": "pass" if passed else "fail",
        "rejection_reason": reason,
        "calendar_counts": {
            "adr_events_found": len(calendar_result.adr_events),
            "brian_adr_blocks_found": len(calendar_result.brian_adr_blocks),
            "do_not_schedule_blocks_found": len(calendar_result.dns_blocks),
            "existing_enrollware_sessions_found": calendar_result.enrollware_existing_sessions_found,
            "class_report_xlsx_found": calendar_result.class_report_found,
            "class_report_sessions_found": calendar_result.class_report_sessions_found,
            "schedule_future_fallback_used": calendar_result.schedule_future_fallback_used,
            "schedule_future_sessions_found": calendar_result.schedule_future_sessions_found,
            "merged_existing_enrollware_sessions": calendar_result.merged_existing_enrollware_sessions,
            "existing_enrollware_source_mode": calendar_result.existing_enrollware_source_mode,
            "duplicate_existing_enrollware_sessions_deduped": calendar_result.duplicate_existing_enrollware_sessions_deduped,
            "hot_sync_active_path_exists": calendar_result.hot_sync_active_path_exists,
            "hot_sync_active_entries_found": calendar_result.hot_sync_active_entries_found,
            "hot_sync_entries_absorbed": calendar_result.hot_sync_entries_absorbed,
            "hot_sync_entries_kept_active": calendar_result.hot_sync_entries_kept_active,
            "hot_sync_entries_flagged_for_review": calendar_result.hot_sync_entries_flagged_for_review,
            "hot_sync_entries_merged": calendar_result.hot_sync_entries_merged,
            "existing_enrollware_sessions_assigned_to_brian": len(calendar_result.enrollware_brian_blocks),
            "existing_enrollware_sessions_ignored_non_brian": calendar_result.enrollware_non_brian_ignored,
        },
        "private_do_not_schedule_url": "[REDACTED]",
        "private_do_not_schedule_url_redacted": True,
    }
    RECHECK_JSON_PATH.write_text(json.dumps(report, indent=2), encoding="utf-8")

    status_lines = [
        "# Free-Time Scheduler Recheck",
        "",
        f"- page_slug: {session['page_slug']}",
        f"- course_key: {session['course_key']}",
        f"- course_title: {session['course_title']}",
        f"- start_time: {session['start_time']}",
        f"- end_time: {session['end_time']}",
        f"- location_name: {session['location_name']}",
        f"- location_address: {session['location_address']}",
        f"- evaluated_at: {evaluated_at}",
        f"- result: {'PASS' if passed else 'FAIL'}",
    ]
    if reason:
        status_lines.append(f"- rejection_reason: {reason}")
    status_lines.extend(
        [
            "",
            "## Calendar Counts",
            f"- ADR events found: {len(calendar_result.adr_events)}",
            f"- Brian ADR blocks found: {len(calendar_result.brian_adr_blocks)}",
            f"- DoNotSchedule blocks found: {len(calendar_result.dns_blocks)}",
            f"- Existing Enrollware sessions found: {calendar_result.enrollware_existing_sessions_found}",
            f"- Class Report XLSX found: {calendar_result.class_report_found}",
            f"- Class Report sessions found: {calendar_result.class_report_sessions_found}",
            f"- schedule_future fallback used: {calendar_result.schedule_future_fallback_used}",
            f"- schedule_future sessions found: {calendar_result.schedule_future_sessions_found}",
            f"- Merged existing Enrollware sessions: {calendar_result.merged_existing_enrollware_sessions}",
            f"- Existing Enrollware source mode: {calendar_result.existing_enrollware_source_mode}",
            f"- Duplicate existing Enrollware sessions deduped: {calendar_result.duplicate_existing_enrollware_sessions_deduped}",
            f"- Hot-sync active file exists: {calendar_result.hot_sync_active_path_exists}",
            f"- Hot-sync active entries found: {calendar_result.hot_sync_active_entries_found}",
            f"- Hot-sync entries absorbed by Class Report: {calendar_result.hot_sync_entries_absorbed}",
            f"- Hot-sync entries kept active: {calendar_result.hot_sync_entries_kept_active}",
            f"- Hot-sync entries flagged for review: {calendar_result.hot_sync_entries_flagged_for_review}",
            f"- Hot-sync entries merged into blocking source: {calendar_result.hot_sync_entries_merged}",
            f"- Existing Enrollware sessions assigned to Brian: {len(calendar_result.enrollware_brian_blocks)}",
            f"- Existing Enrollware sessions ignored because instructor was not Brian: {calendar_result.enrollware_non_brian_ignored}",
            "",
            "## Private URL Handling",
            "- Private DoNotSchedule URL: [REDACTED]",
            "- Redaction confirmed: true",
        ]
    )
    RECHECK_MD_PATH.write_text("\n".join(status_lines) + "\n", encoding="utf-8")


def recheck_proposed_session(page_slug: str) -> int:
    config = load_config()
    sessions = load_proposed_sessions()
    session = next((item for item in sessions if item.get("page_slug") == page_slug), None)
    if not session:
        print("RECHECK FAIL:")
        print(f"Reason: No proposed session found for page_slug '{page_slug}'.")
        return 1

    calendar_result = load_current_calendar_blocks(config)
    if calendar_result.warnings:
        passed = False
        reason = "Calendar fetch warning during recheck: " + " ".join(calendar_result.warnings)
    else:
        course = course_from_proposed_session(session)
        start = datetime.fromisoformat(session["start_time"])
        accepted, rejection = evaluate_candidate_session(course, start, calendar_result.blocks, config)
        passed = accepted is not None
        reason = None if passed else (rejection or {}).get("reason", "Rejected for an unknown scheduling reason.")
    evaluated_at = datetime.now(ZoneInfo(config["timezone"])).isoformat()
    write_recheck_report(session, calendar_result, passed, reason, evaluated_at)

    if passed:
        print("RECHECK PASS:")
        print("The proposed session is still available.")
        print("No current ADR or DoNotSchedule conflict was found.")
        print("Travel and cleanup buffers are satisfied.")
        return 0

    print("RECHECK FAIL:")
    print("The proposed session is no longer available.")
    print(f"Reason: {reason}")
    return 2


EXPANDED_OPEN_DAY_JSON_PATH = DEBUG_DIR / "expanded_open_day_offer_report.json"
EXPANDED_OPEN_DAY_MD_PATH = DEBUG_DIR / "expanded_open_day_offer_report.md"


def block_intersects_day(block: CalendarEvent, day: date) -> bool:
    day_start = datetime.combine(day, time.min, tzinfo=ZoneInfo("America/New_York"))
    return block.start < day_start + timedelta(days=1) and block.end > day_start


def open_day_rejection_reasons(day: date, blocks: list[CalendarEvent]) -> list[str]:
    reasons: list[str] = []
    for block in blocks:
        if not block_intersects_day(block, day):
            continue
        source = str(block.source or "block")
        summary = str(block.summary or "busy")
        if source == "adr":
            reasons.append(f"ADR block: {summary}")
        elif source == "do_not_schedule":
            reasons.append(f"DoNotSchedule block: {summary}")
        elif source == "enrollware_existing_session":
            reasons.append(f"real canonical Enrollware session: {summary}")
        elif source == "hot_sync_delta":
            reasons.append(f"hot-sync active session: {summary}")
        else:
            reasons.append(f"{source} block: {summary}")
    return sorted(set(reasons))


def expanded_course_templates(config: dict[str, Any]) -> list[dict[str, Any]]:
    wanted = {"bls", "bls-renewal", "heartcode-bls-skills"}
    return [course for course in config.get("course_templates", []) if course.get("course_key") in wanted]


def generate_expanded_open_day_report() -> int:
    config = load_config()
    tz = ZoneInfo(config["timezone"])
    strategy = config.get("expanded_offer_strategy", {})
    start_day = date.fromisoformat(str(strategy.get("start_date", "2026-08-01")))
    max_days = int(strategy.get("max_days_shown", 30) or 30)
    max_per_day = int(strategy.get("max_requestable_offers_per_day_per_course", 8) or 8)
    anchor_times = [str(item) for item in strategy.get("anchor_times", [])]
    calendar_result = load_current_calendar_blocks(config)
    report: dict[str, Any] = {
        "generated_at": datetime.now(tz).isoformat(),
        "enabled_publicly": bool(strategy.get("enabled", False)),
        "rendered_on_hubs": False,
        "start_date": start_day.isoformat(),
        "max_days_shown": max_days,
        "calendar_counts": {
            "adr_events_found": len(calendar_result.adr_events),
            "brian_adr_blocks_found": len(calendar_result.brian_adr_blocks),
            "do_not_schedule_blocks_found": len(calendar_result.dns_blocks),
            "merged_existing_enrollware_sessions": calendar_result.merged_existing_enrollware_sessions,
            "hot_sync_entries_merged": calendar_result.hot_sync_entries_merged,
        },
        "by_course_key": {},
        "cross_course_portability_audit": {
            "canonical_class_report_inventory_hubs": [
                "bls",
                "acls",
                "pals",
                "heartsaver",
                "arc",
                "hsi",
                "uscg-elementary-first-aid-cpr",
                "group-training",
            ],
            "requestable_offer_hubs_currently_rendering": ["bls"],
            "course_keys_with_offer_frequency_group": sorted((config.get("offer_strategy", {}).get("course_offer_overrides", {}) or {}).keys()),
            "course_keys_with_enrollware_display_name_mapping": sorted((config.get("course_display_names", {}) or {}).keys()),
            "normalization_alias_scope": "BLS aliases are implemented in free_time_scheduler.py for bls, bls-renewal, and heartcode-bls-skills.",
            "hub_display_strategy_support": "Implemented for BLS tabs only in this pass.",
            "not_yet_safely_wired": "Other hubs use canonical real inventory, but public requestable offer rendering and display-priority tuning are not generalized beyond BLS.",
            "recommended_next_generalization": "ACLS/PALS, because their higher-value classes benefit most from seated-real prioritization and future instructor-priority rules.",
        },
        "amy_parallel_instructor_design": {
            "status": "design_only_not_implemented",
            "instructor_profiles": "Add instructor-specific profiles with name, aliases, home/base location, and eligible calendars.",
            "allowed_courses_by_instructor": "Define which course_keys Amy can teach independently from Brian.",
            "availability_source_by_instructor": "Read Amy availability from her own calendar or availability feed.",
            "preferred_anchor_rules": "Treat broad windows like 12:30-5:00 as ranked opportunity windows, not equal slots.",
            "course_priority_rules": "Prefer ACLS/PALS Renewal/Initial around 1:30 before lower-priority HeartCode skills anchors.",
            "fallback_suppression_rules": "Release lower-priority skills offers only after higher-priority options are no longer viable or intentionally suppressed.",
        },
    }

    for course in expanded_course_templates(config):
        course_key = str(course["course_key"])
        group_name, frequency_group, _reason = resolve_offer_frequency_group(course, config)
        day_rows: list[dict[str, Any]] = []
        offers: list[dict[str, Any]] = []
        rejected_anchor_count = 0
        day = start_day
        while len(day_rows) < max_days and (day - start_day).days <= 180:
            reasons = open_day_rejection_reasons(day, calendar_result.blocks)
            day_record = {"date": day.isoformat(), "is_truly_open": not reasons, "rejection_reasons": reasons}
            if not reasons:
                accepted_for_day = 0
                for anchor in anchor_times:
                    if accepted_for_day >= max_per_day:
                        break
                    hour, minute = [int(part) for part in anchor.split(":", 1)]
                    start_dt = datetime.combine(day, time(hour, minute), tzinfo=tz)
                    candidate, _rejection = evaluate_candidate_session(course, start_dt, calendar_result.blocks, config)
                    if not candidate:
                        rejected_anchor_count += 1
                        continue
                    gap_passed, _gap_reason = same_program_gap_status(course, start_dt, calendar_result.blocks, [], config, frequency_group)
                    if not gap_passed:
                        rejected_anchor_count += 1
                        continue
                    offers.append({
                        "course_key": course_key,
                        "course_display_name": course_display_name(course_key, config) or course.get("course_title"),
                        "start_time": candidate["start_time"],
                        "end_time": candidate["end_time"],
                        "location_name": candidate["location_name"],
                        "offer_slug": slugify_offer(course_key, candidate["location_name"], start_dt),
                        "offer_frequency_group": group_name,
                        "report_only": True,
                    })
                    accepted_for_day += 1
                day_record["expanded_anchor_offers_generated"] = accepted_for_day
            day_rows.append(day_record)
            day += timedelta(days=1)

        reason_counter = Counter(reason.split(":", 1)[0] for row in day_rows for reason in row["rejection_reasons"])
        report["by_course_key"][course_key] = {
            "truly_open_days_found_after_start_date": sum(1 for row in day_rows if row["is_truly_open"]),
            "days_rejected_as_not_truly_open": sum(1 for row in day_rows if not row["is_truly_open"]),
            "rejection_reasons_by_day": [row for row in day_rows if not row["is_truly_open"]],
            "top_rejection_reasons": dict(reason_counter.most_common(10)),
            "expanded_anchor_offers_generated": len(offers),
            "anchor_offers_rejected": rejected_anchor_count,
            "first_10_truly_open_days": [row["date"] for row in day_rows if row["is_truly_open"]][:10],
            "first_20_expanded_offers": offers[:20],
        }

    DEBUG_DIR.mkdir(parents=True, exist_ok=True)
    EXPANDED_OPEN_DAY_JSON_PATH.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    lines = [
        "# Expanded Open-Day Offer Report",
        "",
        f"- Generated at: {report['generated_at']}",
        f"- Publicly enabled: {report['enabled_publicly']}",
        f"- Rendered on hubs: {report['rendered_on_hubs']}",
        f"- Start date: {report['start_date']}",
        "",
        "## Course Summary",
    ]
    for course_key, item in report["by_course_key"].items():
        lines.extend([
            "",
            f"### {course_key}",
            f"- Truly open days found: {item['truly_open_days_found_after_start_date']}",
            f"- Days rejected as not truly open: {item['days_rejected_as_not_truly_open']}",
            f"- Expanded anchor offers generated: {item['expanded_anchor_offers_generated']}",
            f"- Anchor offers rejected: {item['anchor_offers_rejected']}",
            f"- First 10 open days: {', '.join(item['first_10_truly_open_days']) or 'None'}",
            f"- Top rejection reasons: {json.dumps(item['top_rejection_reasons'], ensure_ascii=False)}",
        ])
    lines.extend([
        "",
        "## Amy / Parallel Instructor Design",
        "",
        "Design only; not implemented. Amy should get an instructor profile, allowed course list, independent availability source, ranked anchor rules, course priority rules, and fallback suppression rules so broad availability windows do not produce equally ranked offers for every eligible class.",
    ])
    EXPANDED_OPEN_DAY_MD_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {EXPANDED_OPEN_DAY_JSON_PATH}")
    print(f"Wrote {EXPANDED_OPEN_DAY_MD_PATH}")
    return 0


def appointment_link_case(label: str, func: Any) -> dict[str, Any]:
    try:
        return {"label": label, "passed": True, "value": func(), "error": None}
    except Exception as exc:
        return {"label": label, "passed": False, "value": None, "error": str(exc)}


def generate_appointment_link_test_report() -> int:
    config = load_config()
    day_cases = [
        ("Brian 2026-06-21", "Brian", "2026-06-21", 260670),
        ("Brian 2026-08-04", "Brian", "2026-08-04", 260714),
        ("Brian 2026-08-14", "Brian", "2026-08-14", 260724),
        ("Amy 2026-06-06", "Amy", "2026-06-06", 261393),
        ("Amy 2026-06-07", "Amy", "2026-06-07", 261394),
        ("Amy 2026-08-14", "Amy", "2026-08-14", 261462),
        ("Nick 2026-06-06", "Nick", "2026-06-06", 261013),
        ("Nick 2026-06-07", "Nick", "2026-06-07", 261014),
        ("Nick 2026-08-14", "Nick", "2026-08-14", 261082),
    ]
    day_results = []
    for label, instructor, requested_date, expected in day_cases:
        result = appointment_link_case(label, lambda instructor=instructor, requested_date=requested_date: calculate_appointment_day_id(instructor, requested_date, config))
        result["expected"] = expected
        result["passed"] = result["passed"] and result["value"] == expected
        day_results.append(result)

    url_results = []
    for label, course_key, course_id in [
        ("Brian 2026-08-14 12:45 PM bls", "bls", 209806),
        ("Brian 2026-08-14 12:45 PM bls-renewal", "bls-renewal", 359474),
        ("Brian 2026-08-14 12:45 PM heartcode-bls-skills", "heartcode-bls-skills", 210549),
    ]:
        expected_url = f"https://coastalcprtraining.enrollware.com/enroll?appointmentDayId=260724&startTime=12:45%20PM&courseId={course_id}"
        result = appointment_link_case(
            label,
            lambda course_key=course_key: build_enrollware_appointment_url("Brian", course_key, "2026-08-14T12:45:00-04:00", config, explicit_test_mode=True),
        )
        result["expected"] = expected_url
        result["passed"] = result["passed"] and result["value"] == expected_url
        url_results.append(result)

    fail_cases = [
        ("date before configured start date", lambda: calculate_appointment_day_id("Brian", "2026-06-20", config)),
        ("date after valid_through_date", lambda: calculate_appointment_day_id("Brian", "2027-06-21", config)),
        ("unknown instructor", lambda: calculate_appointment_day_id("Unknown", "2026-08-14", config)),
        ("missing course ID", lambda: build_enrollware_appointment_url("Brian", "heartsaver-fa-cpr-aed", "2026-08-14T12:45:00-04:00", config, explicit_test_mode=True)),
        ("invalid requested_start", lambda: build_enrollware_appointment_url("Brian", "bls", "not-a-date", config, explicit_test_mode=True)),
        ("disabled outside explicit test mode", lambda: build_enrollware_appointment_url("Brian", "bls", "2026-08-14T12:45:00-04:00", config)),
    ]
    fail_results = []
    for label, func in fail_cases:
        result = appointment_link_case(label, func)
        result["passed"] = not result["passed"]
        fail_results.append(result)

    configured_course_ids = appointment_link_config(config).get("course_ids", {})
    missing_course_ids = sorted({"heartsaver-fa-cpr-aed"} - set(configured_course_ids))
    report = {
        "generated_at": datetime.now(ZoneInfo(config["timezone"])).isoformat(),
        "appointment_links_enabled": bool(appointment_link_config(config).get("enabled")),
        "fail_closed": bool(appointment_link_config(config).get("fail_closed")),
        "day_id_formula": "appointmentDayId = appointment_start_day_id + real calendar day offset from appointment_start_date",
        "no_sparse_availability_day_counting": True,
        "no_guessed_ids_used": True,
        "day_id_tests": day_results,
        "url_tests": url_results,
        "fail_closed_tests": fail_results,
        "missing_course_ids": missing_course_ids,
        "configured_course_ids": configured_course_ids,
    }
    DEBUG_DIR.mkdir(parents=True, exist_ok=True)
    APPOINTMENT_LINK_REPORT_JSON_PATH.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    lines = [
        "# Enrollware Appointment Link Test Report",
        "",
        f"- Generated at: {report['generated_at']}",
        f"- Appointment links enabled: {report['appointment_links_enabled']}",
        f"- Fail closed: {report['fail_closed']}",
        f"- Formula: {report['day_id_formula']}",
        f"- No guessed IDs used: {report['no_guessed_ids_used']}",
        "",
        "## Day ID Tests",
    ]
    for item in day_results:
        lines.append(f"- {item['label']}: expected {item['expected']}, got {item['value']} ({'PASS' if item['passed'] else 'FAIL'})")
    lines.extend(["", "## URL Tests"])
    for item in report["url_tests"]:
        lines.append(f"- {item['label']}: {item['value']} ({'PASS' if item['passed'] else 'FAIL'})")
    lines.extend(["", "## Fail-Closed Tests"])
    for item in fail_results:
        lines.append(f"- {item['label']}: {'PASS' if item['passed'] else 'FAIL'} ({item['error']})")
    lines.extend(["", f"Missing course IDs: {', '.join(missing_course_ids) or 'None'}"])
    APPOINTMENT_LINK_REPORT_MD_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {APPOINTMENT_LINK_REPORT_JSON_PATH}")
    print(f"Wrote {APPOINTMENT_LINK_REPORT_MD_PATH}")
    print(f"Generated Brian BLS URL: {url_results[0]['value']}")
    return 0 if all(item["passed"] for item in [*day_results, *url_results, *fail_results]) else 2


def iter_customer_facing_offers() -> list[dict[str, Any]]:
    if not OFFERS_PATH.exists():
        return []
    payload = json.loads(OFFERS_PATH.read_text(encoding="utf-8"))
    rows: list[dict[str, Any]] = []
    for course in payload.get("courses", []):
        if not isinstance(course, dict):
            continue
        for offer in course.get("offered_options", []):
            if not isinstance(offer, dict):
                continue
            row = dict(offer)
            row["course_key"] = row.get("course_key") or course.get("course_key")
            row["course_display_name"] = row.get("course_display_name") or course.get("course_display_name") or course.get("course_title")
            rows.append(row)
    return rows


def requestable_offer_appointment_record(offer: dict[str, Any], config: dict[str, Any], calendar_result: CalendarLoadResult) -> dict[str, Any]:
    instructor = str(offer.get("instructor") or config.get("primary_instructor", {}).get("name") or "Brian")
    course_key = str(offer.get("course_key") or "")
    requested_start = str(offer.get("start_time") or "")
    course_ids = appointment_link_config(config).get("course_ids", {})
    record = {
        "offer_slug": offer.get("offer_slug") or offer.get("page_slug"),
        "course_key": course_key,
        "course_display_name": offer.get("course_display_name") or offer.get("course_title"),
        "instructor": instructor,
        "requested_start": requested_start,
        "appointmentDayId": None,
        "courseId": course_ids.get(course_key),
        "appointment_url": None,
        "valid": False,
        "suppressed_by_build_time_data": False,
        "build_time_suppression_reason": None,
        "fallback_url": config.get("requestable_appointment_url_mode", {}).get("fallback_url"),
        "reason": None,
    }
    try:
        requested = parse_requested_start(requested_start, config)
        record["appointmentDayId"] = calculate_appointment_day_id(instructor, requested.date(), config)
        course = course_from_customer_offer(offer, config)
        accepted, rejection = evaluate_candidate_session(course, requested, calendar_result.blocks, config)
        if not accepted:
            record["suppressed_by_build_time_data"] = True
            record["build_time_suppression_reason"] = (rejection or {}).get("reason", "failed build-time availability validation")
            raise ValueError(record["build_time_suppression_reason"])
        record["appointment_url"] = build_enrollware_appointment_url(
            instructor,
            course_key,
            requested_start,
            config,
            explicit_test_mode=True,
        )
        record["valid"] = True
    except Exception as exc:
        record["reason"] = str(exc)
    return record


def generate_requestable_appointment_url_report() -> int:
    config = load_config()
    mode = config.get("requestable_appointment_url_mode", {})
    calendar_result = load_current_calendar_blocks(config)
    rows = [requestable_offer_appointment_record(offer, config, calendar_result) for offer in iter_customer_facing_offers()]
    by_course: dict[str, dict[str, int]] = {}
    for row in rows:
        course_key = str(row.get("course_key") or "unknown")
        summary = by_course.setdefault(course_key, {"valid": 0, "failed_closed": 0})
        if row.get("valid"):
            summary["valid"] += 1
        else:
            summary["failed_closed"] += 1
    report = {
        "generated_at": datetime.now(ZoneInfo(config["timezone"])).isoformat(),
        "mode_enabled": bool(mode.get("enabled")),
        "report_only": bool(mode.get("report_only", True)),
        "only_confirmed_course_ids": bool(mode.get("only_confirmed_course_ids", True)),
        "public_hrefs_changed": False,
        "no_guessed_ids_used": True,
        "configured_course_ids": appointment_link_config(config).get("course_ids", {}),
        "missing_course_ids": sorted({row["course_key"] for row in rows if str(row.get("reason") or "").startswith("course ID is missing")}),
        "offers_suppressed_by_build_time_availability_data": sum(1 for row in rows if row.get("suppressed_by_build_time_data")),
        "offers_with_exact_enrollware_appointment_links": sum(1 for row in rows if row.get("valid")),
        "offers_lacking_valid_registration_url": sum(1 for row in rows if not row.get("valid")),
        "summary_by_course_key": by_course,
        "offers": rows,
    }
    DEBUG_DIR.mkdir(parents=True, exist_ok=True)
    REQUESTABLE_APPOINTMENT_URL_REPORT_JSON_PATH.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    lines = [
        "# Requestable Appointment URL Report",
        "",
        f"- Generated at: {report['generated_at']}",
        f"- Mode enabled: {report['mode_enabled']}",
        f"- Report only: {report['report_only']}",
        f"- Public hrefs changed: {report['public_hrefs_changed']}",
        f"- No guessed IDs used: {report['no_guessed_ids_used']}",
        f"- Offers suppressed by build-time availability data: {report['offers_suppressed_by_build_time_availability_data']}",
        f"- Offers with exact Enrollware appointment links: {report['offers_with_exact_enrollware_appointment_links']}",
        f"- Offers lacking valid registration URL: {report['offers_lacking_valid_registration_url']}",
        "",
        "## Summary By Course",
    ]
    for course_key, summary in sorted(by_course.items()):
        lines.append(f"- {course_key}: valid {summary['valid']}, failed closed {summary['failed_closed']}")
    lines.extend(["", "## First Valid URLs"])
    for row in [item for item in rows if item["valid"]][:10]:
        lines.append(f"- {row['offer_slug']}: {row['appointment_url']}")
    lines.extend(["", "## First Fail-Closed Records"])
    for row in [item for item in rows if not item["valid"]][:10]:
        lines.append(f"- {row['offer_slug']} ({row['course_key']}): {row['reason']}")
    REQUESTABLE_APPOINTMENT_URL_REPORT_MD_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {REQUESTABLE_APPOINTMENT_URL_REPORT_JSON_PATH}")
    print(f"Wrote {REQUESTABLE_APPOINTMENT_URL_REPORT_MD_PATH}")
    print(f"Valid requestable appointment URLs: {sum(1 for row in rows if row['valid'])}")
    print(f"Failed closed requestable offers: {sum(1 for row in rows if not row['valid'])}")
    return 0


def generate_scheduler_outputs() -> int:
    config = load_config()
    canonical_payload = write_canonical_schedule_from_class_report(config)
    calendar_result = load_current_calendar_blocks(config)

    accepted, rejected = generate_candidates(config, calendar_result.blocks)

    PROPOSED_SESSIONS_PATH.parent.mkdir(parents=True, exist_ok=True)
    PROPOSED_SESSIONS_PATH.write_text(json.dumps(accepted, indent=2), encoding="utf-8")
    generate_preview_pages(accepted)
    generate_reports(
        config,
        calendar_result.adr_events,
        calendar_result.brian_adr_blocks,
        calendar_result.dns_blocks,
        accepted,
        rejected,
        calendar_result.warnings,
        calendar_result.do_not_schedule_source,
        calendar_result.do_not_schedule_private,
        calendar_result.do_not_schedule_loaded,
        calendar_result.enrollware_existing_sessions_found,
        calendar_result.class_report_sessions_found,
        calendar_result.schedule_future_sessions_found,
        calendar_result.merged_existing_enrollware_sessions,
        calendar_result.existing_enrollware_source_mode,
        calendar_result.duplicate_existing_enrollware_sessions_deduped,
        calendar_result.existing_enrollware_source_mismatches,
        calendar_result.hot_sync_active_entries_found,
        calendar_result.hot_sync_entries_absorbed,
        calendar_result.hot_sync_entries_kept_active,
        calendar_result.hot_sync_entries_flagged_for_review,
        calendar_result.hot_sync_entries_merged,
        calendar_result.hot_sync_active_path_exists,
        calendar_result.class_report_found,
        calendar_result.schedule_future_fallback_used,
        calendar_result.enrollware_brian_blocks,
        calendar_result.enrollware_non_brian_ignored,
    )
    print(f"Generated {len(accepted)} proposed sessions and {len(rejected)} rejected candidates.")
    print(f"Canonical schedule source mode: {canonical_payload['build']['source_mode']}")
    print(f"Canonical schedule sessions: {canonical_payload['build']['session_count']}")
    if not calendar_result.do_not_schedule_loaded:
        print("DoNotSchedule calendar did not load successfully.")
    else:
        print(f"DoNotSchedule calendar loaded successfully from {calendar_result.do_not_schedule_source}.")
    print(f"DoNotSchedule blocks found: {len(calendar_result.dns_blocks)}")
    print(f"Existing Enrollware sessions found: {calendar_result.enrollware_existing_sessions_found}")
    print(f"Class Report XLSX found: {calendar_result.class_report_found}")
    print(f"Class Report sessions found: {calendar_result.class_report_sessions_found}")
    print(f"schedule_future.json fallback used: {calendar_result.schedule_future_fallback_used}")
    print(f"schedule_future sessions found: {calendar_result.schedule_future_sessions_found}")
    print(f"Merged existing Enrollware sessions: {calendar_result.merged_existing_enrollware_sessions}")
    print(f"Existing Enrollware source mode: {calendar_result.existing_enrollware_source_mode}")
    print(f"Duplicate existing Enrollware sessions deduped: {calendar_result.duplicate_existing_enrollware_sessions_deduped}")
    print(f"Hot-sync active file exists: {calendar_result.hot_sync_active_path_exists}")
    print(f"Hot-sync active entries found: {calendar_result.hot_sync_active_entries_found}")
    print(f"Hot-sync entries absorbed by Class Report: {calendar_result.hot_sync_entries_absorbed}")
    print(f"Hot-sync entries kept active: {calendar_result.hot_sync_entries_kept_active}")
    print(f"Hot-sync entries flagged for review: {calendar_result.hot_sync_entries_flagged_for_review}")
    print(f"Hot-sync entries merged into blocking source: {calendar_result.hot_sync_entries_merged}")
    print(f"Existing Enrollware sessions assigned to Brian: {len(calendar_result.enrollware_brian_blocks)}")
    print(f"Existing Enrollware sessions ignored because instructor was not Brian: {calendar_result.enrollware_non_brian_ignored}")
    print(
        "Proposed sessions rejected because of Brian Enrollware sessions: "
        f"{sum(1 for item in rejected if item.get('rejection_source') == 'enrollware_existing_session')}"
    )
    if calendar_result.warnings:
        print("Warnings:")
        for warning in calendar_result.warnings:
            print(f"- {warning}")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate and validate 910CPR free-time scheduler previews.")
    parser.add_argument("--recheck-slug", help="Re-fetch calendars and re-evaluate one proposed session by page_slug.")
    parser.add_argument("--list-proposed", action="store_true", help="List proposed sessions from docs/data/proposed_sessions.json.")
    parser.add_argument("--generate-offers", action="store_true", help="Generate curated customer-facing request-time offers.")
    parser.add_argument("--generate-expanded-open-days", action="store_true", help="Generate report-only expanded open-day offer debug output.")
    parser.add_argument("--test-appointment-links", action="store_true", help="Generate deterministic Enrollware appointment link test report.")
    parser.add_argument("--report-requestable-appointment-urls", action="store_true", help="Report which requestable offers can receive deterministic appointment URLs.")
    parser.add_argument("--course-key", default="", help="Optional course_key filter for --generate-offers.")
    parser.add_argument("--date", default="", help="Optional YYYY-MM-DD date filter for --generate-offers.")
    parser.add_argument("--limit", type=int, default=None, help="Maximum proposed sessions to list, or maximum offers per course with --generate-offers.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.list_proposed:
        return list_proposed_sessions(max(args.limit if args.limit is not None else 10, 0))
    if args.recheck_slug:
        return recheck_proposed_session(args.recheck_slug)
    if args.generate_offers:
        return generate_customer_facing_offers(
            course_key_filter=args.course_key.strip(),
            date_filter=args.date.strip(),
            limit=max(args.limit, 0) if args.limit is not None else None,
        )
    if args.generate_expanded_open_days:
        return generate_expanded_open_day_report()
    if args.test_appointment_links:
        return generate_appointment_link_test_report()
    if args.report_requestable_appointment_urls:
        return generate_requestable_appointment_url_report()
    return generate_scheduler_outputs()


if __name__ == "__main__":
    raise SystemExit(main())
