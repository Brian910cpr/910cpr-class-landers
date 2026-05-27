from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ID_RE = re.compile(r"\d+")
LINK_ID_RE = re.compile(r"[?&]id=(\d+)")
CLASS_LINK_RE = re.compile(r"/classes/(\d+)\.html")


@dataclass
class ScanResult:
    report_ids: set[str]
    report_path: Path
    json_ids: set[str]
    schedule_path: Path
    future_ids: set[str]
    schedule_future_path: Path
    class_file_ids: set[str]
    class_file_map: dict[str, list[Path]]
    hub_html_ref_ids: set[str]
    stale_hub_files: dict[str, list[str]]
    stale_hub_file_counts: dict[str, int]
    stale_json_ids: list[str]
    stale_html_ids: list[str]
    retained_past_page_ids: list[str]
    stale_hub_ref_ids: list[str]


def normalize_id(value: Any) -> str:
    digits = "".join(ID_RE.findall(str(value or "")))
    if not digits:
        return ""
    return digits


def resolve_class_report_path(repo_root: Path, requested: str) -> Path:
    requested_path = (repo_root / requested).resolve()
    if requested_path.exists():
        return requested_path
    candidates = [
        repo_root / "data" / "Class Report.xlsx",
        repo_root / "data" / "raw" / "Class Report.xlsx",
        repo_root / "data" / "raw" / "class_report.xlsx",
    ]
    for path in candidates:
        if path.exists():
            return path.resolve()
    return requested_path


def read_class_report_ids(class_report_path: Path) -> set[str]:
    wb = load_workbook(filename=class_report_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return set()
    headers = [str(h or "").strip() for h in rows[0]]
    idx_by_name = {name: i for i, name in enumerate(headers)}
    reg_idx = idx_by_name.get("Registration Link")
    id_idx = idx_by_name.get("ID")
    out: set[str] = set()

    for row in rows[1:]:
        reg_val = row[reg_idx] if reg_idx is not None and reg_idx < len(row) else None
        sid = ""
        if reg_val:
            m = LINK_ID_RE.search(str(reg_val))
            if m:
                sid = m.group(1)
        if not sid and id_idx is not None and id_idx < len(row):
            sid = str(row[id_idx] or "").strip()
        norm = normalize_id(sid)
        if norm:
            out.add(norm)
    return out


def read_schedule_ids(path: Path) -> set[str]:
    if not path.exists():
        return set()
    raw = json.loads(path.read_text(encoding="utf-8"))
    sessions = raw.get("sessions", []) if isinstance(raw, dict) else []
    out: set[str] = set()
    for session in sessions:
        sid = ""
        if isinstance(session, dict):
            sid = str(session.get("session_id") or session.get("id") or "").strip()
        norm = normalize_id(sid)
        if norm:
            out.add(norm)
    return out


def scan_class_html_ids(classes_dir: Path) -> tuple[set[str], dict[str, list[Path]]]:
    out: set[str] = set()
    file_map: dict[str, list[Path]] = {}
    if not classes_dir.exists():
        return out, file_map
    for path in classes_dir.glob("*.html"):
        norm = normalize_id(path.stem)
        if not norm:
            continue
        out.add(norm)
        file_map.setdefault(norm, []).append(path)
    return out, file_map


def is_public_hub_file(path: Path, docs_dir: Path) -> bool:
    try:
        rel = path.relative_to(docs_dir)
    except Exception:
        return False
    parts = rel.parts
    if not parts:
        return False
    # Never treat class detail trees or internal assets/data as hub inputs.
    blocked_roots = {"classes", "data", "assets", "images", "css", "js"}
    if parts[0].lower() in blocked_roots:
        return False
    # Public hub pages live at root and these generated folders.
    if len(parts) == 1:
        return True
    return parts[0].lower() in {"courses", "locations", "course-at-city"}


def scan_hub_html_refs(docs_dir: Path, valid_ids: set[str]) -> tuple[set[str], dict[str, list[str]], dict[str, int]]:
    refs: set[str] = set()
    stale_file_ids: dict[str, list[str]] = {}
    stale_file_counts: dict[str, int] = {}
    if not docs_dir.exists():
        return refs, stale_file_ids, stale_file_counts

    for path in docs_dir.rglob("*.html"):
        if not is_public_hub_file(path, docs_dir):
            continue
        text = path.read_text(encoding="utf-8", errors="ignore")
        stale_hits: list[str] = []
        for m in CLASS_LINK_RE.finditer(text):
            norm = normalize_id(m.group(1))
            if norm:
                refs.add(norm)
                if norm not in valid_ids:
                    stale_hits.append(norm)
        for m in LINK_ID_RE.finditer(text):
            norm = normalize_id(m.group(1))
            if norm:
                refs.add(norm)
                if norm not in valid_ids:
                    stale_hits.append(norm)
        if stale_hits:
            rel = str(path.relative_to(docs_dir)).replace("\\", "/")
            deduped = sorted(set(stale_hits))
            stale_file_ids[rel] = deduped
            stale_file_counts[rel] = len(stale_hits)
    return refs, stale_file_ids, stale_file_counts


def scan_sitemap_refs(docs_dir: Path, valid_ids: set[str]) -> tuple[set[str], dict[str, list[str]], dict[str, int]]:
    sitemap_path = docs_dir / "sitemap.xml"
    refs: set[str] = set()
    stale_file_ids: dict[str, list[str]] = {}
    stale_file_counts: dict[str, int] = {}
    if not sitemap_path.exists():
        return refs, stale_file_ids, stale_file_counts

    text = sitemap_path.read_text(encoding="utf-8", errors="ignore")
    stale_hits: list[str] = []
    for m in CLASS_LINK_RE.finditer(text):
        norm = normalize_id(m.group(1))
        if norm:
            refs.add(norm)
            if norm not in valid_ids:
                stale_hits.append(norm)
    if stale_hits:
        rel = str(sitemap_path.relative_to(docs_dir)).replace("\\", "/")
        stale_file_ids[rel] = sorted(set(stale_hits))
        stale_file_counts[rel] = len(stale_hits)
    return refs, stale_file_ids, stale_file_counts


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def file_mtime(path: Path) -> str | None:
    if not path.exists():
        return None
    return datetime.fromtimestamp(path.stat().st_mtime).isoformat()


def scan(repo_root: Path, class_report_path: Path) -> ScanResult:
    docs_dir = repo_root / "docs"
    classes_dir = docs_dir / "classes"
    schedule_path = docs_dir / "data" / "schedule.json"
    schedule_future_path = docs_dir / "data" / "schedule_future.json"

    report_ids = read_class_report_ids(class_report_path)
    json_ids = read_schedule_ids(schedule_path)
    future_ids = read_schedule_ids(schedule_future_path)
    class_file_ids, class_file_map = scan_class_html_ids(classes_dir)
    hub_html_ref_ids, stale_hub_files, stale_hub_file_counts = scan_hub_html_refs(docs_dir, report_ids)
    sitemap_ref_ids, stale_sitemap_files, stale_sitemap_file_counts = scan_sitemap_refs(docs_dir, report_ids)
    hub_html_ref_ids |= sitemap_ref_ids
    stale_hub_files.update(stale_sitemap_files)
    stale_hub_file_counts.update(stale_sitemap_file_counts)

    combined_json_ids = json_ids | future_ids
    stale_json_ids = sorted(i for i in combined_json_ids if i not in report_ids)
    stale_html_ids = sorted(i for i in class_file_ids if i not in report_ids)
    retained_past_page_ids = sorted(i for i in class_file_ids if i in report_ids and i not in future_ids)
    stale_hub_ref_ids = sorted(i for i in hub_html_ref_ids if i not in report_ids)

    return ScanResult(
        report_ids=report_ids,
        report_path=class_report_path,
        json_ids=json_ids,
        schedule_path=schedule_path,
        future_ids=future_ids,
        schedule_future_path=schedule_future_path,
        class_file_ids=class_file_ids,
        class_file_map=class_file_map,
        hub_html_ref_ids=hub_html_ref_ids,
        stale_hub_files=stale_hub_files,
        stale_hub_file_counts=stale_hub_file_counts,
        stale_json_ids=stale_json_ids,
        stale_html_ids=stale_html_ids,
        retained_past_page_ids=retained_past_page_ids,
        stale_hub_ref_ids=stale_hub_ref_ids,
    )


def quarantine_stale_class_files(repo_root: Path, stale_ids: list[str], file_map: dict[str, list[Path]]) -> int:
    if not stale_ids:
        return 0
    target_root = repo_root / "debug" / "quarantine" / "stale_classes" / now_stamp()
    target_root.mkdir(parents=True, exist_ok=True)
    moved = 0
    for stale_id in stale_ids:
        for src in file_map.get(stale_id, []):
            dst = target_root / src.name
            shutil.move(str(src), str(dst))
            moved += 1
    return moved


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def latest_quarantined_stale_class_files(repo_root: Path, report_ids: set[str]) -> dict[str, Path]:
    quarantine_root = repo_root / "debug" / "quarantine" / "stale_classes"
    if not quarantine_root.exists():
        return {}
    dirs = [path for path in quarantine_root.iterdir() if path.is_dir()]
    if not dirs:
        return {}
    latest = max(dirs, key=lambda path: path.stat().st_mtime)
    out: dict[str, Path] = {}
    for path in latest.glob("*.html"):
        session_id = normalize_id(path.stem)
        if session_id and session_id not in report_ids:
            out[session_id] = path
    return out


def write_stale_summary_csv(repo_root: Path, result: ScanResult) -> Path:
    path = repo_root / "debug" / "stale_sessions_summary.csv"
    path.parent.mkdir(parents=True, exist_ok=True)
    referenced_by: dict[str, list[str]] = {}
    for rel_path, ids in result.stale_hub_files.items():
        for session_id in ids:
            referenced_by.setdefault(session_id, []).append(rel_path)

    quarantined_files = latest_quarantined_stale_class_files(repo_root, result.report_ids)
    ids = sorted(
        set(result.stale_html_ids)
        | set(result.retained_past_page_ids)
        | set(result.stale_hub_ref_ids)
        | set(quarantined_files)
    )
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "session_id",
                "html_path",
                "reason_stale",
                "found_in_class_report",
                "found_in_schedule_future",
                "referenced_by_files",
            ],
        )
        writer.writeheader()
        for session_id in ids:
            in_report = session_id in result.report_ids
            in_future = session_id in result.future_ids
            refs = sorted(set(referenced_by.get(session_id, [])))
            html_paths = [
                str(path.relative_to(repo_root)).replace("\\", "/")
                for path in result.class_file_map.get(session_id, [])
            ]
            if session_id in quarantined_files:
                html_paths.append(str(quarantined_files[session_id].relative_to(repo_root)).replace("\\", "/"))
            reasons: list[str] = []
            if session_id in result.stale_html_ids:
                reasons.append("obsolete_html_not_in_class_report")
            if session_id in quarantined_files:
                reasons.append("obsolete_html_quarantined")
            if session_id in result.retained_past_page_ids:
                reasons.append("retained_past_or_non_public_page")
            if session_id in result.stale_hub_ref_ids:
                reasons.append("invalid_public_reference")
            writer.writerow(
                {
                    "session_id": session_id,
                    "html_path": ";".join(html_paths),
                    "reason_stale": ";".join(reasons),
                    "found_in_class_report": str(in_report).lower(),
                    "found_in_schedule_future": str(in_future).lower(),
                    "referenced_by_files": ";".join(refs),
                }
            )
    return path


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit and quarantine stale sessions not present in Class Report.xlsx")
    parser.add_argument("--repo-root", default=".")
    parser.add_argument("--class-report", default="data/Class Report.xlsx")
    parser.add_argument("--warn-only", action="store_true", help="Always return exit code 0")
    parser.add_argument("--cleanup", action="store_true", help="Move stale class pages into quarantine")
    args = parser.parse_args()

    repo_root = Path(args.repo_root).resolve()
    class_report_path = resolve_class_report_path(repo_root, args.class_report)
    if not class_report_path.exists():
        raise SystemExit(f"Class report not found: {class_report_path}")

    result = scan(repo_root, class_report_path)
    quarantine_count = 0
    post_cleanup: ScanResult | None = None

    if args.cleanup and result.stale_html_ids:
        quarantine_count = quarantine_stale_class_files(repo_root, result.stale_html_ids, result.class_file_map)
        post_cleanup = scan(repo_root, class_report_path)

    effective = post_cleanup or result
    audit_path = repo_root / "debug" / "stale_sessions_audit.json"
    health_path = repo_root / "debug" / "latest_build_health.json"
    summary_source = result if args.cleanup else effective
    summary_path = write_stale_summary_csv(repo_root, summary_source)

    payload = {
        "timestamp": datetime.now().isoformat(),
        "repo_path": str(repo_root),
        "class_report": {
            "path": str(class_report_path),
            "modified_at": file_mtime(class_report_path),
            "valid_session_id_count": len(result.report_ids),
        },
        "schedule": {
            "path": str(result.schedule_path),
            "modified_at": file_mtime(result.schedule_path),
            "session_count": len(result.json_ids),
        },
        "schedule_future": {
            "path": str(result.schedule_future_path),
            "modified_at": file_mtime(result.schedule_future_path),
            "session_count": len(result.future_ids),
        },
        "classes_html_count": len(effective.class_file_ids),
        "stale_json_ids": effective.stale_json_ids,
        "stale_html_ids": effective.stale_html_ids,
        "retained_past_page_ids": effective.retained_past_page_ids,
        "stale_hub_ref_ids": effective.stale_hub_ref_ids,
        "stale_hub_files": effective.stale_hub_files,
        "stale_hub_file_counts": effective.stale_hub_file_counts,
        "cleanup": {
            "requested": bool(args.cleanup),
            "quarantine_count": quarantine_count,
        },
        "rules": {
            "public_session_source": "docs/data/schedule_future.json",
            "stale_hub_ref_ids": "Hard failure: public hubs or sitemap reference session IDs absent from Class Report.xlsx.",
            "stale_html_ids": "Hard failure only when obsolete HTML files remain for session IDs absent from Class Report.xlsx; --cleanup quarantines them.",
            "retained_past_page_ids": "Class pages present in Class Report.xlsx but absent from schedule_future.json are retained past/non-public pages and are not failures unless publicly referenced.",
        },
        "summary_csv": str(summary_path),
    }
    write_json(audit_path, payload)
    write_json(health_path, payload)

    stale_found = bool(effective.stale_json_ids or effective.stale_html_ids or effective.stale_hub_ref_ids)
    print(f"Class Report IDs: {len(result.report_ids)}")
    print(f"schedule.json IDs: {len(result.json_ids)}")
    print(f"schedule_future.json IDs: {len(result.future_ids)}")
    print(f"docs/classes HTML files: {len(effective.class_file_ids)}")
    if args.cleanup:
        print(f"Quarantined stale class files: {quarantine_count}")
    print(f"stale_json_ids: {len(effective.stale_json_ids)}")
    print(f"stale_html_ids: {len(effective.stale_html_ids)}")
    print(f"retained_past_page_ids: {len(effective.retained_past_page_ids)}")
    print(f"stale_hub_ref_ids: {len(effective.stale_hub_ref_ids)}")
    print(f"Stale summary CSV: {summary_path}")
    print(f"Audit JSON: {audit_path}")
    print(f"Build health JSON: {health_path}")
    print("AUDIT RESULT: FAIL" if stale_found else "AUDIT RESULT: PASS")

    if stale_found and not args.warn_only:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
