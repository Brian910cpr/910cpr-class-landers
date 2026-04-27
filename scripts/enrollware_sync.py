from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, time
from html import unescape
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


TZ = ZoneInfo("America/New_York")
HEADER_ALIASES = {
    "id": "id",
    "course": "course",
    "course name": "course",
    "start date / time": "start",
    "start": "start",
    "end date / time": "end",
    "end": "end",
    "location": "location",
    "client": "client",
    "instructor": "instructor",
    "assistants": "assistants",
    "students": "students",
    "seats": "seats",
    "hours": "hours",
    "registration link": "registration_link",
    "registration url": "registration_link",
}
DEFAULT_RULES_PATH = Path("data/config/enrollware_sync_rules.example.json")
DEFAULT_OUTPUT_DIR = Path("data/runtime/enrollware_sync")
DEFAULT_MASTER_PATH = Path("data/Class Report.xlsx")
DEFAULT_EXPORT_PATH = Path("data/enrollware_export.xlsx")


@dataclass
class ValidationIssue:
    code: str
    severity: str
    message: str
    session_key: str | None = None
    related_session_keys: list[str] | None = None


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_space(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_string(value)).strip()


def strip_html(value: Any) -> str:
    text = clean_string(value)
    if not text:
        return ""
    text = re.sub(r"<br\s*/?>", " ", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    return normalize_space(unescape(text))


def normalize_header(value: Any) -> str:
    return normalize_space(value).lower()


def clean_location(value: Any) -> str:
    text = strip_html(value)
    if text.startswith("::"):
        text = text[2:].strip()
    if "::" in text:
        text = text.rsplit("::", 1)[-1].strip()
    return text


def canonical_name(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", normalize_space(value).lower()).strip()


def parse_int(value: Any) -> int | None:
    text = clean_string(value).replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except Exception:
        return None


def parse_float(value: Any) -> float | None:
    text = clean_string(value).replace(",", "").replace("$", "")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def parse_dt(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        dt = value
        if dt.tzinfo is None:
            return dt.replace(tzinfo=TZ)
        return dt.astimezone(TZ)
    if isinstance(value, date):
        return datetime.combine(value, time.min, tzinfo=TZ)

    text = clean_string(value)
    if not text:
        return None
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M",
        "%m/%d/%y %I:%M %p",
        "%m/%d/%y %H:%M",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=TZ)
        except Exception:
            continue
    try:
        dt = datetime.fromisoformat(text)
    except Exception:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=TZ)
    return dt.astimezone(TZ)


def dt_to_iso(value: datetime | None) -> str | None:
    return value.astimezone(TZ).isoformat() if value else None


def duration_hours(start_dt: datetime | None, end_dt: datetime | None, hours_value: Any) -> float | None:
    explicit = parse_float(hours_value)
    if explicit is not None:
        return explicit
    if start_dt and end_dt:
        return round((end_dt - start_dt).total_seconds() / 3600.0, 4)
    return None


def extract_course_meta(course_raw: Any) -> dict[str, Any]:
    raw = clean_string(course_raw)
    clean = strip_html(raw)
    lower = clean.lower()
    longdesc_match = re.search(r'longdesc=["\']([^"\']+)["\']', raw, flags=re.I)
    tokens: dict[str, str] = {}
    if longdesc_match:
        for token in longdesc_match.group(1).split("|"):
            if ":" not in token:
                continue
            key, value = token.split(":", 1)
            tokens[key.strip().lower()] = value.strip()

    course_number = tokens.get("r")
    course_code = tokens.get("t")
    body = tokens.get("cb")
    delivery = tokens.get("d")
    if not course_number:
        name_attr = re.search(r'\bname=["\'](\d+)["\']', raw, flags=re.I)
        if name_attr:
            course_number = name_attr.group(1)

    normalized_key = canonical_name(course_code or clean)
    if "heartcode" in lower or "skills session" in lower:
        delivery = delivery or "BL"
    elif "in-person" in lower or "in person" in lower or "classroom" in lower:
        delivery = delivery or "IP"

    return {
        "course_name_raw": raw,
        "course_name_clean": clean,
        "course_number": course_number,
        "course_code": course_code,
        "certifying_body": body,
        "delivery_mode": delivery,
        "normalized_course_key": normalized_key,
    }


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_tabular_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            return [{HEADER_ALIASES.get(normalize_header(k), normalize_header(k)): v for k, v in row.items()} for row in reader]

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [HEADER_ALIASES.get(normalize_header(value), normalize_header(value)) for value in rows[0]]
        output: list[dict[str, Any]] = []
        for row in rows[1:]:
            if row is None:
                continue
            record = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
            if any(value not in (None, "") for value in record.values()):
                output.append(record)
        return output

    raise ValueError(f"Unsupported input format: {path.suffix}")


def normalize_session_record(row: dict[str, Any], *, source_name: str, row_index: int) -> dict[str, Any]:
    meta = extract_course_meta(row.get("course"))
    start_dt = parse_dt(row.get("start"))
    end_dt = parse_dt(row.get("end"))
    location_clean = clean_location(row.get("location"))
    instructor = normalize_space(row.get("instructor"))
    assistants = normalize_space(row.get("assistants"))
    students = parse_int(row.get("students"))
    seats = parse_int(row.get("seats"))
    hours = duration_hours(start_dt, end_dt, row.get("hours"))
    source_id = clean_string(row.get("id"))
    registration_url = clean_string(row.get("registration_link"))

    fingerprint_parts = [
        meta.get("course_number") or meta.get("course_code") or meta.get("normalized_course_key"),
        dt_to_iso(start_dt) or "",
        dt_to_iso(end_dt) or "",
        canonical_name(location_clean),
    ]
    fingerprint = "|".join(fingerprint_parts)

    return {
        "source_name": source_name,
        "source_row_index": row_index,
        "source_id": source_id or None,
        "session_key": fingerprint,
        "course_name_clean": meta["course_name_clean"],
        "course_name_raw": meta["course_name_raw"],
        "course_number": meta["course_number"],
        "course_code": meta["course_code"],
        "certifying_body": meta["certifying_body"],
        "delivery_mode": meta["delivery_mode"],
        "normalized_course_key": meta["normalized_course_key"],
        "start_at": dt_to_iso(start_dt),
        "end_at": dt_to_iso(end_dt),
        "location": location_clean,
        "client": normalize_space(row.get("client")),
        "lead_instructor": instructor or None,
        "assistants": assistants or None,
        "students": students,
        "seats": seats,
        "hours": hours,
        "registration_url": registration_url or None,
    }


def filter_sessions_by_time(sessions: list[dict[str, Any]], *, include_past: bool) -> list[dict[str, Any]]:
    if include_past:
        return sessions
    now = datetime.now(TZ)
    return [session for session in sessions if (parse_dt(session.get("start_at")) or now) >= now]


def load_rules(path: Path | None) -> dict[str, Any]:
    if not path or not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def sessions_overlap(left: dict[str, Any], right: dict[str, Any]) -> bool:
    left_start = parse_dt(left.get("start_at"))
    left_end = parse_dt(left.get("end_at"))
    right_start = parse_dt(right.get("start_at"))
    right_end = parse_dt(right.get("end_at"))
    if not left_start or not left_end or not right_start or not right_end:
        return False
    return left_start < right_end and right_start < left_end


def validate_sessions(sessions: list[dict[str, Any]], rules: dict[str, Any]) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_location: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_instructor: dict[str, list[dict[str, Any]]] = defaultdict(list)
    closed_dates = {clean_string(item) for item in rules.get("closed_dates", []) if clean_string(item)}
    closed_weekdays = {int(item) for item in rules.get("closed_weekdays", []) if str(item).strip() != ""}
    duration_rules = rules.get("expected_duration_hours", {})
    instructor_unavailable = rules.get("instructor_unavailable", {})
    location_unavailable = rules.get("location_unavailable", {})

    for session in sessions:
        key = session["session_key"]
        by_key[key].append(session)

        start_dt = parse_dt(session.get("start_at"))
        end_dt = parse_dt(session.get("end_at"))
        location = canonical_name(session.get("location"))
        instructor = canonical_name(session.get("lead_instructor"))
        if location:
            by_location[location].append(session)
        if instructor:
            by_instructor[instructor].append(session)

        if not start_dt or not end_dt:
            issues.append(
                ValidationIssue(
                    code="missing_datetime",
                    severity="error",
                    message="Session is missing a start or end time.",
                    session_key=key,
                )
            )
            continue
        if end_dt <= start_dt:
            issues.append(
                ValidationIssue(
                    code="impossible_time_window",
                    severity="error",
                    message="Session end time must be after start time.",
                    session_key=key,
                )
            )

        if start_dt.date().isoformat() in closed_dates or start_dt.weekday() in closed_weekdays:
            issues.append(
                ValidationIssue(
                    code="closed_day",
                    severity="error",
                    message=f"Session falls on a closed day: {start_dt.date().isoformat()}",
                    session_key=key,
                )
            )

        hours = session.get("hours")
        duration_rule = duration_rules.get(session.get("course_code")) or duration_rules.get(session.get("course_number"))
        if duration_rule and hours is not None:
            min_hours = duration_rule.get("min")
            max_hours = duration_rule.get("max")
            if min_hours is not None and hours < float(min_hours):
                issues.append(
                    ValidationIssue(
                        code="course_duration_short",
                        severity="warning",
                        message=f"Session duration {hours}h is below expected minimum {min_hours}h.",
                        session_key=key,
                    )
                )
            if max_hours is not None and hours > float(max_hours):
                issues.append(
                    ValidationIssue(
                        code="course_duration_long",
                        severity="warning",
                        message=f"Session duration {hours}h is above expected maximum {max_hours}h.",
                        session_key=key,
                    )
                )

        for block in instructor_unavailable.get(session.get("lead_instructor") or "", []):
            block_start = parse_dt(block.get("start"))
            block_end = parse_dt(block.get("end"))
            if block_start and block_end and start_dt < block_end and block_start < end_dt:
                issues.append(
                    ValidationIssue(
                        code="instructor_unavailable",
                        severity="error",
                        message=f"Instructor unavailable during requested window: {session.get('lead_instructor')}",
                        session_key=key,
                    )
                )
        for block in location_unavailable.get(session.get("location") or "", []):
            block_start = parse_dt(block.get("start"))
            block_end = parse_dt(block.get("end"))
            if block_start and block_end and start_dt < block_end and block_start < end_dt:
                issues.append(
                    ValidationIssue(
                        code="location_unavailable",
                        severity="error",
                        message=f"Location unavailable during requested window: {session.get('location')}",
                        session_key=key,
                    )
                )

    for key, bucket in by_key.items():
        if len(bucket) > 1:
            issues.append(
                ValidationIssue(
                    code="duplicate_slot",
                    severity="error",
                    message=f"Duplicate desired slot detected {len(bucket)} times.",
                    session_key=key,
                    related_session_keys=[item["session_key"] for item in bucket],
                )
            )

    def scan_overlaps(bucket: list[dict[str, Any]], *, code: str, label_field: str) -> None:
        active: list[dict[str, Any]] = []
        ordered = sorted(bucket, key=lambda item: item.get("start_at") or "")
        for session in ordered:
            session_start = parse_dt(session.get("start_at"))
            session_end = parse_dt(session.get("end_at"))
            if not session_start or not session_end:
                continue
            active = [
                candidate
                for candidate in active
                if (parse_dt(candidate.get("end_at")) or session_start) > session_start
            ]
            for candidate in active:
                issues.append(
                    ValidationIssue(
                        code=code,
                        severity="error",
                        message=f"{label_field.replace('_', ' ').title()} overlap for {session.get(label_field)}.",
                        session_key=candidate["session_key"],
                        related_session_keys=[session["session_key"]],
                    )
                )
            active.append(session)

    for bucket in by_location.values():
        scan_overlaps(bucket, code="location_overlap", label_field="location")

    for bucket in by_instructor.values():
        scan_overlaps(bucket, code="instructor_overlap", label_field="lead_instructor")

    return issues


def preferred_identifier(session: dict[str, Any]) -> str:
    return (
        session.get("course_number")
        or session.get("course_code")
        or session.get("normalized_course_key")
        or session.get("course_name_clean")
        or "unknown-course"
    )


def compare_sessions(desired: list[dict[str, Any]], actual: list[dict[str, Any]]) -> dict[str, Any]:
    actual_by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for session in actual:
        actual_by_key[session["session_key"]].append(session)

    desired_by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for session in desired:
        desired_by_key[session["session_key"]].append(session)

    report: dict[str, list[dict[str, Any]]] = {
        "would_create": [],
        "would_update": [],
        "would_skip": [],
        "duplicate_risk": [],
        "missing_in_enrollware": [],
        "extra_in_enrollware": [],
        "manual_review_needed": [],
    }

    for desired_session in desired:
        key = desired_session["session_key"]
        matches = actual_by_key.get(key, [])
        if not matches:
            item = {
                "session_key": key,
                "desired_identifier": preferred_identifier(desired_session),
                "reason": "No exact Enrollware match found for desired session.",
                "desired": desired_session,
            }
            report["would_create"].append(item)
            report["missing_in_enrollware"].append(item)
            continue

        if len(matches) > 1:
            report["duplicate_risk"].append(
                {
                    "session_key": key,
                    "desired_identifier": preferred_identifier(desired_session),
                    "reason": "Multiple Enrollware sessions match this desired slot.",
                    "desired": desired_session,
                    "actual_matches": matches,
                }
            )
            continue

        actual_session = matches[0]
        diffs = {}
        for field in ("course_name_clean", "lead_instructor", "assistants", "students", "seats", "hours", "location", "registration_url"):
            if desired_session.get(field) != actual_session.get(field):
                diffs[field] = {"desired": desired_session.get(field), "actual": actual_session.get(field)}
        if diffs:
            report["would_update"].append(
                {
                    "session_key": key,
                    "desired_identifier": preferred_identifier(desired_session),
                    "actual_id": actual_session.get("source_id"),
                    "diffs": diffs,
                    "desired": desired_session,
                    "actual": actual_session,
                }
            )
        else:
            report["would_skip"].append(
                {
                    "session_key": key,
                    "desired_identifier": preferred_identifier(desired_session),
                    "actual_id": actual_session.get("source_id"),
                }
            )

    for actual_session in actual:
        key = actual_session["session_key"]
        if key not in desired_by_key:
            report["extra_in_enrollware"].append(
                {
                    "session_key": key,
                    "actual_id": actual_session.get("source_id"),
                    "reason": "Enrollware contains a class not present in desired sessions.",
                    "actual": actual_session,
                }
            )

    return report


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def build_markdown_report(payload: dict[str, Any]) -> str:
    counts = payload["summary"]["counts"]
    lines = [
        "# Enrollware Sync Dry Run",
        "",
        f"- Generated: `{payload['generated_at']}`",
        f"- Master schedule: `{payload['inputs']['master_schedule_path']}`",
        f"- Enrollware export: `{payload['inputs']['enrollware_export_path']}`",
        "",
        "## Summary",
        "",
        f"- Desired sessions: `{counts['desired_sessions']}`",
        f"- Enrollware sessions: `{counts['enrollware_sessions']}`",
        f"- Validation issues: `{counts['validation_issues']}`",
        f"- Would create: `{counts['would_create']}`",
        f"- Would update: `{counts['would_update']}`",
        f"- Would skip: `{counts['would_skip']}`",
        f"- Duplicate risk: `{counts['duplicate_risk']}`",
        f"- Missing in Enrollware: `{counts['missing_in_enrollware']}`",
        f"- Extra in Enrollware: `{counts['extra_in_enrollware']}`",
        "",
    ]

    if payload["validation_issues"]:
        lines.extend(["## Validation Issues", ""])
        for issue in payload["validation_issues"][:25]:
            lines.append(f"- `{issue['severity']}` `{issue['code']}`: {issue['message']}")
        if len(payload["validation_issues"]) > 25:
            lines.append(f"- ... and `{len(payload['validation_issues']) - 25}` more")
        lines.append("")

    for section in ("would_create", "would_update", "duplicate_risk", "extra_in_enrollware"):
        items = payload["report"].get(section, [])
        if not items:
            continue
        lines.extend([f"## {section.replace('_', ' ').title()}", ""])
        for item in items[:20]:
            label = item.get("desired_identifier") or item.get("actual_id") or item.get("session_key")
            reason = item.get("reason") or ("fields changed" if section == "would_update" else "")
            lines.append(f"- `{label}` {reason}".rstrip())
        if len(items) > 20:
            lines.append(f"- ... and `{len(items) - 20}` more")
        lines.append("")

    lines.extend(
        [
            "## Safety",
            "",
            "- No classes were created, edited, or deleted in this phase.",
            "- Extra Enrollware classes are flagged for manual review only.",
            "",
        ]
    )
    return "\n".join(lines)


def dry_run(args: argparse.Namespace) -> int:
    repo_root = Path(args.repo_root).resolve()
    master_path = (repo_root / args.master).resolve()
    export_path = (repo_root / args.enrollware_export).resolve()
    rules_path = (repo_root / args.rules).resolve() if args.rules else None
    output_dir = (repo_root / args.output_dir).resolve()
    timestamp = datetime.now(TZ).strftime("%Y%m%d-%H%M%S")
    run_dir = output_dir / timestamp

    rules = load_rules(rules_path)
    desired_rows = read_tabular_rows(master_path)
    enrollware_rows = read_tabular_rows(export_path)

    desired_sessions = [
        normalize_session_record(row, source_name="master_schedule", row_index=index)
        for index, row in enumerate(desired_rows, start=2)
    ]
    enrollware_sessions = [
        normalize_session_record(row, source_name="enrollware_export", row_index=index)
        for index, row in enumerate(enrollware_rows, start=2)
    ]
    desired_sessions = filter_sessions_by_time(desired_sessions, include_past=args.include_past)
    enrollware_sessions = filter_sessions_by_time(enrollware_sessions, include_past=args.include_past)

    validation_issues = [asdict(issue) for issue in validate_sessions(desired_sessions, rules)]
    comparison = compare_sessions(desired_sessions, enrollware_sessions)
    manual_review_needed = []
    if validation_issues:
        manual_review_needed.extend(validation_issues)
    manual_review_needed.extend(comparison["duplicate_risk"])
    comparison["manual_review_needed"] = manual_review_needed

    desired_payload = {
        "generated_at": datetime.now(TZ).isoformat(),
        "source_file": str(master_path),
        "source_file_sha256": file_sha256(master_path),
        "future_only_mode": not args.include_past,
        "sessions": desired_sessions,
    }
    desired_path = run_dir / "desired_sessions.json"
    write_json(desired_path, desired_payload)

    report_payload = {
        "generated_at": datetime.now(TZ).isoformat(),
        "phase": "dry_run",
        "inputs": {
            "master_schedule_path": str(master_path),
            "master_schedule_sha256": file_sha256(master_path),
            "enrollware_export_path": str(export_path),
            "enrollware_export_sha256": file_sha256(export_path),
            "rules_path": str(rules_path) if rules_path else None,
        },
        "summary": {
            "counts": {
                "desired_sessions": len(desired_sessions),
                "enrollware_sessions": len(enrollware_sessions),
                "future_only_mode": not args.include_past,
                "validation_issues": len(validation_issues),
                "would_create": len(comparison["would_create"]),
                "would_update": len(comparison["would_update"]),
                "would_skip": len(comparison["would_skip"]),
                "duplicate_risk": len(comparison["duplicate_risk"]),
                "missing_in_enrollware": len(comparison["missing_in_enrollware"]),
                "extra_in_enrollware": len(comparison["extra_in_enrollware"]),
                "manual_review_needed": len(comparison["manual_review_needed"]),
            }
        },
        "validation_issues": validation_issues,
        "report": comparison,
        "outputs": {
            "desired_sessions_path": str(desired_path),
        },
    }
    report_path = run_dir / "sync_report.json"
    write_json(report_path, report_payload)
    markdown_path = run_dir / "sync_report.md"
    markdown_path.write_text(build_markdown_report(report_payload), encoding="utf-8")

    print(f"Desired sessions written: {desired_path}")
    print(f"Sync report written: {report_path}")
    print(f"Markdown summary written: {markdown_path}")
    print(json.dumps(report_payload["summary"]["counts"], indent=2))
    return 0


def validate_apply_gates(args: argparse.Namespace, dry_run_report: dict[str, Any]) -> list[str]:
    blockers: list[str] = []
    if not args.apply:
        blockers.append("Live writes are disabled because --apply was not passed.")
    if args.limit is None:
        blockers.append("Live writes are disabled because --limit was not passed.")
    if not args.enrollware_snapshot:
        blockers.append("Live writes are disabled because a timestamped --enrollware-snapshot export was not supplied.")
    elif not Path(args.enrollware_snapshot).exists():
        blockers.append("Live writes are disabled because the supplied --enrollware-snapshot file does not exist.")
    if dry_run_report["inputs"]["master_schedule_sha256"] != file_sha256(Path(args.master).resolve()):
        blockers.append("Dry-run report does not match the current master schedule file.")
    return blockers


def playwright_scaffold(args: argparse.Namespace) -> int:
    dry_run_report_path = Path(args.dry_run_report).resolve() if args.dry_run_report else None
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(TZ).strftime("%Y%m%d-%H%M%S")
    log_path = output_dir / f"playwright_scaffold_{timestamp}.json"
    report: dict[str, Any] = {
        "generated_at": datetime.now(TZ).isoformat(),
        "phase": "playwright_scaffold",
        "apply_requested": bool(args.apply),
        "actions_attempted": [],
        "blockers": [],
    }

    if not dry_run_report_path or not dry_run_report_path.exists():
        report["blockers"].append("A dry-run report is required before browser automation can proceed.")
        write_json(log_path, report)
        print(f"Scaffold report written: {log_path}")
        return 1

    dry_run_report = json.loads(dry_run_report_path.read_text(encoding="utf-8"))
    report["dry_run_report"] = str(dry_run_report_path)

    for blocker in validate_apply_gates(args, dry_run_report):
        if args.apply:
            report["blockers"].append(blocker)

    rules_path = Path(args.rules).resolve() if args.rules else Path(dry_run_report["inputs"].get("rules_path") or "")
    rules = load_rules(rules_path if rules_path.exists() else None)
    playwright_rules = rules.get("playwright", {})
    create_url = playwright_rules.get("create_url")
    selectors = playwright_rules.get("selectors", {})
    if not create_url:
        report["blockers"].append("Missing Playwright config: create_url")
    if not selectors:
        report["blockers"].append("Missing Playwright config: selectors")

    try:
        from playwright.sync_api import sync_playwright  # type: ignore
    except Exception as exc:  # pragma: no cover
        report["blockers"].append(f"Playwright is not available: {exc}")
        write_json(log_path, report)
        print(f"Scaffold report written: {log_path}")
        return 1

    actions = dry_run_report.get("report", {}).get("would_create", []) + dry_run_report.get("report", {}).get("would_update", [])
    if args.limit is not None:
        actions = actions[: args.limit]

    if report["blockers"]:
        write_json(log_path, report)
        print(f"Scaffold report written: {log_path}")
        return 1

    with sync_playwright() as playwright:  # pragma: no cover
        browser = playwright.chromium.launch(headless=not args.headed)
        context_kwargs: dict[str, Any] = {}
        if args.storage_state and Path(args.storage_state).exists():
            context_kwargs["storage_state"] = str(Path(args.storage_state).resolve())
        context = browser.new_context(**context_kwargs)
        page = context.new_page()

        for action in actions:
            desired = action.get("desired") or action.get("actual") or {}
            attempted = {
                "mode": "update" if "diffs" in action else "create",
                "session_key": desired.get("session_key"),
                "course": desired.get("course_name_clean"),
                "start_at": desired.get("start_at"),
                "status": "stopped_before_save",
            }
            try:
                page.goto(create_url, wait_until="domcontentloaded")
                attempted["page_opened"] = create_url
                # Selectors are intentionally config-driven; we do not guess around auth or controls.
                title_selector = selectors.get("course_title")
                start_selector = selectors.get("start_at")
                end_selector = selectors.get("end_at")
                location_selector = selectors.get("location")
                instructor_selector = selectors.get("instructor")
                if not all([title_selector, start_selector, end_selector]):
                    attempted["status"] = "blocked"
                    attempted["blocker"] = "Missing required field selectors in Playwright config."
                else:
                    page.locator(title_selector).fill(desired.get("course_name_clean") or "")
                    page.locator(start_selector).fill(desired.get("start_at") or "")
                    page.locator(end_selector).fill(desired.get("end_at") or "")
                    if location_selector:
                        page.locator(location_selector).fill(desired.get("location") or "")
                    if instructor_selector:
                        page.locator(instructor_selector).fill(desired.get("lead_instructor") or "")
                    if args.apply:
                        save_selector = selectors.get("save")
                        if not save_selector:
                            attempted["status"] = "blocked"
                            attempted["blocker"] = "Missing save selector, so live write cannot continue."
                        else:
                            page.locator(save_selector).click()
                            attempted["status"] = "save_clicked"
                    else:
                        attempted["status"] = "filled_no_save"
            except Exception as exc:
                attempted["status"] = "blocked"
                attempted["blocker"] = str(exc)
            report["actions_attempted"].append(attempted)

        context.close()
        browser.close()

    write_json(log_path, report)
    print(f"Scaffold report written: {log_path}")
    return 0 if not report["blockers"] else 1


def reconcile(args: argparse.Namespace) -> int:
    desired_path = Path(args.desired_sessions).resolve()
    export_path = Path(args.enrollware_export).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    desired_payload = json.loads(desired_path.read_text(encoding="utf-8"))
    desired_sessions = desired_payload.get("sessions", [])
    include_past = not bool(desired_payload.get("future_only_mode", True))
    actual_rows = read_tabular_rows(export_path)
    actual_sessions = [
        normalize_session_record(row, source_name="enrollware_export", row_index=index)
        for index, row in enumerate(actual_rows, start=2)
    ]
    actual_sessions = filter_sessions_by_time(actual_sessions, include_past=include_past)

    comparison = compare_sessions(desired_sessions, actual_sessions)
    reconciled_sessions = []
    actual_by_key = defaultdict(list)
    for session in actual_sessions:
        actual_by_key[session["session_key"]].append(session)

    for desired in desired_sessions:
        matches = actual_by_key.get(desired["session_key"], [])
        match = matches[0] if len(matches) == 1 else None
        reconciled_sessions.append(
            {
                **desired,
                "enrollware_class_id": match.get("source_id") if match else None,
                "registration_url": match.get("registration_url") if match else desired.get("registration_url"),
            }
        )

    payload = {
        "generated_at": datetime.now(TZ).isoformat(),
        "phase": "reconciliation",
        "inputs": {
            "desired_sessions_path": str(desired_path),
            "enrollware_export_path": str(export_path),
            "enrollware_export_sha256": file_sha256(export_path),
        },
        "counts": {
            "reconciled_sessions": len(reconciled_sessions),
            "remaining_missing_in_enrollware": len(comparison["missing_in_enrollware"]),
            "remaining_extra_in_enrollware": len(comparison["extra_in_enrollware"]),
            "remaining_duplicate_risk": len(comparison["duplicate_risk"]),
        },
        "reconciled_sessions": reconciled_sessions,
        "comparison": comparison,
    }
    timestamp = datetime.now(TZ).strftime("%Y%m%d-%H%M%S")
    out_path = output_dir / f"reconciliation_{timestamp}.json"
    write_json(out_path, payload)
    print(f"Reconciliation report written: {out_path}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Spreadsheet-controlled Enrollware sync workflow.")
    parser.add_argument("--repo-root", default=".", help="Repo root path. Defaults to current working directory.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    dry = subparsers.add_parser("dry-run", help="Build desired sessions, validate them, and compare against Enrollware export.")
    dry.add_argument("--master", default=str(DEFAULT_MASTER_PATH), help="Master schedule CSV/XLSX path relative to repo root.")
    dry.add_argument("--enrollware-export", default=str(DEFAULT_EXPORT_PATH), help="Enrollware export CSV/XLSX path relative to repo root.")
    dry.add_argument("--rules", default=str(DEFAULT_RULES_PATH), help="Rules JSON path relative to repo root.")
    dry.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Output directory for desired sessions and reports.")
    dry.add_argument("--include-past", action="store_true", help="Include historical sessions instead of defaulting to future-only sync.")
    dry.set_defaults(handler=dry_run)

    scaffold = subparsers.add_parser("playwright-scaffold", help="Open Enrollware pages and fill fields, but stop before save unless --apply is explicitly passed.")
    scaffold.add_argument("--dry-run-report", required=True, help="Path to sync_report.json from a prior dry run.")
    scaffold.add_argument("--master", default=str(DEFAULT_MASTER_PATH), help="Master schedule path used for the dry run.")
    scaffold.add_argument("--rules", default=str(DEFAULT_RULES_PATH), help="Rules JSON path with Playwright selectors.")
    scaffold.add_argument("--storage-state", help="Optional Playwright storage-state JSON from an authorized session.")
    scaffold.add_argument("--headed", action="store_true", help="Launch browser visibly for manual observation.")
    scaffold.add_argument("--apply", action="store_true", help="Allow final save click if all safety gates pass.")
    scaffold.add_argument("--limit", type=int, help="Maximum number of create/update actions to attempt.")
    scaffold.add_argument("--enrollware-snapshot", help="Timestamped Enrollware export path required for live writes.")
    scaffold.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Output directory for scaffold logs.")
    scaffold.set_defaults(handler=playwright_scaffold)

    reconcile_parser = subparsers.add_parser("reconcile", help="Recompare desired sessions against a fresh Enrollware export after any write.")
    reconcile_parser.add_argument("--desired-sessions", required=True, help="Path to desired_sessions.json")
    reconcile_parser.add_argument("--enrollware-export", required=True, help="Fresh Enrollware export CSV/XLSX path")
    reconcile_parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Output directory for reconciliation reports.")
    reconcile_parser.set_defaults(handler=reconcile)
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    repo_root = Path(args.repo_root).resolve()
    args.repo_root = str(repo_root)
    if hasattr(args, "master"):
        args.master = str((repo_root / args.master).resolve()) if not Path(args.master).is_absolute() else str(Path(args.master).resolve())
    if hasattr(args, "enrollware_export") and args.enrollware_export:
        args.enrollware_export = str((repo_root / args.enrollware_export).resolve()) if not Path(args.enrollware_export).is_absolute() else str(Path(args.enrollware_export).resolve())
    if hasattr(args, "rules") and args.rules:
        args.rules = str((repo_root / args.rules).resolve()) if not Path(args.rules).is_absolute() else str(Path(args.rules).resolve())
    if hasattr(args, "output_dir") and args.output_dir:
        args.output_dir = str((repo_root / args.output_dir).resolve()) if not Path(args.output_dir).is_absolute() else str(Path(args.output_dir).resolve())
    return args.handler(args)


if __name__ == "__main__":
    raise SystemExit(main())
