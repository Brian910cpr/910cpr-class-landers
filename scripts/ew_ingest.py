from __future__ import annotations
import json, re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import openpyxl

META_RE = re.compile(r"<!--\s*META:\s*(\{.*?\})\s*-->", re.DOTALL | re.IGNORECASE)
IMG_ATTR_RE = re.compile(r'(\w+)\s*=\s*"([^"]*)"')
TAG_RE = re.compile(r"<[^>]+>")
BR_RE = re.compile(r"<br\s*/?>", re.IGNORECASE)

@dataclass
class CourseRecord:
    row_number: int
    name: str
    description: str
    discipline: str
    add_ons: str
    class_price_text: str
    shipping_price_text: str
    card_type: str
    secondary_card_type: str
    ecard_code: str
    meta: Dict[str, Any]
    meta_key: str
    course_id: Optional[str]
    family: str
    certifying_body: str
    delivery: str
    price: Optional[float]
    slug: str

@dataclass
class SessionRecord:
    session_id: str
    course_html: str
    course_text: str
    start_dt: Optional[datetime]
    end_dt: Optional[datetime]
    location: str
    instructor: str
    students: int
    seats: int
    hours: float
    registration_link: str
    source_course_id: Optional[str]
    source_meta_key: str
    source_certifying_body: str
    source_delivery: str
    source_price: Optional[float]
    matched_course: Optional[CourseRecord]

def slugify(value: str) -> str:
    value = value.lower().strip().replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-{2,}", "-", value)
    return value.strip("-") or "item"

def clean_html_text(html: str) -> str:
    html = BR_RE.sub(" ", html or "")
    html = TAG_RE.sub(" ", html)
    return re.sub(r"\s+", " ", html).strip()

def extract_meta_json(description: str) -> Dict[str, Any]:
    m = META_RE.search(description or "")
    if not m: return {}
    try: return json.loads(m.group(1))
    except Exception: return {}

def parse_money(value: Any) -> Optional[float]:
    if value in (None, ""): return None
    if isinstance(value, (int, float)): return float(value)
    text = str(value).strip().replace("$","").replace(",","")
    try: return float(text)
    except Exception: return None

def parse_int(value: Any, default: int = 0) -> int:
    try: return int(value)
    except Exception: return default

def parse_float(value: Any, default: float = 0.0) -> float:
    try: return float(value)
    except Exception: return default

def normalize_location(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^:+\s*", "", text)
    return re.sub(r"\s+", " ", text)

def load_first_sheet_rows(xlsx_path: Path) -> Tuple[List[str], List[Tuple[Any, ...]]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(v or "").strip() for v in next(rows)]
    return headers, list(rows)

def parse_course_export(xlsx_path: Path) -> List[CourseRecord]:
    headers, rows = load_first_sheet_rows(xlsx_path)
    idx = {name:i for i,name in enumerate(headers)}
    out = []
    for row_no, row in enumerate(rows, start=2):
        name = str(row[idx.get("Name",-1)] or "").strip()
        if not name: continue
        description = str(row[idx.get("Description",-1)] or "")
        discipline = str(row[idx.get("Discipline",-1)] or "").strip()
        meta = extract_meta_json(description)
        course_id = str(meta.get("course_id")) if meta.get("course_id") not in (None, "") else None
        meta_key = str(meta.get("meta_key") or "").strip()
        family = str(meta.get("family") or discipline or "").strip()
        cert = str(meta.get("certifying_body") or "").strip()
        delivery = str(meta.get("delivery") or "").strip()
        price = parse_money(meta.get("price"))
        if price is None: price = parse_money(row[idx.get("Class Price",-1)])
        slug = slugify(f"{course_id or meta_key or name}-{family or cert or 'course'}")[:80]
        out.append(CourseRecord(
            row_number=row_no,
            name=name,
            description=description,
            discipline=discipline,
            add_ons=str(row[idx.get("Add-ons",-1)] or "").strip(),
            class_price_text=str(row[idx.get("Class Price",-1)] or "").strip(),
            shipping_price_text=str(row[idx.get("Shipping Price",-1)] or "").strip(),
            card_type=str(row[idx.get("Card Type",-1)] or "").strip(),
            secondary_card_type=str(row[idx.get("Secondary Card Type",-1)] or "").strip(),
            ecard_code=str(row[idx.get("eCard Code",-1)] or "").strip(),
            meta=meta,
            meta_key=meta_key,
            course_id=course_id,
            family=family,
            certifying_body=cert,
            delivery=delivery,
            price=price,
            slug=slug,
        ))
    return out

def parse_course_html_metadata(course_html: str) -> Dict[str, Any]:
    out={}
    m = re.search(r"<img\b([^>]+)>", course_html or "", re.IGNORECASE)
    if not m: return out
    attrs = dict((k.lower(), v) for k,v in IMG_ATTR_RE.findall(m.group(1)))
    if attrs.get("name"): out["course_id"] = attrs["name"].strip()
    for piece in attrs.get("longdesc","").split("|"):
        if ":" in piece:
            k,v = piece.split(":",1)
            out[k.strip()] = v.strip()
    return out

def classify_topic(course_text: str, cert_body: str, meta_key: str) -> str:
    text=(course_text or "").upper(); body=(cert_body or "").upper(); meta=(meta_key or "").upper()
    if "BLS" in text or "BLS" in meta: return "bls"
    if "ACLS" in text or "ACLS" in meta: return "acls"
    if "PALS" in text or "PALS" in meta: return "pals"
    if "HEARTSAVER" in text or "HS_" in meta: return "heartsaver"
    if "RED CROSS" in text or body=="ARC": return "red-cross"
    if "HSI" in text or body=="HSI": return "hsi"
    if "USCG" in text or "COAST GUARD" in text: return "uscg"
    if "INSTRUCTOR" in text: return "instructor"
    if "FAMILY" in text and "FRIENDS" in text: return "family-friends"
    if "STOP THE BLEED" in text: return "stop-the-bleed"
    if "AED" in text and "CLASS" not in text: return "aed"
    if "FIRST AID" in text: return "first-aid"
    return "misc"

def course_to_dict(c: CourseRecord) -> Dict[str, Any]:
    return c.__dict__

def session_to_dict(s: SessionRecord) -> Dict[str, Any]:
    course=s.matched_course
    title=course.name if course else s.course_text
    topic=classify_topic(s.course_text, s.source_certifying_body, s.source_meta_key)
    return {
        "session_id": s.session_id,
        "title": title,
        "course_text": s.course_text,
        "course_html": s.course_html,
        "start": s.start_dt.isoformat() if s.start_dt else None,
        "end": s.end_dt.isoformat() if s.end_dt else None,
        "year": s.start_dt.year if s.start_dt else None,
        "location": s.location,
        "location_slug": slugify(s.location) if s.location else "unknown-location",
        "instructor": s.instructor,
        "students": s.students,
        "seats": s.seats,
        "hours": s.hours,
        "registration_link": s.registration_link,
        "source_course_id": s.source_course_id,
        "source_meta_key": s.source_meta_key,
        "source_certifying_body": s.source_certifying_body,
        "source_delivery": s.source_delivery,
        "source_price": s.source_price,
        "topic": topic,
        "course": course_to_dict(course) if course else None,
        "course_slug": course.slug if course else slugify(title)[:80],
        "course_page": f"courses/{course.slug if course else slugify(title)[:80]}.html",
        "class_page": f"classes/{s.session_id}.html",
    }

def parse_class_report(xlsx_path: Path, courses: List[CourseRecord]) -> List[SessionRecord]:
    headers, rows = load_first_sheet_rows(xlsx_path)
    idx={name:i for i,name in enumerate(headers)}
    by_id={c.course_id:c for c in courses if c.course_id}
    by_meta={c.meta_key:c for c in courses if c.meta_key}
    by_name={slugify(c.name):c for c in courses if c.name}
    out=[]
    for row in rows:
        sid = str(row[idx.get("ID",-1)] or "").strip()
        if not sid: continue
        course_html = str(row[idx.get("Course",-1)] or "")
        course_text = clean_html_text(course_html)
        meta=parse_course_html_metadata(course_html)
        source_course_id = str(meta.get("course_id")) if meta.get("course_id") not in (None, "") else None
        source_meta_key = str(meta.get("t") or "").strip()
        matched = by_id.get(source_course_id) or by_meta.get(source_meta_key) or by_name.get(slugify(course_text))
        out.append(SessionRecord(
            session_id=sid,
            course_html=course_html,
            course_text=course_text,
            start_dt=row[idx.get("Start Date / Time",-1)] if isinstance(row[idx.get("Start Date / Time",-1)], datetime) else None,
            end_dt=row[idx.get("End Date / Time",-1)] if isinstance(row[idx.get("End Date / Time",-1)], datetime) else None,
            location=normalize_location(row[idx.get("Location",-1)]),
            instructor=str(row[idx.get("Instructor",-1)] or "").strip(),
            students=parse_int(row[idx.get("Students",-1)], 0),
            seats=parse_int(row[idx.get("Seats",-1)], 0),
            hours=parse_float(row[idx.get("Hours",-1)], 0.0),
            registration_link=str(row[idx.get("Registration Link",-1)] or "").strip(),
            source_course_id=source_course_id,
            source_meta_key=source_meta_key,
            source_certifying_body=str(meta.get("cb") or "").strip(),
            source_delivery=str(meta.get("d") or "").strip(),
            source_price=parse_money(meta.get("p")),
            matched_course=matched,
        ))
    return out

def load_normalized(course_export_path: Path, class_report_path: Path) -> Dict[str, Any]:
    courses=parse_course_export(course_export_path)
    sessions=parse_class_report(class_report_path, courses)
    return {"courses":[course_to_dict(c) for c in courses], "sessions":[session_to_dict(s) for s in sessions]}
