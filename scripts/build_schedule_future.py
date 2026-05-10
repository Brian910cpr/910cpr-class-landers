from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from scripts.build_status import BuildStatusReporter
from supervisor.status_snapshot import write_status_snapshot
from typing import Any, Optional
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from scripts.course_identity_resolver import (
    load_aliases,
    normalize_title,
    resolve_course_identity,
    strip_html_tags,
)


TZ = ZoneInfo("America/New_York")
ID_FROM_LINK_RE = re.compile(r"[?&]id=(\d+)")


def clean_string(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def parse_iso_dt(value: Any) -> Optional[datetime]:
    s = clean_string(value)
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=TZ)
        else:
            dt = dt.astimezone(TZ)
        return dt
    except Exception:
        return None


def first_present(mapping: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        if key in mapping and mapping.get(key) not in (None, ""):
            return mapping.get(key)
    return None


def session_start_candidate(session: dict[str, Any]) -> Any:
    timing = session.get("timing", {}) if isinstance(session.get("timing"), dict) else {}
    start_val = first_present(
        session,
        [
            "start",
            "start_datetime",
            "Start Date / Time",
            "start_at",
        ],
    )
    if start_val is not None:
        return start_val
    return first_present(
        timing,
        [
            "start",
            "start_datetime",
            "Start Date / Time",
            "start_at",
        ],
    )


def session_end_candidate(session: dict[str, Any]) -> Any:
    timing = session.get("timing", {}) if isinstance(session.get("timing"), dict) else {}
    end_val = first_present(
        session,
        [
            "end",
            "end_datetime",
            "End Date / Time",
            "end_at",
        ],
    )
    if end_val is not None:
        return end_val
    return first_present(
        timing,
        [
            "end",
            "end_datetime",
            "End Date / Time",
            "end_at",
        ],
    )


def debug_location_value(session: dict[str, Any]) -> str:
    location = session.get("location", {}) if isinstance(session.get("location"), dict) else {}
    value = (
        location.get("location_display")
        or location.get("location_name")
        or session.get("location_display")
        or session.get("location_name")
        or (session.get("location") if not isinstance(session.get("location"), dict) else None)
        or ""
    )
    return clean_string(value) or ""


def resolve_class_report_path(repo_root: Path, requested: str) -> Path:
    requested_path = (repo_root / requested).resolve()
    if requested_path.exists():
        return requested_path
    candidates = [
        repo_root / "data" / "Class Report.xlsx",
        repo_root / "data" / "raw" / "Class Report.xlsx",
        repo_root / "data" / "raw" / "class_report.xlsx",
    ]
    for path in candidates:
        if path.exists():
            return path.resolve()
    return requested_path


def extract_id_from_report_row(row: tuple[Any, ...], reg_idx: int | None, id_idx: int | None) -> str:
    sid = ""
    if reg_idx is not None and reg_idx < len(row):
        reg_val = row[reg_idx]
        if reg_val:
            m = ID_FROM_LINK_RE.search(str(reg_val))
            if m:
                sid = m.group(1)
    if not sid and id_idx is not None and id_idx < len(row):
        raw_id = row[id_idx]
        sid = str(raw_id or "").strip()
    return sid


def read_class_report_ids(path: Path) -> set[str]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return set()
    headers = [str(h or "").strip() for h in rows[0]]
    reg_idx = headers.index("Registration Link") if "Registration Link" in headers else None
    id_idx = headers.index("ID") if "ID" in headers else None
    ids: set[str] = set()
    for row in rows[1:]:
        sid = extract_id_from_report_row(row, reg_idx, id_idx)
        if sid:
            ids.add(sid)
    return ids


def build_public_future_session(session: dict[str, Any], course_identity: dict[str, Any] | None = None) -> dict[str, Any]:
    course = session.get("course", {})
    timing = session.get("timing", {})
    location = session.get("location", {})
    staffing = session.get("staffing", {})
    capacity = session.get("capacity", {})
    commerce = session.get("commerce", {})
    status = session.get("status", {})

    start_raw = session_start_candidate(session)
    end_raw = session_end_candidate(session)
    raw_course_name = course.get("course_name_raw") or course.get("course_name_primary_raw") or course.get("course_name_primary_clean")
    identity = course_identity or {}
    return {
        "session_id": session.get("session_id"),
        "course_id": course.get("course_id"),
        "course_key": identity.get("course_key") or session.get("course_key"),
        "official_course_name": identity.get("official_course_name") or session.get("official_course_name"),
        "raw_course_name": raw_course_name,
        "course_match_method": identity.get("match_method") or session.get("course_match_method"),
        "course_match_confidence": identity.get("match_confidence") or session.get("course_match_confidence"),
        "course_needs_review": identity.get("needs_review") if "needs_review" in identity else session.get("course_needs_review"),
        "included_by_mapping_method": session.get("_included_by_mapping_method"),
        "course_name": course.get("course_name_primary_clean"),
        "course_subtitle": course.get("course_subtitle_text"),
        "course_number": course.get("course_number"),
        "course_code": course.get("course_code_hint"),
        "certifying_body": course.get("certifying_body_hint"),
        "delivery_mode": course.get("delivery_mode_hint"),
        "start_at": start_raw,
        "end_at": end_raw,
        "timezone": timing.get("timezone"),
        "location_name": location.get("location_name"),
        "location_display": location.get("location_display"),
        "client": location.get("client"),
        "lead_instructor_name": staffing.get("lead_instructor_name"),
        "price": commerce.get("price"),
        "max_students": capacity.get("max_students"),
        "registered_count": capacity.get("registered_count"),
        "enrolled_count": capacity.get("registered_count"),
        "available_seats": capacity.get("available_seats"),
        "is_full": capacity.get("is_full"),
        "registration_url": commerce.get("registration_url"),
        "session_status": status.get("session_status"),
        "mapped_family": session.get("mapped_family") or course.get("mapped_family"),
        "mapped_subtype": session.get("mapped_subtype") or course.get("mapped_subtype"),
        "mapped_certifying_body": session.get("mapped_certifying_body") or course.get("mapped_certifying_body"),
        "mapped_delivery_mode": session.get("mapped_delivery_mode") or course.get("mapped_delivery_mode"),
        "mapped_logo_key": session.get("mapped_logo_key") or course.get("mapped_logo_key"),
        "mapped_price": session.get("mapped_price") if session.get("mapped_price") is not None else course.get("mapped_price"),
        "mapped_clean_title": session.get("mapped_clean_title") or course.get("mapped_clean_title"),
        "mapped_short_description": session.get("mapped_short_description") or course.get("mapped_short_description"),
        "mapped_long_description": session.get("mapped_long_description") or course.get("mapped_long_description"),
        "mapped_who_class_for": session.get("mapped_who_class_for") or course.get("mapped_who_class_for"),
        "mapped_prerequisites": session.get("mapped_prerequisites") or course.get("mapped_prerequisites"),
        "mapped_what_to_expect": session.get("mapped_what_to_expect") or course.get("mapped_what_to_expect"),
        "mapped_certification_card": session.get("mapped_certification_card") or course.get("mapped_certification_card"),
        "mapped_renewal_info": session.get("mapped_renewal_info") or course.get("mapped_renewal_info"),
        "mapped_official_description_source": session.get("mapped_official_description_source") or course.get("mapped_official_description_source"),
        "mapped_description_status": session.get("mapped_description_status") or course.get("mapped_description_status"),
        "mapping_status": session.get("mapping_status") or course.get("mapping_status") or "unmapped",
        "mapping_notes": session.get("mapping_notes") or course.get("mapping_notes") or [],
    }


def identity_can_include_session(identity: dict[str, Any]) -> bool:
    confidence = str(identity.get("match_confidence") or "").strip().lower()
    course_key = str(identity.get("course_key") or "").strip()
    alias_status = str(identity.get("alias_status") or "").strip().lower()
    needs_review = bool(identity.get("needs_review"))
    return bool(course_key) and confidence in {"high", "medium"} and (not needs_review or alias_status == "active_alias_needs_review")


def main() -> int:
    reporter = BuildStatusReporter("build_schedule_future")
    parser = argparse.ArgumentParser(description="Build docs/data/schedule_future.json from data/sessions_current.json")
    parser.add_argument(
        "--repo-root",
        default=".",
        help="Path to repo root. Defaults to current working directory.",
    )
    parser.add_argument(
        "--input",
        default="data/sessions_current.json",
        help="Relative path from repo root to sessions_current.json",
    )
    parser.add_argument(
        "--output",
        default="docs/data/schedule_future.json",
        help="Relative path from repo root to output future schedule JSON",
    )
    parser.add_argument(
        "--class-report",
        default="data/Class Report.xlsx",
        help="Relative path from repo root to Class Report.xlsx used for hard reconciliation",
    )
    parser.add_argument(
        "--course-aliases",
        default="data/course_aliases.json",
        help="Relative path from repo root to legacy course alias mapping JSON",
    )
    parser.add_argument(
        "--course-map",
        default="data/config/course_map.json",
        help="Relative path from repo root to structured course reference JSON",
    )
    parser.add_argument(
        "--include-past",
        action="store_true",
        help="Include past sessions instead of filtering them out. Used for full class-page rebuilds.",
    )
    parser.add_argument(
        "--fail-on-unmapped",
        action="store_true",
        help="Fail build if any output session is unmapped.",
    )
    args = parser.parse_args()

    repo_root = Path(args.repo_root).resolve()
    input_path = (repo_root / args.input).resolve()
    output_path = (repo_root / args.output).resolve()
    class_report_path = resolve_class_report_path(repo_root, args.class_report)
    aliases_path = (repo_root / args.course_aliases).resolve()
    course_map_path = (repo_root / args.course_map).resolve()
    unmatched_path = repo_root / "debug" / "unmatched_courses.json"
    match_report_path = repo_root / "debug" / "course_identity_match_report.json"
    reporter.set_context(inputs=[input_path, class_report_path, aliases_path, course_map_path], outputs=[output_path, unmatched_path, match_report_path])

    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")
    if not class_report_path.exists():
        raise SystemExit(f"Class report not found: {class_report_path}")
    if not course_map_path.exists():
        raise SystemExit(f"Course map not found: {course_map_path}")

    reporter.waiting(total=0)
    try:
        class_report_ids = read_class_report_ids(class_report_path)
        aliases = load_aliases(aliases_path)
        course_reference = json.loads(course_map_path.read_text(encoding="utf-8"))
        raw = json.loads(input_path.read_text(encoding="utf-8"))
        sessions = raw.get("sessions", [])
        reporter.start(total=len(sessions))
        print(f"Loaded {len(sessions)} sessions from {input_path}")
        print(f"Loaded {len(class_report_ids)} session IDs from {class_report_path}")
        print("Filtering future sessions" if not args.include_past else "Classifying all sessions")

        now_dt = datetime.now(TZ)
        now_iso = now_dt.isoformat()

        output_sessions_raw: list[dict[str, Any]] = []
        identities_by_session_id: dict[str, dict[str, Any]] = {}
        unmatched_rows: list[dict[str, Any]] = []
        match_counts = {
            "course_id": 0,
            "metadata": 0,
            "legacy_alias": 0,
            "normalized_title": 0,
            "unmatched": 0,
        }
        needs_review_count = 0
        skipped_missing_start = 0
        skipped_past = 0
        skipped_orphan = 0
        included_by_legacy_mapping = 0
        included_by_course_identity_resolver = 0
        skipped_unmapped_after_resolver = 0

        for index, session in enumerate(sessions, start=1):
            session_id = str(session.get("session_id") or "").strip()
            identity = resolve_course_identity(session, aliases, course_reference)
            identities_by_session_id[session_id] = identity
            method = str(identity.get("match_method") or "unmatched")
            match_counts[method] = match_counts.get(method, 0) + 1
            if identity.get("needs_review"):
                needs_review_count += 1
            if identity.get("needs_review") or identity.get("match_method") == "unmatched" or identity.get("match_confidence") in {"low", "none"}:
                course = session.get("course", {}) if isinstance(session.get("course"), dict) else {}
                timing = session.get("timing", {}) if isinstance(session.get("timing"), dict) else {}
                capacity = session.get("capacity", {}) if isinstance(session.get("capacity"), dict) else {}
                raw_course = course.get("course_name_raw") or course.get("course_name_primary_raw") or course.get("course_name_primary_clean") or session.get("course_name")
                unmatched_rows.append(
                    {
                        "raw_course": raw_course or "",
                        "normalized_course": normalize_title(raw_course),
                        "session_id": session_id,
                        "start_datetime": timing.get("start_at") or session.get("start_at") or session.get("start_datetime") or session.get("start") or "",
                        "location": debug_location_value(session),
                        "students": capacity.get("registered_count") or capacity.get("students_count_raw") or session.get("registered_count") or 0,
                        "match_method": identity.get("match_method"),
                        "match_confidence": identity.get("match_confidence"),
                        "matched_alias": identity.get("matched_alias"),
                        "reason": "no alias or course_key match" if identity.get("match_method") == "unmatched" else "low confidence or needs review",
                    }
                )
            if not session_id:
                print("ORPHAN SESSION DETECTED: missing session_id in sessions_current row")
                skipped_orphan += 1
                reporter.update(current=index, total=len(sessions))
                continue
            if session_id not in class_report_ids:
                print(f"ORPHAN SESSION DETECTED: {session_id} not present in current Class Report")
                skipped_orphan += 1
                reporter.update(current=index, total=len(sessions))
                continue
            start_dt = parse_iso_dt(session_start_candidate(session))
            if start_dt is None:
                skipped_missing_start += 1
                reporter.update(current=index, total=len(sessions))
                continue
            is_past = start_dt < now_dt
            if is_past and not args.include_past:
                skipped_past += 1
                reporter.update(current=index, total=len(sessions))
                continue
            mapping_status = str(session.get("mapping_status") or "").strip().lower()
            if mapping_status == "mapped":
                session["_included_by_mapping_method"] = "legacy_mapping"
                included_by_legacy_mapping += 1
            elif identity_can_include_session(identity):
                session["_included_by_mapping_method"] = "course_identity_resolver"
                included_by_course_identity_resolver += 1
            else:
                print(f"UNMAPPED COURSE DETECTED: session_id={session_id} mapping_status={session.get('mapping_status')}")
                skipped_unmapped_after_resolver += 1
                reporter.update(current=index, total=len(sessions))
                continue
            session["_build_classification"] = "past" if is_past else "future"
            output_sessions_raw.append(session)
            reporter.update(current=index, total=len(sessions))

        output_sessions_raw.sort(
            key=lambda s: (
                s.get("timing", {}).get("start_at") or "",
                s.get("session_id") or "",
            )
        )

        future_sessions = []
        for s in output_sessions_raw:
            row = build_public_future_session(s, identities_by_session_id.get(str(s.get("session_id") or "").strip()))
            row["build_classification"] = s.get("_build_classification", "future")
            future_sessions.append(row)

        output = {
            "build": {
                "generated_at": now_iso,
                "source_file": str(input_path),
                "counts": {
                    "sessions_input": len(sessions),
                    "sessions_output": len(future_sessions),
                    "skipped_missing_start": skipped_missing_start,
                    "skipped_past": skipped_past,
                    "skipped_orphan": skipped_orphan,
                    "included_by_legacy_mapping": included_by_legacy_mapping,
                    "included_by_course_identity_resolver": included_by_course_identity_resolver,
                    "skipped_unmapped_after_resolver": skipped_unmapped_after_resolver,
                    "include_past": args.include_past,
                },
                "class_report": str(class_report_path),
            },
            "sessions": future_sessions,
        }

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(output, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        debug_generated_at = datetime.now(TZ).isoformat()
        unmatched_payload = {
            "generated_at": debug_generated_at,
            "unmatched": unmatched_rows,
        }
        match_report = {
            "generated_at": debug_generated_at,
            "total_sessions": len(sessions),
            "matched_by_course_id": match_counts.get("course_id", 0),
            "matched_by_metadata": match_counts.get("metadata", 0),
            "matched_by_legacy_alias": match_counts.get("legacy_alias", 0),
            "matched_by_normalized_title": match_counts.get("normalized_title", 0),
            "unmatched": match_counts.get("unmatched", 0),
            "needs_review": needs_review_count,
        }
        unmatched_path.parent.mkdir(parents=True, exist_ok=True)
        unmatched_path.write_text(json.dumps(unmatched_payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        match_report_path.write_text(json.dumps(match_report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

        warnings = []
        files_needing_review = []
        if skipped_orphan:
            warnings.append(f"{skipped_orphan} sessions were skipped because they were not present in the current Class Report.")
        if skipped_unmapped_after_resolver:
            warnings.append(f"{skipped_unmapped_after_resolver} sessions were skipped because course mapping was missing after resolver fallback.")
            files_needing_review.append(repo_root / "data" / "audit" / "unmapped_courses.json")
        reporter.done(
            current=len(sessions),
            total=len(sessions),
            last_output_file=output_path,
            counts={
                "sessions_input": len(sessions),
                "sessions_output": len(future_sessions),
                "skipped_missing_start": skipped_missing_start,
                "skipped_past": skipped_past,
                "skipped_orphan": skipped_orphan,
                "included_by_legacy_mapping": included_by_legacy_mapping,
                "included_by_course_identity_resolver": included_by_course_identity_resolver,
                "skipped_unmapped_after_resolver": skipped_unmapped_after_resolver,
                "course_identity_unmatched": match_counts.get("unmatched", 0),
                "course_identity_needs_review": needs_review_count,
            },
            warnings=warnings,
            files_needing_review=files_needing_review,
        )
        write_status_snapshot()
        print("Future schedule build complete")
        print(f"Wrote {output_path}")
        print(f"Future sessions written: {len(future_sessions)}")
        print(f"Skipped missing start: {skipped_missing_start}")
        print(f"Skipped past: {skipped_past}")
        print(f"Skipped orphan: {skipped_orphan}")
        print(f"Included by legacy mapping: {included_by_legacy_mapping}")
        print(f"Included by course identity resolver: {included_by_course_identity_resolver}")
        print(f"Skipped unmapped after resolver: {skipped_unmapped_after_resolver}")
        print(f"Course identity report: {match_report_path}")
        print(f"Unmatched course report: {unmatched_path}")

        if skipped_unmapped_after_resolver > 0:
            print("UNMAPPED COURSE DETECTED: one or more sessions were excluded due to missing structured mapping.")
            if args.fail_on_unmapped:
                return 2

        return 0
    except Exception:
        reporter.error(last_output_file=output_path if output_path.exists() else None)
        raise


if __name__ == "__main__":
    raise SystemExit(main())
