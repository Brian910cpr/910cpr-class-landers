from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "data" / "private" / "enrollware" / "student_report_latest.xlsx"
DEFAULT_OUTPUT = ROOT / "data" / "enrollware_student_snapshot.json"
TZ = ZoneInfo("America/New_York")


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def course_key(value: Any) -> str:
    text = clean(value).lower().replace("®", "")
    if "bls" in text and "heartcode" in text:
        return "aha_bls_heartcode"
    if "bls" in text and "provider" in text:
        return "aha_bls_provider_renewal" if "renew" in text else "aha_bls_provider_initial"
    if "acls" in text and "heartcode" in text:
        return "aha_acls_heartcode"
    if "acls" in text and "provider" in text:
        return "aha_acls_provider_renewal" if "renew" in text else "aha_acls_provider_initial"
    if "heartsaver" in text and "first aid" in text:
        return "aha_heartsaver_first_aid_cpr_aed"
    if "heartsaver" in text and "cpr" in text:
        return "aha_heartsaver_cpr_aed"
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def read_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    headers = [clean(item) for item in next(values)]
    return [dict(zip(headers, row)) for row in values if any(item not in (None, "") for item in row)]


def build_snapshot(rows: list[dict[str, Any]], *, source_name: str, source_modified_at: str) -> dict[str, Any]:
    grouped: dict[tuple[str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        start = row.get("Course Date")
        if not isinstance(start, datetime):
            continue
        key = (start.replace(tzinfo=TZ).isoformat(), course_key(row.get("Course")), clean(row.get("Instructor")), clean(row.get("Course Location")))
        grouped[key].append(row)
    classes = []
    for (start_at, normalized_course, instructor, location), members in sorted(grouped.items()):
        statuses = Counter(clean(row.get("Status")) or "UNKNOWN" for row in members)
        class_ids = sorted({clean(row.get("Class ID")) for row in members if clean(row.get("Class ID"))})
        classes.append({
            "class_id": class_ids[0] if len(class_ids) == 1 else None,
            "class_ids": class_ids,
            "start_at": start_at,
            "course_key": normalized_course,
            "course_name": clean(members[0].get("Course")),
            "location_name": location,
            "instructor_name": instructor,
            "seated_count": len(members),
            "status_counts": dict(sorted(statuses.items())),
        })
    covered_dates = [str(item["start_at"])[:10] for item in classes]
    return {
        "schema_version": "1.0",
        "source": source_name,
        "source_modified_at": source_modified_at,
        "privacy": "Class-level counts only; student PII intentionally excluded.",
        "counts": {"student_rows": len(rows), "classes_with_students": len(classes)},
        "class_date_coverage": {
            "start": min(covered_dates) if covered_dates else None,
            "end": max(covered_dates) if covered_dates else None,
        },
        "classes": classes,
    }


def apply_snapshot_to_sessions(sessions: list[dict[str, Any]], snapshot: Any) -> dict[str, int]:
    evidence = snapshot.get("classes", []) if isinstance(snapshot, dict) else []
    by_key = {
        (str(item.get("start_at")), str(item.get("course_key"))): item
        for item in evidence if isinstance(item, dict)
    }
    matched = 0
    seated_students = 0
    covered_without_students = 0
    coverage = snapshot.get("class_date_coverage", {}) if isinstance(snapshot, dict) else {}
    coverage_start = str(coverage.get("start") or "")
    coverage_end = str(coverage.get("end") or "")
    for session in sessions:
        timing = session.get("timing", {}) if isinstance(session.get("timing"), dict) else {}
        course = session.get("course", {}) if isinstance(session.get("course"), dict) else {}
        start = session.get("start_at") or session.get("start") or timing.get("start_at") or timing.get("start")
        title = session.get("course_name") or session.get("mapped_clean_title") or course.get("mapped_clean_title") or course.get("course_name_primary_clean") or course.get("course_name_raw")
        item = by_key.get((str(start), course_key(title)))
        if not item:
            day = str(start)[:10]
            if coverage_start and coverage_end and coverage_start <= day <= coverage_end:
                capacity = session.setdefault("capacity", {})
                capacity["registered_count"] = 0
                capacity["students_count_raw"] = 0
                session["registered_count"] = 0
                session["enrolled_count"] = 0
                session["enrollment_evidence"] = {
                    "source": "enrollware_student_snapshot",
                    "status": "no_student_row_in_snapshot",
                    "as_of": snapshot.get("source_modified_at"),
                }
                covered_without_students += 1
            continue
        count = int(item.get("seated_count") or 0)
        capacity = session.setdefault("capacity", {})
        capacity["registered_count"] = count
        capacity["students_count_raw"] = count
        session["registered_count"] = count
        session["enrolled_count"] = count
        session["enrollment_evidence"] = {
            "source": "enrollware_student_snapshot",
            "status": "student_rows_present",
            "class_id": item.get("class_id"),
            "as_of": snapshot.get("source_modified_at"),
            "status_counts": item.get("status_counts", {}),
        }
        matched += 1
        seated_students += count
    return {
        "student_snapshot_classes": len(evidence),
        "student_snapshot_matches": matched,
        "seated_students_matched": seated_students,
        "covered_classes_without_students": covered_without_students,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Create a PII-free class enrollment snapshot from an Enrollware student export.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()
    source = Path(args.input).resolve()
    output = Path(args.output).resolve()
    payload = build_snapshot(
        read_rows(source),
        source_name=source.name,
        source_modified_at=datetime.fromtimestamp(source.stat().st_mtime, TZ).isoformat(),
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Imported {payload['counts']['student_rows']} student rows across {payload['counts']['classes_with_students']} classes.")
    print(f"Published PII-free snapshot: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
