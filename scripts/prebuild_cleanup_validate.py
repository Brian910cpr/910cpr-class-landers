from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ID_FROM_LINK_RE = re.compile(r"[?&]id=(\d+)")


def resolve_class_report_path(repo_root: Path, requested: str) -> Path:
    requested_path = (repo_root / requested).resolve()
    if requested_path.exists():
        return requested_path
    candidates = [
        repo_root / "data" / "Class Report.xlsx",
        repo_root / "data" / "raw" / "Class Report.xlsx",
        repo_root / "data" / "raw" / "class_report.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()
    return requested_path


def read_class_report_ids(path: Path) -> set[str]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return set()
    headers = [str(h or "").strip() for h in rows[0]]
    reg_idx = headers.index("Registration Link") if "Registration Link" in headers else None
    id_idx = headers.index("ID") if "ID" in headers else None
    out: set[str] = set()
    for row in rows[1:]:
        sid = ""
        if reg_idx is not None and reg_idx < len(row):
            reg_val = row[reg_idx]
            if reg_val:
                m = ID_FROM_LINK_RE.search(str(reg_val))
                if m:
                    sid = m.group(1)
        if not sid and id_idx is not None and id_idx < len(row):
            sid = str(row[id_idx] or "").strip()
        if sid:
            out.add(sid)
    return out


def clean_docs_data_json(docs_data_dir: Path) -> list[str]:
    deleted: list[str] = []
    if not docs_data_dir.exists():
        return deleted
    for path in docs_data_dir.glob("*.json"):
        path.unlink(missing_ok=True)
        deleted.append(str(path))
    return deleted


def clean_cached_schedule_files(repo_root: Path) -> list[str]:
    deleted: list[str] = []
    candidates = [
        repo_root / "data" / "runtime" / "build_schedule_future.json",
        repo_root / "data" / "runtime" / "build_schedule.json",
        repo_root / "data" / "runtime" / "build_sessions_current.json",
    ]
    for path in candidates:
        if path.exists():
            path.unlink(missing_ok=True)
            deleted.append(str(path))
    return deleted


def run_step(repo_root: Path, args: list[str]) -> None:
    print(f"Running: {' '.join(args)}")
    result = subprocess.run([sys.executable, *args], cwd=repo_root)
    if result.returncode != 0:
        raise SystemExit(result.returncode)


def load_session_ids(path: Path) -> set[str]:
    if not path.exists():
        return set()
    payload = json.loads(path.read_text(encoding="utf-8"))
    out: set[str] = set()
    for session in payload.get("sessions", []):
        sid = str(session.get("session_id") or "").strip()
        if sid:
            out.add(sid)
    return out


def write_report(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Pre-build cleanup and session integrity validation")
    parser.add_argument("--repo-root", default=".")
    parser.add_argument("--class-report", default="data/Class Report.xlsx")
    args = parser.parse_args()

    repo_root = Path(args.repo_root).resolve()
    class_report_path = resolve_class_report_path(repo_root, args.class_report)
    if not class_report_path.exists():
        raise SystemExit(f"Class report not found: {class_report_path}")

    deleted_docs_data = clean_docs_data_json(repo_root / "docs" / "data")
    deleted_cache = clean_cached_schedule_files(repo_root)
    print(f"Deleted docs/data JSON files: {len(deleted_docs_data)}")
    print(f"Deleted cached schedule files: {len(deleted_cache)}")

    run_step(repo_root, ["-m", "scripts.build_sessions_current", "--class-report", str(class_report_path)])
    run_step(
        repo_root,
        ["-m", "scripts.build_schedule_future", "--class-report", str(class_report_path)],
    )

    expected_ids = read_class_report_ids(class_report_path)
    built_ids = load_session_ids(repo_root / "data" / "sessions_current.json")
    future_ids = load_session_ids(repo_root / "docs" / "data" / "schedule_future.json")

    missing_sessions = sorted(expected_ids - built_ids)
    extra_sessions = sorted(built_ids - expected_ids)
    future_orphans = sorted(future_ids - expected_ids)

    for sid in extra_sessions + future_orphans:
        print(f"ORPHAN SESSION DETECTED: {sid}")

    report = {
        "timestamp": datetime.now().isoformat(),
        "class_report_path": str(class_report_path),
        "total_sessions_expected": len(expected_ids),
        "total_sessions_built": len(built_ids),
        "total_future_sessions_built": len(future_ids),
        "missing_sessions": missing_sessions,
        "extra_sessions": extra_sessions,
        "future_orphan_sessions": future_orphans,
        "deleted_docs_data_json": deleted_docs_data,
        "deleted_cached_schedule_files": deleted_cache,
    }
    report_path = repo_root / "debug" / "session_integrity_report.json"
    write_report(report_path, report)
    print(f"Wrote {report_path}")

    mismatch_count = len(missing_sessions) + len(extra_sessions) + len(future_orphans)
    if mismatch_count > 0:
        print(f"Integrity check FAILED: mismatch_count={mismatch_count}")
        return 2

    print("Integrity check PASSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
