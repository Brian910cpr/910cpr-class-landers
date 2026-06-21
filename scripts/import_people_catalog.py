from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit("openpyxl is required to import the instructor roster workbook.") from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK_NAME = "910CPR Instructor Roster - Allied 100 Transfer.xlsx"
PEOPLE_CATALOG_PATH = ROOT / "data" / "config" / "people_catalog.json"
IMPORT_REPORT_PATH = ROOT / "data" / "audit" / "people_catalog_import_report.md"
IMPORT_SUMMARY_PATH = ROOT / "data" / "audit" / "people_catalog_import_summary.json"
UNKNOWN = "UNKNOWN"

CREDENTIAL_COLUMNS = {
    "HS": "AHA_HEARTSAVER_INSTRUCTOR",
    "BLS": "AHA_BLS_INSTRUCTOR",
    "ACLS": "AHA_ACLS_INSTRUCTOR",
    "PALS": "AHA_PALS_INSTRUCTOR",
    "PEARS": "AHA_PEARS_INSTRUCTOR",
}

HEADER_ALIASES = {
    "name": ["name"],
    "email": ["email"],
    "phone": ["phone"],
    "instructor_id": ["instructor id", "instructor_id"],
    "job_association": ["job / association", "job association", "association"],
    "prior_tc": ["prior tc"],
    "hs": ["hs"],
    "bls": ["bls"],
    "acls": ["acls"],
    "pals": ["pals"],
    "pears": ["pears"],
    "transfer_status": ["transfer status"],
    "last_verified": ["last verified"],
    "alt_email": ["alt email", "alternate email"],
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean_text(value).lower()).strip()


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", clean_text(value).lower()).strip("_")


def serialize_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean_text(value)


def find_default_workbook() -> Path | None:
    candidates = [
        Path.cwd() / DEFAULT_WORKBOOK_NAME,
        ROOT / DEFAULT_WORKBOOK_NAME,
        Path.home() / "Downloads" / DEFAULT_WORKBOOK_NAME,
        Path("D:/Users/ten77/Downloads") / DEFAULT_WORKBOOK_NAME,
        Path("C:/Users/ten77/Downloads") / DEFAULT_WORKBOOK_NAME,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def resolve_workbook(path_text: str | None) -> Path:
    if path_text:
        path = Path(path_text).expanduser()
        if path.exists():
            return path
        raise FileNotFoundError(f"Workbook not found: {path}")
    found = find_default_workbook()
    if found:
        return found
    raise FileNotFoundError(f"Workbook not found in default locations: {DEFAULT_WORKBOOK_NAME}")


def find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    best_index = -1
    best_score = 0
    best_headers: dict[str, int] = {}
    alias_lookup = {
        alias: field
        for field, aliases in HEADER_ALIASES.items()
        for alias in aliases
    }
    for index, row in enumerate(rows[:25]):
        headers: dict[str, int] = {}
        for col_index, value in enumerate(row):
            normalized = normalize_header(value)
            if normalized in alias_lookup:
                headers[alias_lookup[normalized]] = col_index
        score = len(headers)
        if score > best_score:
            best_index = index
            best_score = score
            best_headers = headers
    if best_index < 0 or "name" not in best_headers:
        raise ValueError("Could not identify a usable header row with a Name column.")
    return best_index, best_headers


def cell(row: tuple[Any, ...], headers: dict[str, int], field: str) -> Any:
    index = headers.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def split_emails(value: Any) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    parts = re.split(r"[/,;]|\s+and\s+", text, flags=re.I)
    emails = []
    for part in parts:
        part = part.strip()
        if "@" in part:
            emails.append(part.lower())
    return emails


def person_id_for(name: str, email: str, instructor_id: str) -> str:
    if instructor_id:
        return f"instructor_{normalize_key(instructor_id)}"
    if email:
        return f"person_{normalize_key(email.split('@')[0])}"
    return f"person_{normalize_key(name)}"


def is_positive_credential(value: Any) -> bool:
    if value is None or clean_text(value) == "":
        return False
    if isinstance(value, (datetime, date)):
        return True
    text = clean_text(value).lower()
    return text in {"✓", "✔", "x", "yes", "y", "true", "1", "complete", "completed"} or bool(
        re.fullmatch(r"\d{4}-\d{2}-\d{2}(?: 00:00:00)?", text)
    )


def is_negative_or_blank(value: Any) -> bool:
    if value is None or clean_text(value) == "":
        return True
    return clean_text(value).lower() in {"no", "n", "false", "0", "none", "na", "n/a"}


def certification_from_cell(column_name: str, value: Any, row_number: int) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    code = CREDENTIAL_COLUMNS[column_name]
    if is_positive_credential(value):
        observed = serialize_value(value)
        expiration = observed if isinstance(value, (datetime, date)) or re.fullmatch(r"\d{4}-\d{2}-\d{2}(?: 00:00:00)?", observed) else UNKNOWN
        return {
            "certification_code": code,
            "provider": "AHA",
            "family": column_name,
            "expiration_date": expiration,
            "source": "910CPR Instructor Roster - Allied 100 Transfer.xlsx",
            "source_column": column_name,
            "source_row": row_number,
            "observed_value": observed or "checked",
        }, None
    if is_negative_or_blank(value):
        return None, None
    return None, {
        "row": row_number,
        "column": column_name,
        "value": serialize_value(value),
        "message": "Credential value was neither blank/negative nor a recognized positive marker.",
    }


def normalize_people(workbook_path: Path) -> tuple[dict[str, Any], dict[str, Any], str]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_index, headers = find_header_row(rows)
    header_row_number = header_index + 1
    people: list[dict[str, Any]] = []
    unclear_values: list[dict[str, Any]] = []
    email_counter: Counter[str] = Counter()
    name_counter: Counter[str] = Counter()
    cert_counter: Counter[str] = Counter()
    missing_emails: list[dict[str, str]] = []

    for row_index, row in enumerate(rows[header_index + 1 :], start=header_row_number + 1):
        name = clean_text(cell(row, headers, "name"))
        if not name:
            continue
        email_values = split_emails(cell(row, headers, "email"))
        alt_email_values = split_emails(cell(row, headers, "alt_email"))
        all_emails = []
        for email in [*email_values, *alt_email_values]:
            if email not in all_emails:
                all_emails.append(email)
        primary_email = all_emails[0] if all_emails else UNKNOWN
        if primary_email == UNKNOWN:
            missing_emails.append({"row": str(row_index), "display_name": name})
        for email in all_emails:
            email_counter[email] += 1
        name_counter[name.lower()] += 1

        instructor_id = clean_text(cell(row, headers, "instructor_id"))
        certifications: list[dict[str, Any]] = []
        for column_name in CREDENTIAL_COLUMNS:
            field = column_name.lower()
            if field not in headers:
                continue
            certification, unclear = certification_from_cell(column_name, cell(row, headers, field), row_index)
            if certification:
                certifications.append(certification)
                cert_counter[certification["certification_code"]] += 1
            if unclear:
                unclear["display_name"] = name
                unclear_values.append(unclear)

        person = {
            "person_id": person_id_for(name, primary_email if primary_email != UNKNOWN else "", instructor_id),
            "display_name": name,
            "email": primary_email,
            "phone": clean_text(cell(row, headers, "phone")) or UNKNOWN,
            "external_ids": {
                "instructor_id": instructor_id or UNKNOWN,
                "prior_tc": clean_text(cell(row, headers, "prior_tc")) or UNKNOWN,
                "aligned_tc": UNKNOWN,
            },
            "relationship_type": UNKNOWN,
            "scheduler_status": UNKNOWN,
            "active_for_dynamic_scheduling": False,
            "job_association": clean_text(cell(row, headers, "job_association")) or UNKNOWN,
            "transfer_status": clean_text(cell(row, headers, "transfer_status")) or UNKNOWN,
            "last_verified": serialize_value(cell(row, headers, "last_verified")) or UNKNOWN,
            "alternate_emails": all_emails[1:],
            "certifications": certifications,
            "source": {
                "workbook": str(workbook_path),
                "worksheet": ws.title,
                "row": row_index,
            },
        }
        people.append(person)

    duplicate_emails = [{"email": email, "count": count} for email, count in sorted(email_counter.items()) if count > 1]
    duplicate_names = [{"display_name": name, "count": count} for name, count in sorted(name_counter.items()) if count > 1]

    catalog = {
        "schema_version": "0.1",
        "description": "Read-only People/Instructor qualification model imported from Brian's instructor roster workbook. This does not modify public pages, Enrollware behavior, appointment URLs, Worker behavior, or solver behavior.",
        "unknown_value": UNKNOWN,
        "source_workbook": str(workbook_path),
        "worksheets": sheet_names,
        "header_row": header_row_number,
        "mapped_columns": {field: headers[field] + 1 for field in sorted(headers)},
        "people": people,
    }
    summary = {
        "source_workbook": str(workbook_path),
        "worksheets": sheet_names,
        "primary_worksheet": ws.title,
        "header_row": header_row_number,
        "mapped_columns": {field: headers[field] + 1 for field in sorted(headers)},
        "rows_processed": len(people),
        "people_created": len(people),
        "duplicate_emails": duplicate_emails,
        "duplicate_names": duplicate_names,
        "certifications_found_by_type": dict(sorted(cert_counter.items())),
        "missing_emails": missing_emails,
        "unclear_credential_values": unclear_values,
        "active_for_dynamic_scheduling_count": 0,
        "public_site_affected": False,
        "enrollware_behavior_changed": False,
        "appointment_urls_changed": False,
        "worker_creation_enabled": False,
        "recommended_next_step": "Brian should review people_catalog.json and confirm which people should be active for dynamic scheduling before any solver integration.",
    }
    return catalog, summary, ws.title


def render_report(summary: dict[str, Any]) -> str:
    lines = [
        "# People Catalog Import Report",
        "",
        "This is a read-only People/Instructor qualification import. It did not modify public pages, Enrollware behavior, appointment URLs, Worker behavior, generated HTML, or solver behavior.",
        "",
        "## Workbook Inspection",
        "",
        f"- Source workbook: `{summary['source_workbook']}`",
        f"- Worksheets: {', '.join(f'`{name}`' for name in summary['worksheets'])}",
        f"- Imported worksheet: `{summary['primary_worksheet']}`",
        f"- Header row: {summary['header_row']}",
        "",
        "## Column Mapping",
        "",
        "| Field | Column Number |",
        "| --- | ---: |",
    ]
    for field, col in summary["mapped_columns"].items():
        lines.append(f"| `{field}` | {col} |")

    lines.extend([
        "",
        "## Import Summary",
        "",
        f"- Rows processed: {summary['rows_processed']}",
        f"- People created: {summary['people_created']}",
        f"- Duplicate emails: {len(summary['duplicate_emails'])}",
        f"- Duplicate names: {len(summary['duplicate_names'])}",
        f"- Missing emails: {len(summary['missing_emails'])}",
        f"- Unclear credential values: {len(summary['unclear_credential_values'])}",
        f"- Active for dynamic scheduling: {summary['active_for_dynamic_scheduling_count']}",
        "",
        "## Certifications Found By Type",
        "",
    ])
    if summary["certifications_found_by_type"]:
        lines.extend(["| Certification Code | Count |", "| --- | ---: |"])
        for code, count in summary["certifications_found_by_type"].items():
            lines.append(f"| `{code}` | {count} |")
    else:
        lines.append("- None")

    lines.extend(["", "## Duplicate Emails", ""])
    if summary["duplicate_emails"]:
        lines.extend(["| Email | Count |", "| --- | ---: |"])
        for item in summary["duplicate_emails"]:
            lines.append(f"| {item['email']} | {item['count']} |")
    else:
        lines.append("- None")

    lines.extend(["", "## Duplicate Names", ""])
    if summary["duplicate_names"]:
        lines.extend(["| Display Name | Count |", "| --- | ---: |"])
        for item in summary["duplicate_names"]:
            lines.append(f"| {item['display_name']} | {item['count']} |")
    else:
        lines.append("- None")

    lines.extend(["", "## Missing Emails", ""])
    if summary["missing_emails"]:
        lines.extend(["| Row | Display Name |", "| ---: | --- |"])
        for item in summary["missing_emails"]:
            lines.append(f"| {item['row']} | {item['display_name']} |")
    else:
        lines.append("- None")

    lines.extend(["", "## Unclear Credential Values", ""])
    if summary["unclear_credential_values"]:
        lines.extend(["| Row | Name | Column | Value | Message |", "| ---: | --- | --- | --- | --- |"])
        for item in summary["unclear_credential_values"]:
            lines.append(f"| {item['row']} | {item.get('display_name', UNKNOWN)} | {item['column']} | {item['value']} | {item['message']} |")
    else:
        lines.append("- None")

    lines.extend([
        "",
        "## Recommended Next Step",
        "",
        f"- {summary['recommended_next_step']}",
        "- Do not activate anyone for dynamic scheduling until Brian confirms scheduler eligibility separately from credential presence.",
        "",
    ])
    return "\n".join(lines)


def run_import(workbook_path: Path) -> dict[str, Any]:
    catalog, summary, _ = normalize_people(workbook_path)
    PEOPLE_CATALOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    IMPORT_REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    PEOPLE_CATALOG_PATH.write_text(json.dumps(catalog, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    IMPORT_SUMMARY_PATH.write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    IMPORT_REPORT_PATH.write_text(render_report(summary), encoding="utf-8")
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Brian's instructor roster workbook into a read-only People catalog.")
    parser.add_argument("workbook", nargs="?", help="Path to 910CPR Instructor Roster - Allied 100 Transfer.xlsx")
    args = parser.parse_args()
    workbook_path = resolve_workbook(args.workbook)
    summary = run_import(workbook_path)
    print("People catalog import complete (READ ONLY).")
    print("No public pages, Enrollware behavior, appointment URLs, Worker settings, generated HTML, or solver behavior were changed.")
    print("")
    print(f"Workbook: {summary['source_workbook']}")
    print(f"Worksheets: {', '.join(summary['worksheets'])}")
    print(f"Header row: {summary['header_row']}")
    print(f"Rows processed: {summary['rows_processed']}")
    print(f"People created: {summary['people_created']}")
    print(f"Duplicate emails: {len(summary['duplicate_emails'])}")
    print(f"Duplicate names: {len(summary['duplicate_names'])}")
    print(f"Missing emails: {len(summary['missing_emails'])}")
    print(f"Unclear credential values: {len(summary['unclear_credential_values'])}")
    print("")
    print("Output files:")
    print(f"- {PEOPLE_CATALOG_PATH}")
    print(f"- {IMPORT_REPORT_PATH}")
    print(f"- {IMPORT_SUMMARY_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
