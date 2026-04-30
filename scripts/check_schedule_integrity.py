from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ID_FROM_LINK_RE = re.compile(r"[?&]id=(\d+)")
TZ_SUFFIX_RE = re.compile(r"(Z|[+-]\d{2}:\d{2})$")


def resolve_class_report_path(repo_root: Path) -> Path:
    candidates = [
        repo_root / "data" / "Class Report.xlsx",
        repo_root / "data" / "raw" / "Class Report.xlsx",
        repo_root / "data" / "raw" / "class_report.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def extract_session_id(registration_link: Any, fallback_id: Any) -> str:
    sid = ""
    if registration_link:
        m = ID_FROM_LINK_RE.search(str(registration_link))
        if m:
            sid = m.group(1).strip()
    if not sid and fallback_id:
        sid = str(fallback_id).strip()
    return sid


def parse_report_dt(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    formats = [
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw)
    except ValueError:
        return None


def parse_iso_dt(value: Any) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    if raw.endswith("Z"):
        raw = raw[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(raw)
        return dt.replace(tzinfo=None) if dt.tzinfo else dt
    except ValueError:
        return None


def load_class_report_sessions(class_report_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(filename=class_report_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(h or "").strip() for h in rows[0]]
    idx = {name: i for i, name in enumerate(headers)}
    reg_idx = idx.get("Registration Link")
    id_idx = idx.get("ID")
    start_idx = idx.get("Start Date / Time")
    end_idx = idx.get("End Date / Time")

    sessions: dict[str, dict[str, Any]] = {}
    for row in rows[1:]:
        reg_val = row[reg_idx] if reg_idx is not None and reg_idx < len(row) else None
        id_val = row[id_idx] if id_idx is not None and id_idx < len(row) else None
        sid = extract_session_id(reg_val, id_val)
        if not sid:
            continue
        start_val = row[start_idx] if start_idx is not None and start_idx < len(row) else None
        end_val = row[end_idx] if end_idx is not None and end_idx < len(row) else None
        sessions[sid] = {
            "start": parse_report_dt(start_val),
            "end": parse_report_dt(end_val),
            "raw_start": str(start_val or "").strip(),
            "raw_end": str(end_val or "").strip(),
        }
    return sessions


def load_schedule_future_sessions(schedule_future_path: Path) -> dict[str, dict[str, Any]]:
    if not schedule_future_path.exists():
        return {}
    payload = json.loads(schedule_future_path.read_text(encoding="utf-8"))
    items = payload.get("sessions", []) if isinstance(payload, dict) else []
    sessions: dict[str, dict[str, Any]] = {}
    for item in items:
        if not isinstance(item, dict):
            continue
        sid = str(item.get("session_id") or "").strip()
        if not sid:
            continue
        sessions[sid] = {
            "start": parse_iso_dt(item.get("start_at")),
            "end": parse_iso_dt(item.get("end_at")),
            "raw_start": str(item.get("start_at") or "").strip(),
            "raw_end": str(item.get("end_at") or "").strip(),
        }
    return sessions


def fmt_dt(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S")


def main() -> int:
    repo_root = Path(__file__).resolve().parents[1]
    class_report_path = resolve_class_report_path(repo_root)
    schedule_future_path = repo_root / "docs" / "data" / "schedule_future.json"

    if not class_report_path.exists():
        print(f"FAIL: Class Report not found: {class_report_path}")
        return 1
    if not schedule_future_path.exists():
        print(f"FAIL: schedule_future.json not found: {schedule_future_path}")
        return 1

    report_sessions = load_class_report_sessions(class_report_path)
    schedule_sessions = load_schedule_future_sessions(schedule_future_path)

    report_ids = set(report_sessions.keys())
    schedule_ids = set(schedule_sessions.keys())

    missing_sessions = sorted(report_ids - schedule_ids)
    extra_sessions = sorted(schedule_ids - report_ids)

    mismatched_times: list[dict[str, str]] = []
    for sid in sorted(report_ids & schedule_ids):
        report_start = report_sessions[sid]["start"]
        report_end = report_sessions[sid]["end"]
        sched_start = schedule_sessions[sid]["start"]
        sched_end = schedule_sessions[sid]["end"]
        if report_start is None or report_end is None or sched_start is None or sched_end is None:
            continue
        if report_start != sched_start or report_end != sched_end:
            mismatched_times.append(
                {
                    "session_id": sid,
                    "class_report_start": fmt_dt(report_start),
                    "class_report_end": fmt_dt(report_end),
                    "schedule_start": fmt_dt(sched_start),
                    "schedule_end": fmt_dt(sched_end),
                }
            )

    print("=== Schedule Integrity Check ===")
    print(f"Class Report path: {class_report_path}")
    print(f"schedule_future path: {schedule_future_path}")
    print(f"Class Report sessions: {len(report_ids)}")
    print(f"schedule_future sessions: {len(schedule_ids)}")
    print(f"Missing sessions: {len(missing_sessions)}")
    print(f"Extra sessions: {len(extra_sessions)}")
    print(f"Mismatched dates/times: {len(mismatched_times)}")

    if missing_sessions:
        print("\nMissing session IDs (in Class Report, not in schedule_future):")
        for sid in missing_sessions:
            print(f"- {sid}")

    if extra_sessions:
        print("\nExtra session IDs (in schedule_future, not in Class Report):")
        for sid in extra_sessions:
            print(f"- {sid}")

    if mismatched_times:
        print("\nMismatched dates/times:")
        for row in mismatched_times:
            print(
                f"- {row['session_id']}: "
                f"ClassReport {row['class_report_start']} -> {row['class_report_end']} | "
                f"Schedule {row['schedule_start']} -> {row['schedule_end']}"
            )

    if missing_sessions or extra_sessions or mismatched_times:
        print("\nFAIL: Schedule integrity check found mismatches.")
        return 2

    print("\nPASS: schedule_future.json matches Class Report session IDs and date/time values.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
