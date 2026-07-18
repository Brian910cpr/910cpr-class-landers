from __future__ import annotations

import argparse
import csv
import html
import hashlib
import json
import os
import re
import ssl
import urllib.error
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from scripts.course_identity_resolver import normalize_title
from scripts.build_status import BuildStatusReporter
from scripts.local_data_paths import missing_live_input_message, print_resolved_path, resolve_live_input_path


TZ = ZoneInfo("America/New_York")
DESCRIPTION_FIELDS = [
    "short_description",
    "long_description",
    "who_class_for",
    "prerequisites",
    "what_to_expect",
    "certification_card",
    "renewal_info",
    "official_description_source",
    "description_status",
]
DESCRIPTION_REQUIRED_FIELDS = [
    "short_description",
    "long_description",
    "who_class_for",
]
DESCRIPTION_PLACEHOLDER_PATTERN = re.compile(
    r"^(tbd|todo|placeholder|lorem ipsum|coming soon|n/?a|na|unknown)$",
    flags=re.IGNORECASE,
)


CLASS_REPORT_LEGACY_PATHS = [
    "data/Class Report.xlsx",
    "data/raw/Class Report.xlsx",
    "data/raw/class_report.xlsx",
    "Class Report (37).xlsx",
    "class-report.xlsx",
]
CLASSES_CSV_LEGACY_PATHS = ["data/raw/classes_raw_live.csv"]
STUDENTS_CSV_LEGACY_PATHS = ["data/raw/students_raw_live.csv"]
DEFAULT_ENROLLWARE_ICAL_URL = "https://www.enrollware.com/calendar/ical.ashx?Bsd2bwNowUkB4RQAmjnCBA=="
ENROLLWARE_ICAL_USER_AGENT = "910CPR-Lander-Build/1.0 (+https://www.910cpr.com)"
_COURSE_CONSUMPTION_RULES_CACHE: dict[str, int] | None = None


def clean_string(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def clean_whitespace(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    return re.sub(r"\s+", " ", value).strip() or None


def normalize_description_status(value: Optional[str]) -> str:
    status = clean_string(value)
    if not status:
        return "missing"
    status = status.strip().lower()
    if status in {"official", "draft", "needs_review", "missing"}:
        return status
    if status in {"need_review", "needs review"}:
        return "needs_review"
    return "needs_review"


def is_placeholder_text(value: Optional[str]) -> bool:
    text = clean_whitespace(value)
    if not text:
        return False
    return bool(DESCRIPTION_PLACEHOLDER_PATTERN.match(text))


def parse_int(value: Any) -> Optional[int]:
    s = clean_string(value)
    if s is None:
        return None
    s = s.replace(",", "").strip()
    if s == "":
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def parse_float(value: Any) -> Optional[float]:
    s = clean_string(value)
    if s is None:
        return None
    s = s.replace("$", "").replace(",", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None


def strip_tags(text: Optional[str]) -> Optional[str]:
    if text is None:
        return None
    no_tags = re.sub(r"<[^>]+>", " ", text)
    no_tags = html.unescape(no_tags)
    return clean_whitespace(no_tags)


def extract_enroll_id_from_url(url: Optional[str]) -> Optional[str]:
    if not url:
        return None
    m = re.search(r"[?&]id=(\d+)", url)
    if m:
        return m.group(1)
    return None


def parse_datetime_flexible(value: Any) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, datetime):
        dt = value
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=TZ)
        else:
            dt = dt.astimezone(TZ)
        return dt.isoformat()

    s = clean_string(value)
    if s is None:
        return None

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%y %I:%M %p",
        "%m/%d/%y %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt).replace(tzinfo=TZ)
            return dt.isoformat()
        except Exception:
            continue

    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=TZ)
        else:
            dt = dt.astimezone(TZ)
        return dt.isoformat()
    except Exception:
        return None


def unfold_ical_lines(text: str) -> list[str]:
    unfolded: list[str] = []
    for raw_line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        if not raw_line:
            continue
        if raw_line.startswith((" ", "\t")) and unfolded:
            unfolded[-1] += raw_line[1:]
        else:
            unfolded.append(raw_line)
    return unfolded


def parse_ical_property(line: str) -> tuple[str, dict[str, str], str] | None:
    if ":" not in line:
        return None
    name_and_params, value = line.split(":", 1)
    parts = name_and_params.split(";")
    name = parts[0].upper()
    params: dict[str, str] = {}
    for part in parts[1:]:
        if "=" not in part:
            continue
        key, param_value = part.split("=", 1)
        params[key.upper()] = param_value.strip('"')
    return name, params, value


def ical_unescape(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\\n", "\n").replace("\\N", "\n")
    text = text.replace("\\,", ",").replace("\\;", ";").replace("\\\\", "\\")
    return html.unescape(text).strip()


def parse_ical_datetime(value: str, params: dict[str, str]) -> Optional[str]:
    raw = clean_string(value)
    if not raw:
        return None
    raw = raw.strip()
    try:
        if re.fullmatch(r"\d{8}", raw):
            dt = datetime.strptime(raw, "%Y%m%d").replace(tzinfo=TZ)
        elif raw.endswith("Z"):
            dt = datetime.strptime(raw, "%Y%m%dT%H%M%SZ").replace(tzinfo=timezone.utc).astimezone(TZ)
        else:
            fmt = "%Y%m%dT%H%M%S" if len(raw) >= 15 else "%Y%m%dT%H%M"
            dt = datetime.strptime(raw, fmt)
            dt = dt.replace(tzinfo=TZ)
        return dt.isoformat()
    except Exception:
        return None


def parse_ical_events(text: str) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for line in unfold_ical_lines(text):
        if line == "BEGIN:VEVENT":
            current = {"raw_properties": {}}
            continue
        if line == "END:VEVENT":
            if current is not None:
                events.append(current)
            current = None
            continue
        if current is None:
            continue
        parsed = parse_ical_property(line)
        if not parsed:
            continue
        name, params, value = parsed
        current["raw_properties"][name] = {"params": params, "value": value}
        if name in {"SUMMARY", "DESCRIPTION", "LOCATION", "UID", "URL", "STATUS", "DTSTAMP", "LAST-MODIFIED"}:
            current[name.lower().replace("-", "_")] = ical_unescape(value)
        elif name in {"DTSTART", "DTEND"}:
            current[name.lower()] = parse_ical_datetime(value, params)
            current[f"{name.lower()}_raw"] = value
            current[f"{name.lower()}_params"] = params
    return events


def fetch_ical_text(url: str) -> str:
    request = urllib.request.Request(url, headers={"User-Agent": ENROLLWARE_ICAL_USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=45) as response:
            raw = response.read()
    except urllib.error.URLError as exc:
        reason = getattr(exc, "reason", None)
        if not isinstance(reason, ssl.SSLError):
            raise
        insecure_context = ssl._create_unverified_context()
        with urllib.request.urlopen(request, timeout=45, context=insecure_context) as response:
            raw = response.read()
    return decode_ical_bytes(raw)


def fetch_enrollment_page_text(url: str) -> str:
    request = urllib.request.Request(url, headers={"User-Agent": ENROLLWARE_ICAL_USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            raw = response.read()
    except urllib.error.URLError as exc:
        reason = getattr(exc, "reason", None)
        if not isinstance(reason, ssl.SSLError):
            raise
        insecure_context = ssl._create_unverified_context()
        with urllib.request.urlopen(request, timeout=20, context=insecure_context) as response:
            raw = response.read()
    return decode_ical_bytes(raw)


def decode_ical_bytes(raw: bytes) -> str:
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError:
        return raw.decode("cp1252", errors="replace")


def enrollment_page_unavailable_reason(text: str) -> Optional[str]:
    visible_text = re.sub(r"<script[\s\S]*?</script>", " ", text, flags=re.IGNORECASE)
    visible_text = re.sub(r"<style[\s\S]*?</style>", " ", visible_text, flags=re.IGNORECASE)
    visible_text = re.sub(r"<[^>]+>", " ", visible_text)
    visible_text = html.unescape(re.sub(r"\s+", " ", visible_text)).strip().lower()
    if "registration for this class is closed" in visible_text:
        return "enrollware_registration_closed"
    if "this class is currently full" in visible_text:
        return "enrollware_class_full"
    return None


def enrollment_url_unavailable_reason(url: Optional[str]) -> Optional[str]:
    clean_url = clean_string(url)
    if not clean_url:
        return None
    try:
        return enrollment_page_unavailable_reason(fetch_enrollment_page_text(clean_url))
    except Exception as exc:
        return f"enrollware_registration_status_check_failed:{exc.__class__.__name__}"


def registration_status_from_unavailable_reason(reason: Optional[str]) -> str:
    if reason == "enrollware_registration_closed":
        return "closed"
    if reason == "enrollware_class_full":
        return "full"
    return "open"


def course_consumption_minutes(course_map: dict[str, Any], course_id: Optional[str]) -> Optional[int]:
    global _COURSE_CONSUMPTION_RULES_CACHE
    clean_course_id = clean_string(course_id)
    if not clean_course_id:
        return None
    if _COURSE_CONSUMPTION_RULES_CACHE is None:
        rules: dict[str, int] = {}
        course_map_path = clean_string(course_map.get("_path"))
        if course_map_path:
            rules_path = Path(course_map_path).resolve().parents[1] / "inventory" / "course_consumption_rules.json"
            if rules_path.exists():
                payload = json.loads(rules_path.read_text(encoding="utf-8"))
                defaults = payload.get("defaults", {}) if isinstance(payload, dict) and isinstance(payload.get("defaults"), dict) else {}
                for raw_rule in payload.get("rules", []) if isinstance(payload, dict) else []:
                    if not isinstance(raw_rule, dict):
                        continue
                    rule = {**defaults, **raw_rule}
                    rule_course_id = clean_string(rule.get("course_id"))
                    if not rule_course_id:
                        continue
                    duration = parse_int(rule.get("duration_minutes")) or 0
                    setup = parse_int(rule.get("setup_buffer_minutes")) or 0
                    cleanup = parse_int(rule.get("cleanup_buffer_minutes")) or 0
                    minimum = parse_int(rule.get("minimum_reservation_block_minutes")) or 0
                    minutes = max(duration + setup + cleanup, minimum)
                    if minutes > 0:
                        rules[rule_course_id] = minutes
        _COURSE_CONSUMPTION_RULES_CACHE = rules
    return _COURSE_CONSUMPTION_RULES_CACHE.get(clean_course_id)


def inferred_ical_end(
    *,
    start_at: Optional[str],
    end_at: Optional[str],
    course_map: dict[str, Any],
    course_id: Optional[str],
) -> tuple[Optional[str], Optional[str], Optional[int]]:
    if not start_at:
        return end_at, None, None
    start_iso = parse_datetime_flexible(start_at)
    end_iso = parse_datetime_flexible(end_at)
    if start_iso and end_iso and end_iso > start_iso:
        return end_at, None, None
    minutes = course_consumption_minutes(course_map, course_id)
    if not minutes:
        return end_at, "missing_course_consumption_rule_for_zero_duration_ical_event", None
    start = datetime.fromisoformat(start_iso)
    return (start + timedelta(minutes=minutes)).isoformat(), "inferred_from_course_consumption_rule_for_zero_duration_ical_event", minutes


def extract_lead_instructor(description: Optional[str]) -> Optional[str]:
    text = ical_unescape(description)
    if not text:
        return None
    match = re.search(r"Instructors:\s*(.*?)(?:\n\s*\n|Directions:|$)", text, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        return None
    lines = [clean_whitespace(re.sub(r"\(lead\)", "", line, flags=re.IGNORECASE)) for line in match.group(1).splitlines()]
    lines = [line for line in lines if line]
    for line in lines:
        if "(lead)" in line.lower():
            return clean_whitespace(re.sub(r"\(lead\)", "", line, flags=re.IGNORECASE))
    return lines[0] if lines else None


def stable_ical_session_id(event: dict[str, Any]) -> str:
    uid = clean_string(event.get("uid"))
    if uid:
        return uid
    key = "|".join(
        clean_string(event.get(name)) or ""
        for name in ["summary", "dtstart", "dtend", "location", "url"]
    )
    return "ical_" + hashlib.sha1(key.encode("utf-8")).hexdigest()[:16]


def previous_sessions_by_id(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    sessions = payload.get("sessions", []) if isinstance(payload, dict) else []
    return {
        str(session.get("session_id")): session
        for session in sessions
        if isinstance(session, dict) and session.get("session_id") is not None
    }


def build_session_from_ical_event(
    event: dict[str, Any],
    now_iso: str,
    course_map: dict[str, Any],
    previous_session: Optional[dict[str, Any]] = None,
) -> Optional[dict[str, Any]]:
    session_id = stable_ical_session_id(event)
    start_at = clean_string(event.get("dtstart"))
    end_at = clean_string(event.get("dtend"))
    if not session_id or not start_at:
        return None

    raw_summary = clean_string(event.get("summary"))
    course_title = strip_tags(raw_summary)
    parsed = parse_course_blob(course_title)
    mapped, mapping_status, mapping_notes = resolve_course_mapping(
        course_map,
        course_id=None,
        course_number=None,
        raw_course_title=parsed.course_name_primary_clean or course_title,
    )
    course_id = clean_string(mapped.get("course_id")) if mapped else None
    course_number = clean_string(mapped.get("course_number") or course_id) if mapped else None
    end_at, end_inference_reason, inferred_consumption_minutes = inferred_ical_end(
        start_at=start_at,
        end_at=end_at,
        course_map=course_map,
        course_id=course_id,
    )

    mapped_family = clean_string(mapped.get("family")) if mapped else None
    mapped_subtype = clean_string(mapped.get("subtype")) if mapped else None
    mapped_certifying_body = clean_string(mapped.get("certifying_body")) if mapped else None
    mapped_delivery_mode = clean_string(mapped.get("delivery_mode")) if mapped else None
    mapped_course_code = clean_string(mapped.get("course_code")) if mapped else None
    mapped_logo_key = clean_string(mapped.get("logo_key")) if mapped else None
    mapped_price = parse_float(mapped.get("price")) if mapped else None
    mapped_clean_title = clean_string(mapped.get("clean_title")) if mapped else None
    mapped_short_description = clean_string(mapped.get("short_description")) if mapped else None
    mapped_long_description = clean_string(mapped.get("long_description")) if mapped else None
    mapped_who_class_for = clean_string(mapped.get("who_class_for") or mapped.get("who_for")) if mapped else None
    mapped_prerequisites = clean_string(mapped.get("prerequisites")) if mapped else None
    mapped_what_to_expect = clean_string(mapped.get("what_to_expect")) if mapped else None
    mapped_certification_card = clean_string(mapped.get("certification_card")) if mapped else None
    mapped_renewal_info = clean_string(mapped.get("renewal_info")) if mapped else None
    mapped_official_description_source = clean_string(mapped.get("official_description_source")) if mapped else None
    mapped_description_status = normalize_description_status(mapped.get("description_status") if mapped else None)

    previous_capacity = previous_session.get("capacity", {}) if isinstance(previous_session, dict) and isinstance(previous_session.get("capacity"), dict) else {}
    previous_commerce = previous_session.get("commerce", {}) if isinstance(previous_session, dict) and isinstance(previous_session.get("commerce"), dict) else {}
    registration_url = clean_string(event.get("url")) or clean_string(previous_commerce.get("registration_url"))

    location_name = clean_string(event.get("location"))
    lead_instructor_name = extract_lead_instructor(clean_string(event.get("description")))

    return {
        "session_id": session_id,
        "source": "enrollware_ical",
        "start": start_at,
        "end": end_at,
        "start_datetime": start_at,
        "end_datetime": end_at,
        "source_keys": {
            "enrollware_ical_uid": clean_string(event.get("uid")),
            "enrollware_ical_url": registration_url,
        },
        "course": {
            "course_id": course_id,
            "course_name_raw": raw_summary,
            "course_name_primary_raw": course_title,
            "course_name_primary_clean": parsed.course_name_primary_clean or course_title,
            "course_tail_raw": parsed.course_tail_raw,
            "course_subtitle_text": parsed.course_subtitle_text,
            "course_number": course_number,
            "course_code_hint": parsed.course_code_hint or mapped_course_code,
            "certifying_body_hint": parsed.certifying_body_hint or mapped_certifying_body,
            "source_hint": "enrollware_ical",
            "price_hint": parsed.price_hint,
            "delivery_mode_hint": parsed.delivery_mode_hint or mapped_delivery_mode,
            "image_src": parsed.image_src,
            "parse_confidence": parsed.parse_confidence,
            "parse_notes": parsed.parse_notes,
            "mapped_family": mapped_family,
            "mapped_subtype": mapped_subtype,
            "mapped_certifying_body": mapped_certifying_body,
            "mapped_delivery_mode": mapped_delivery_mode,
            "mapped_logo_key": mapped_logo_key,
            "mapped_price": mapped_price,
            "mapped_clean_title": mapped_clean_title,
            "mapped_short_description": mapped_short_description,
            "mapped_long_description": mapped_long_description,
            "mapped_who_class_for": mapped_who_class_for,
            "mapped_prerequisites": mapped_prerequisites,
            "mapped_what_to_expect": mapped_what_to_expect,
            "mapped_certification_card": mapped_certification_card,
            "mapped_renewal_info": mapped_renewal_info,
            "mapped_official_description_source": mapped_official_description_source,
            "mapped_description_status": mapped_description_status,
            "mapping_status": mapping_status,
            "mapping_notes": mapping_notes,
        },
        "mapped_family": mapped_family,
        "mapped_subtype": mapped_subtype,
        "mapped_certifying_body": mapped_certifying_body,
        "mapped_delivery_mode": mapped_delivery_mode,
        "mapped_logo_key": mapped_logo_key,
        "mapped_price": mapped_price,
        "mapped_clean_title": mapped_clean_title,
        "mapped_short_description": mapped_short_description,
        "mapped_long_description": mapped_long_description,
        "mapped_who_class_for": mapped_who_class_for,
        "mapped_prerequisites": mapped_prerequisites,
        "mapped_what_to_expect": mapped_what_to_expect,
        "mapped_certification_card": mapped_certification_card,
        "mapped_renewal_info": mapped_renewal_info,
        "mapped_official_description_source": mapped_official_description_source,
        "mapped_description_status": mapped_description_status,
        "mapping_status": mapping_status,
        "mapping_notes": mapping_notes,
        "timing": {
            "start_at": start_at,
            "end_at": end_at,
            "start": start_at,
            "end": end_at,
            "start_datetime": start_at,
            "end_datetime": end_at,
            "start_source": clean_string(event.get("dtstart_raw")),
            "end_source": clean_string(event.get("dtend_raw")),
            "end_inference_reason": end_inference_reason,
            "inferred_scheduler_consumption_minutes": inferred_consumption_minutes,
            "timezone": "America/New_York",
            "is_future": parse_datetime_flexible(start_at) >= now_iso if parse_datetime_flexible(start_at) else None,
            "is_past": parse_datetime_flexible(start_at) < now_iso if parse_datetime_flexible(start_at) else None,
        },
        "location": {
            "location_name": location_name,
            "location_display": location_name,
            "client": None,
        },
        "staffing": {
            "lead_instructor_name": lead_instructor_name,
        },
        "capacity": {
            "max_students": previous_capacity.get("max_students"),
            "students_count_raw": previous_capacity.get("students_count_raw"),
            "registered_count": previous_capacity.get("registered_count"),
            "available_seats": previous_capacity.get("available_seats"),
            "is_full": previous_capacity.get("is_full"),
        },
        "commerce": {
            "price": mapped_price if mapped_price is not None else previous_commerce.get("price"),
            "registration_url": registration_url,
        },
        "status": {
            "session_status": "active",
            "registration_status": "open",
            "public_direct_booking": True,
            "source_of_truth": "enrollware_ical",
            "last_updated_at": now_iso,
        },
        "flags": {
            "needs_review": mapping_status != "mapped",
            "has_missing_start_time": start_at is None,
            "has_missing_registration_url": registration_url is None,
            "has_missing_course_name": not bool(parsed.course_name_primary_clean or course_title),
        },
    }


def build_sessions_from_enrollware_ical(
    *,
    repo_root: Path,
    ical_url: str,
    output_path: Path,
    audit_dir: Path,
    course_map: dict[str, Any],
) -> int:
    previous_by_id = previous_sessions_by_id(output_path)
    now_iso = datetime.now(TZ).isoformat()
    text = fetch_ical_text(ical_url)
    events = parse_ical_events(text)
    sessions: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    unmapped_rows: list[dict[str, Any]] = []
    unavailable_rows: list[dict[str, Any]] = []

    for event in events:
        session_id = stable_ical_session_id(event)
        session = build_session_from_ical_event(
            event,
            now_iso,
            course_map,
            previous_session=previous_by_id.get(session_id),
        )
        if not session:
            skipped.append(
                {
                    "uid": clean_string(event.get("uid")),
                    "summary": strip_tags(event.get("summary")),
                    "reason": "missing_uid_or_start",
                }
            )
            continue
        registration_unavailable_reason = enrollment_url_unavailable_reason(session.get("commerce", {}).get("registration_url"))
        registration_status = registration_status_from_unavailable_reason(registration_unavailable_reason)
        if registration_status in {"closed", "full"}:
            status = session.setdefault("status", {})
            status["registration_status"] = registration_status
            status["public_direct_booking"] = False
            status["registration_status_source"] = "enrollware_registration_page"
            status["registration_status_reason"] = registration_unavailable_reason
            session["registration_status"] = registration_status
            session["public_direct_booking"] = False
            session["registration_status_source"] = "enrollware_registration_page"
            session["registration_status_reason"] = registration_unavailable_reason
            unavailable_rows.append(
                {
                    "session_id": session.get("session_id"),
                    "raw_course": session.get("course", {}).get("course_name_primary_clean"),
                    "start_at": session.get("timing", {}).get("start_at"),
                    "location": session.get("location", {}).get("location_display"),
                    "registration_url": session.get("commerce", {}).get("registration_url"),
                    "reason": registration_unavailable_reason,
                }
            )
        sessions.append(session)
        if session.get("mapping_status") != "mapped":
            unmapped_rows.append(
                {
                    "session_id": session.get("session_id"),
                    "raw_course": session.get("course", {}).get("course_name_primary_clean"),
                    "start_at": session.get("timing", {}).get("start_at"),
                    "location": session.get("location", {}).get("location_display"),
                    "mapping_status": session.get("mapping_status"),
                    "mapping_notes": session.get("mapping_notes", []),
                }
            )

    sessions.sort(key=lambda item: (item.get("timing", {}).get("start_at") or "", item.get("session_id") or ""))
    enrollment_counts = {"student_snapshot_classes": 0, "student_snapshot_matches": 0, "seated_students_matched": 0}
    student_snapshot_path = repo_root / "data" / "enrollware_student_snapshot.json"
    if student_snapshot_path.exists():
        from scripts.import_enrollware_student_report import apply_snapshot_to_sessions
        enrollment_counts = apply_snapshot_to_sessions(
            sessions,
            json.loads(student_snapshot_path.read_text(encoding="utf-8")),
        )
    current_ids = {str(session.get("session_id")) for session in sessions if session.get("session_id") is not None}
    prior_ids = set(previous_by_id)
    removed_ids = sorted(prior_ids - current_ids)
    added_ids = sorted(current_ids - prior_ids)

    output = {
        "build": {
            "generated_at": now_iso,
            "repo_root": str(repo_root),
            "source_mode": "enrollware_ical_authoritative",
            "source": "enrollware_ical",
            "course_map_path": course_map.get("_path"),
            "inputs": {
                "enrollware_ical_url": ical_url,
                "prior_sessions_current": str(output_path),
            },
            "counts": {
                "ical_events_read": len(events),
                "sessions_written": len(sessions),
                "skipped_events": len(skipped),
                "registration_unavailable_sessions_marked_not_direct_bookable": len(unavailable_rows),
                "unmapped_sessions": len(unmapped_rows),
                "prior_sessions_read": len(previous_by_id),
                "sessions_added_compared_with_prior_source": len(added_ids),
                "classes_removed_compared_with_prior_source": len(removed_ids),
                "stale_sheet_or_class_report_sessions_excluded": len(removed_ids),
                **enrollment_counts,
            },
        },
        "sessions": sessions,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8")

    audit_dir.mkdir(parents=True, exist_ok=True)
    removed_examples = [
        {
            "session_id": session_id,
            "course_name": (previous_by_id.get(session_id, {}).get("course", {}) or {}).get("course_name_primary_clean"),
            "start_at": (previous_by_id.get(session_id, {}).get("timing", {}) or {}).get("start_at"),
            "location": (previous_by_id.get(session_id, {}).get("location", {}) or {}).get("location_display"),
        }
        for session_id in removed_ids[:50]
    ]
    summary = {
        "generated_at": now_iso,
        "source": "enrollware_ical",
        "ical_events_read": len(events),
        "public_sessions_created": len(sessions),
        "skipped_events": len(skipped),
        "registration_unavailable_sessions_marked_not_direct_bookable": len(unavailable_rows),
        "unmapped_sessions": len(unmapped_rows),
        "prior_sessions_read": len(previous_by_id),
        "sessions_added_compared_with_prior_source": len(added_ids),
        "classes_removed_compared_with_prior_source": len(removed_ids),
        "stale_zapier_sheet_sessions_excluded": 0,
        "stale_manual_class_report_sessions_excluded": len(removed_ids),
        **enrollment_counts,
        "removed_examples": removed_examples,
    }
    (audit_dir / "enrollware_ical_import_summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    report_lines = [
        "# Enrollware iCal Import Report",
        "",
        f"- Generated at: `{now_iso}`",
        "- Source: `enrollware_ical`",
        f"- iCal events read: `{len(events)}`",
        f"- Public sessions created: `{len(sessions)}`",
        f"- Skipped events: `{len(skipped)}`",
        f"- Registration unavailable sessions marked not direct-bookable: `{len(unavailable_rows)}`",
        f"- Unmapped sessions: `{len(unmapped_rows)}`",
        f"- Prior sessions read: `{len(previous_by_id)}`",
        f"- Classes removed compared with prior source: `{len(removed_ids)}`",
        f"- Stale manual/Class Report sessions excluded: `{len(removed_ids)}`",
        "- Stale Zapier/Google Sheet public schedule sessions excluded: `0`",
        "",
        "Zapier/Google Sheet registration events are not used as the public schedule source in this build path.",
        "They remain separate registration-signal/audit inputs only.",
        "",
        "## Removed Examples",
        "",
    ]
    if removed_examples:
        report_lines.append("| Session ID | Course | Start | Location |")
        report_lines.append("|---|---|---|---|")
        for item in removed_examples[:20]:
            report_lines.append(
                f"| `{item.get('session_id')}` | {item.get('course_name') or ''} | {item.get('start_at') or ''} | {item.get('location') or ''} |"
            )
    else:
        report_lines.append("- None")
    report_lines.extend(["", "## Registration Unavailable Exclusions", ""])
    if unavailable_rows:
        report_lines.append("| Session ID | Reason | Course | Start | Location |")
        report_lines.append("|---|---|---|---|---|")
        for item in unavailable_rows[:50]:
            report_lines.append(
                f"| `{item.get('session_id')}` | {item.get('reason') or ''} | {item.get('raw_course') or ''} | {item.get('start_at') or ''} | {item.get('location') or ''} |"
            )
    else:
        report_lines.append("- None")
    report_lines.extend(["", "## Unmapped Examples", ""])
    if unmapped_rows:
        for row in unmapped_rows[:20]:
            report_lines.append(f"- `{row.get('session_id')}`: {row.get('raw_course')} ({row.get('start_at')})")
    else:
        report_lines.append("- None")
    (audit_dir / "enrollware_ical_import_report.md").write_text("\n".join(report_lines) + "\n", encoding="utf-8")

    print(f"Wrote {output_path}")
    print(f"iCal events read: {len(events)}")
    print(f"Public sessions created: {len(sessions)}")
    print(f"Registration unavailable sessions marked not direct-bookable: {len(unavailable_rows)}")
    print(f"Classes removed compared with prior source: {len(removed_ids)}")
    print(f"Wrote {audit_dir / 'enrollware_ical_import_summary.json'}")
    print(f"Wrote {audit_dir / 'enrollware_ical_import_report.md'}")
    return 0


@dataclass
class ParsedCourseBlob:
    course_name_raw: Optional[str]
    course_name_primary_raw: Optional[str]
    course_name_primary_clean: Optional[str]
    course_tail_raw: Optional[str]
    course_subtitle_text: Optional[str]
    course_number: Optional[str]
    course_code_hint: Optional[str]
    certifying_body_hint: Optional[str]
    source_hint: Optional[str]
    price_hint: Optional[float]
    delivery_mode_hint: Optional[str]
    image_src: Optional[str]
    parse_confidence: str
    parse_notes: list[str]


def parse_longdesc_tokens(longdesc: Optional[str]) -> dict[str, str]:
    result: dict[str, str] = {}
    if not longdesc:
        return result
    for token in longdesc.split("|"):
        if ":" not in token:
            continue
        key, value = token.split(":", 1)
        result[key.strip().lower()] = value.strip()
    return result


def parse_course_blob(raw: Optional[str]) -> ParsedCourseBlob:
    notes: list[str] = []
    raw = clean_string(raw)

    if raw is None:
        return ParsedCourseBlob(
            course_name_raw=None,
            course_name_primary_raw=None,
            course_name_primary_clean=None,
            course_tail_raw=None,
            course_subtitle_text=None,
            course_number=None,
            course_code_hint=None,
            certifying_body_hint=None,
            source_hint=None,
            price_hint=None,
            delivery_mode_hint=None,
            image_src=None,
            parse_confidence="low",
            parse_notes=["empty_course_name"],
        )

    parts = re.split(r"<br\s*/?>", raw, maxsplit=1, flags=re.IGNORECASE)
    if len(parts) == 2:
        primary_raw, tail_raw = parts
    else:
        primary_raw, tail_raw = raw, None
        notes.append("no_br_found")

    primary_clean = strip_tags(primary_raw)

    subtitle_text = None
    if tail_raw:
        no_img_tail = re.sub(r"<img\b[^>]*>", " ", tail_raw, flags=re.IGNORECASE)
        subtitle_text = strip_tags(no_img_tail)
        if subtitle_text == primary_clean:
            subtitle_text = None

    image_src = None
    img_match = re.search(r'<img\b[^>]*\bsrc=["\']([^"\']+)["\']', raw, flags=re.IGNORECASE)
    if img_match:
        image_src = img_match.group(1).strip()

    longdesc = None
    longdesc_match = re.search(r'\blongdesc=["\']([^"\']+)["\']', raw, flags=re.IGNORECASE)
    if longdesc_match:
        longdesc = longdesc_match.group(1).strip()
    else:
        notes.append("no_longdesc_found")

    tokens = parse_longdesc_tokens(longdesc)
    course_number = tokens.get("r")
    course_code_hint = tokens.get("t")
    certifying_body_hint = tokens.get("cb")
    source_hint = tokens.get("src")
    delivery_mode_hint = tokens.get("d")

    price_hint = None
    if "p" in tokens:
        try:
            price_hint = float(tokens["p"])
        except Exception:
            notes.append("price_hint_parse_failed")

    if not course_number:
        notes.append("course_number_missing")

    confidence = "low"
    if primary_clean and course_number and (course_code_hint or certifying_body_hint):
        confidence = "high"
    elif primary_clean:
        confidence = "medium"

    return ParsedCourseBlob(
        course_name_raw=raw,
        course_name_primary_raw=clean_whitespace(primary_raw),
        course_name_primary_clean=primary_clean,
        course_tail_raw=clean_whitespace(tail_raw) if tail_raw else None,
        course_subtitle_text=subtitle_text,
        course_number=course_number,
        course_code_hint=course_code_hint,
        certifying_body_hint=certifying_body_hint,
        source_hint=source_hint,
        price_hint=price_hint,
        delivery_mode_hint=delivery_mode_hint,
        image_src=image_src,
        parse_confidence=confidence,
        parse_notes=notes,
    )


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [dict(row) for row in csv.DictReader(f)]


def read_class_report_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_string(h) or f"col_{i}" for i, h in enumerate(rows[0])]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        rec = {}
        for i, header in enumerate(headers):
            rec[header] = row[i] if i < len(row) else None
        out.append(rec)
    return out


def index_students_by_class_id(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        class_id = clean_string(row.get("Class ID"))
        if class_id:
            grouped[class_id].append(row)
    return grouped


def dedupe_classes_rows(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    deduped: dict[str, dict[str, Any]] = {}
    for row in rows:
        internal_id = clean_string(row.get("Internal ID"))
        if not internal_id:
            internal_id = extract_enroll_id_from_url(clean_string(row.get("Registration Link")))
        if internal_id:
            deduped[internal_id] = row
    return deduped


def compute_registered_count(student_rows: list[dict[str, Any]]) -> int:
    return len(student_rows)


def load_course_map(repo_root: Path, relative_path: str) -> dict[str, Any]:
    path = (repo_root / relative_path).resolve()
    if not path.exists():
        raise SystemExit(f"Required course map not found: {path}")
    payload = json.loads(path.read_text(encoding="utf-8"))
    payload["_path"] = str(path)
    payload.setdefault("courses_by_id", {})
    payload.setdefault("courses_by_number", {})
    return payload


def resolve_course_mapping(
    course_map: dict[str, Any],
    *,
    course_id: Optional[str],
    course_number: Optional[str],
    raw_course_title: Optional[str] = None,
) -> tuple[Optional[dict[str, Any]], str, list[str]]:
    by_id = course_map.get("courses_by_id", {})
    by_number = course_map.get("courses_by_number", {})
    notes: list[str] = []
    id_key = clean_string(course_id)
    num_key = clean_string(course_number)
    map_by_id = by_id.get(id_key) if id_key else None
    map_by_number = by_number.get(num_key) if num_key else None
    if isinstance(map_by_number, str):
        map_by_number = by_id.get(map_by_number)
    if map_by_id and map_by_number:
        if clean_string(map_by_id.get("course_id")) != clean_string(map_by_number.get("course_id")):
            notes.append("course_map_conflict:id_and_number_resolve_different_courses")
            return map_by_id, "mapped_conflict", notes
        return map_by_id, "mapped", notes
    if map_by_id:
        return map_by_id, "mapped", notes
    if map_by_number:
        return map_by_number, "mapped", notes
    title_key = normalize_title(raw_course_title)
    if title_key:
        for entry in by_id.values():
            if not isinstance(entry, dict) or entry.get("active") is False:
                continue
            title_values = [
                entry.get("official_title"),
                entry.get("clean_title"),
                entry.get("course_key"),
                *(entry.get("title_aliases", []) or []),
            ]
            if any(normalize_title(value) == title_key for value in title_values if value):
                notes.append("course_map_title_alias")
                return entry, "mapped", notes
    return None, "unmapped", notes


def build_session_from_class_report(
    report_row: dict[str, Any],
    class_patch: Optional[dict[str, Any]],
    student_rows: list[dict[str, Any]],
    now_iso: str,
    course_map: dict[str, Any],
) -> Optional[dict[str, Any]]:
    registration_url_report = clean_string(report_row.get("Registration Link"))
    session_id = extract_enroll_id_from_url(registration_url_report)

    if not session_id:
        session_id = clean_string(report_row.get("ID"))

    if not session_id:
        return None

    course_raw = None
    if class_patch:
        course_raw = clean_string(class_patch.get("Course Name"))
    if not course_raw:
        course_raw = clean_string(report_row.get("Course"))

    parsed = parse_course_blob(course_raw)

    course_id = clean_string(class_patch.get("Course Id")) if class_patch else None
    if not course_id:
        course_id = parsed.course_number
    course_number = parsed.course_number or course_id

    mapped, mapping_status, mapping_notes = resolve_course_mapping(
        course_map,
        course_id=course_id,
        course_number=course_number,
        raw_course_title=parsed.course_name_primary_clean or parsed.course_name_raw,
    )
    if mapped and not course_id:
        course_id = clean_string(mapped.get("course_id"))
    if mapped and not course_number:
        course_number = clean_string(mapped.get("course_number") or mapped.get("course_id"))

    mapped_family = clean_string(mapped.get("family")) if mapped else None
    mapped_subtype = clean_string(mapped.get("subtype")) if mapped else None
    mapped_certifying_body = clean_string(mapped.get("certifying_body")) if mapped else None
    mapped_delivery_mode = clean_string(mapped.get("delivery_mode")) if mapped else None
    mapped_course_code = clean_string(mapped.get("course_code")) if mapped else None
    mapped_logo_key = clean_string(mapped.get("logo_key")) if mapped else None
    mapped_price = parse_float(mapped.get("price")) if mapped else None
    mapped_clean_title = clean_string(mapped.get("clean_title")) if mapped else None
    mapped_short_description = clean_string(mapped.get("short_description")) if mapped else None
    mapped_long_description = clean_string(mapped.get("long_description")) if mapped else None
    mapped_who_class_for = clean_string(mapped.get("who_class_for") or mapped.get("who_for")) if mapped else None
    mapped_prerequisites = clean_string(mapped.get("prerequisites")) if mapped else None
    mapped_what_to_expect = clean_string(mapped.get("what_to_expect")) if mapped else None
    mapped_certification_card = clean_string(mapped.get("certification_card")) if mapped else None
    mapped_renewal_info = clean_string(mapped.get("renewal_info")) if mapped else None
    mapped_official_description_source = clean_string(mapped.get("official_description_source")) if mapped else None
    mapped_description_status = normalize_description_status(mapped.get("description_status") if mapped else None)

    start_source = report_row.get("Start Date / Time")
    end_source = report_row.get("End Date / Time")
    start_at = parse_datetime_flexible(start_source)
    end_at = parse_datetime_flexible(end_source)

    max_students = None
    if class_patch:
        max_students = parse_int(class_patch.get("Max Students"))
    if max_students is None:
        # In class report, Seats appears to be total seat capacity in this export
        max_students = parse_int(report_row.get("Seats"))

    students_count_raw = parse_int(report_row.get("Students"))
    registered_count = compute_registered_count(student_rows)

    # Fallback to class report student count if no student rows exist
    if registered_count == 0 and students_count_raw is not None:
        registered_count = students_count_raw

    available_seats = None
    is_full = None
    if max_students is not None:
        available_seats = max_students - registered_count
        if available_seats < 0:
            available_seats = 0
        is_full = available_seats <= 0

    price = mapped_price
    if price is None and class_patch:
        price = parse_float(class_patch.get("Price"))
    if price is None:
        price = parsed.price_hint

    registration_url = None
    if class_patch:
        registration_url = clean_string(class_patch.get("Registration Link"))
    if not registration_url:
        registration_url = registration_url_report

    location_name = None
    if class_patch:
        location_name = clean_string(class_patch.get("Location Name"))
    if not location_name:
        location_name = clean_string(report_row.get("Location"))

    client = None
    if class_patch:
        client = clean_string(class_patch.get("Client"))
    if not client:
        client = clean_string(report_row.get("Client"))

    lead_instructor_name = None
    if class_patch:
        lead_instructor_name = clean_string(class_patch.get("Lead Instructor Name"))
    if not lead_instructor_name:
        lead_instructor_name = clean_string(report_row.get("Instructor"))

    is_future = None
    is_past = None
    if start_at:
        try:
            start_dt = datetime.fromisoformat(start_at)
            now_dt = datetime.fromisoformat(now_iso)
            is_future = start_dt >= now_dt
            is_past = start_dt < now_dt
        except Exception:
            pass

    return {
        "session_id": session_id,
        "start": start_at,
        "end": end_at,
        "start_datetime": start_at,
        "end_datetime": end_at,
        "source_keys": {
            "class_report_id": clean_string(report_row.get("ID")),
            "class_sheet_internal_id": clean_string(class_patch.get("Internal ID")) if class_patch else None,
            "student_sheet_class_id": session_id,
        },
        "course": {
            "course_id": course_id,
            "course_name_raw": parsed.course_name_raw,
            "course_name_primary_raw": parsed.course_name_primary_raw,
            "course_name_primary_clean": parsed.course_name_primary_clean,
            "course_tail_raw": parsed.course_tail_raw,
            "course_subtitle_text": parsed.course_subtitle_text,
            "course_number": course_number,
            "course_code_hint": parsed.course_code_hint or mapped_course_code,
            "certifying_body_hint": parsed.certifying_body_hint or mapped_certifying_body,
            "source_hint": parsed.source_hint,
            "price_hint": parsed.price_hint,
            "delivery_mode_hint": parsed.delivery_mode_hint or mapped_delivery_mode,
            "image_src": parsed.image_src,
            "parse_confidence": parsed.parse_confidence,
            "parse_notes": parsed.parse_notes,
            "mapped_family": mapped_family,
            "mapped_subtype": mapped_subtype,
            "mapped_certifying_body": mapped_certifying_body,
            "mapped_delivery_mode": mapped_delivery_mode,
            "mapped_logo_key": mapped_logo_key,
            "mapped_price": mapped_price,
            "mapped_clean_title": mapped_clean_title,
            "mapped_short_description": mapped_short_description,
            "mapped_long_description": mapped_long_description,
            "mapped_who_class_for": mapped_who_class_for,
            "mapped_prerequisites": mapped_prerequisites,
            "mapped_what_to_expect": mapped_what_to_expect,
            "mapped_certification_card": mapped_certification_card,
            "mapped_renewal_info": mapped_renewal_info,
            "mapped_official_description_source": mapped_official_description_source,
            "mapped_description_status": mapped_description_status,
            "mapping_status": mapping_status,
            "mapping_notes": mapping_notes,
        },
        "mapped_family": mapped_family,
        "mapped_subtype": mapped_subtype,
        "mapped_certifying_body": mapped_certifying_body,
        "mapped_delivery_mode": mapped_delivery_mode,
        "mapped_logo_key": mapped_logo_key,
        "mapped_price": mapped_price,
        "mapped_clean_title": mapped_clean_title,
        "mapped_short_description": mapped_short_description,
        "mapped_long_description": mapped_long_description,
        "mapped_who_class_for": mapped_who_class_for,
        "mapped_prerequisites": mapped_prerequisites,
        "mapped_what_to_expect": mapped_what_to_expect,
        "mapped_certification_card": mapped_certification_card,
        "mapped_renewal_info": mapped_renewal_info,
        "mapped_official_description_source": mapped_official_description_source,
        "mapped_description_status": mapped_description_status,
        "mapping_status": mapping_status,
        "mapping_notes": mapping_notes,
        "timing": {
            "start_at": start_at,
            "end_at": end_at,
            "start": start_at,
            "end": end_at,
            "start_datetime": start_at,
            "end_datetime": end_at,
            "start_source": clean_string(start_source),
            "end_source": clean_string(end_source),
            "timezone": "America/New_York",
            "is_future": is_future,
            "is_past": is_past,
        },
        "location": {
            "location_name": location_name,
            "location_display": location_name,
            "client": client,
        },
        "staffing": {
            "lead_instructor_name": lead_instructor_name,
        },
        "capacity": {
            "max_students": max_students,
            "students_count_raw": students_count_raw,
            "registered_count": registered_count,
            "available_seats": available_seats,
            "is_full": is_full,
        },
        "commerce": {
            "price": price,
            "registration_url": registration_url,
        },
        "status": {
            "session_status": "active",
            "source_of_truth": "class_report_backbone",
            "last_updated_at": now_iso,
        },
        "flags": {
            "needs_review": False,
            "has_missing_start_time": start_at is None,
            "has_missing_registration_url": registration_url is None,
            "has_missing_course_name": parsed.course_name_primary_clean is None,
        },
    }


def main() -> int:
    reporter = BuildStatusReporter("build_sessions_current")
    parser = argparse.ArgumentParser(description="Build merged sessions_current.json from raw files.")
    parser.add_argument("--repo-root", default=".", help="Path to repo root.")
    parser.add_argument(
        "--source",
        choices=["enrollware_ical", "class_report"],
        default="enrollware_ical",
        help="Public schedule source. Defaults to Enrollware iCal freshness source.",
    )
    parser.add_argument(
        "--ical-url",
        default=None,
        help="Enrollware iCal feed URL. Defaults to LANDER_ENROLLWARE_ICAL_URL or the configured 910CPR feed.",
    )
    parser.add_argument("--class-report", default=None)
    parser.add_argument("--classes-csv", default=None)
    parser.add_argument("--students-csv", default=None)
    parser.add_argument("--output", default="data/sessions_current.json")
    parser.add_argument("--course-map", default="data/config/course_map.json")
    parser.add_argument("--audit-dir", default="data/audit")
    parser.add_argument("--fail-on-unmapped", action="store_true")
    args = parser.parse_args()

    repo_root = Path(args.repo_root).resolve()
    output_path = (repo_root / args.output).resolve()
    audit_dir = (repo_root / args.audit_dir).resolve()
    course_map = load_course_map(repo_root, args.course_map)
    if args.source == "enrollware_ical":
        ical_url = clean_string(args.ical_url) or clean_string(os.environ.get("LANDER_ENROLLWARE_ICAL_URL")) or DEFAULT_ENROLLWARE_ICAL_URL
        reporter.set_context(
            inputs=[repo_root / args.course_map],
            outputs=[
                output_path,
                audit_dir / "enrollware_ical_import_summary.json",
                audit_dir / "enrollware_ical_import_report.md",
            ],
        )
        reporter.waiting(total=0)
        return build_sessions_from_enrollware_ical(
            repo_root=repo_root,
            ical_url=ical_url,
            output_path=output_path,
            audit_dir=audit_dir,
            course_map=course_map,
        )

    class_report = resolve_live_input_path(
        repo_root,
        label="Class report",
        cli_path=args.class_report,
        env_var="LANDER_CLASS_REPORT_PATH",
        private_path="data/private/enrollware/Class Report.xlsx",
        legacy_paths=CLASS_REPORT_LEGACY_PATHS,
    )
    classes_csv = resolve_live_input_path(
        repo_root,
        label="Classes CSV",
        cli_path=args.classes_csv,
        env_var="LANDER_CLASSES_CSV_PATH",
        private_path="data/private/raw/classes_raw_live.csv",
        legacy_paths=CLASSES_CSV_LEGACY_PATHS,
    )
    students_csv = resolve_live_input_path(
        repo_root,
        label="Students CSV",
        cli_path=args.students_csv,
        env_var="LANDER_STUDENTS_CSV_PATH",
        private_path="data/private/raw/students_raw_live.csv",
        legacy_paths=STUDENTS_CSV_LEGACY_PATHS,
    )
    class_report_path = class_report.path
    classes_csv_path = classes_csv.path
    students_csv_path = students_csv.path
    reporter.set_context(
        inputs=[class_report_path, classes_csv_path, students_csv_path, repo_root / args.course_map],
        outputs=[
            output_path,
            audit_dir / "unmapped_courses.json",
            audit_dir / "course_map_conflicts.json",
            audit_dir / "missing_descriptions.json",
            audit_dir / "missing_descriptions_report.md",
            audit_dir / "name_dependency_report.md",
        ],
    )

    missing_messages = []
    for resolved, private_path, env_var, cli_flag, legacy_paths in [
        (class_report, "data/private/enrollware/Class Report.xlsx", "LANDER_CLASS_REPORT_PATH", "--class-report", CLASS_REPORT_LEGACY_PATHS),
        (classes_csv, "data/private/raw/classes_raw_live.csv", "LANDER_CLASSES_CSV_PATH", "--classes-csv", CLASSES_CSV_LEGACY_PATHS),
        (students_csv, "data/private/raw/students_raw_live.csv", "LANDER_STUDENTS_CSV_PATH", "--students-csv", STUDENTS_CSV_LEGACY_PATHS),
    ]:
        print_resolved_path(resolved)
        if not resolved.path.exists():
            missing_messages.append(
                missing_live_input_message(
                    resolved,
                    private_path=private_path,
                    env_var=env_var,
                    cli_flag=cli_flag,
                    legacy_paths=legacy_paths,
                )
            )
    if missing_messages:
        raise SystemExit("\n\n".join(missing_messages))

    reporter.waiting(total=0)
    print(f"Reading class report: {class_report_path}")
    class_report_rows = read_class_report_rows(class_report_path)

    print(f"Reading classes csv: {classes_csv_path}")
    classes_rows = read_csv_rows(classes_csv_path)

    print(f"Reading students csv: {students_csv_path}")
    students_rows = read_csv_rows(students_csv_path)

    print("Indexing sources...")
    classes_index = dedupe_classes_rows(classes_rows)
    students_index = index_students_by_class_id(students_rows)

    now_iso = datetime.now(TZ).isoformat()

    sessions: list[dict[str, Any]] = []
    skipped_no_session_id = 0
    unmapped_rows: list[dict[str, Any]] = []
    conflict_rows: list[dict[str, Any]] = []
    mapped_course_rollup: dict[str, dict[str, Any]] = {}

    reporter.start(total=len(class_report_rows))
    for index, report_row in enumerate(class_report_rows, start=1):
        report_reg_link = clean_string(report_row.get("Registration Link"))
        session_id = extract_enroll_id_from_url(report_reg_link)
        if not session_id:
            session_id = clean_string(report_row.get("ID"))

        if not session_id:
            skipped_no_session_id += 1
            reporter.update(current=index, total=len(class_report_rows))
            continue

        class_patch = classes_index.get(session_id)
        student_rows_for_session = students_index.get(session_id, [])

        session = build_session_from_class_report(
            report_row=report_row,
            class_patch=class_patch,
            student_rows=student_rows_for_session,
            now_iso=now_iso,
            course_map=course_map,
        )
        if session:
            sessions.append(session)
            if session.get("mapping_status") == "mapped":
                course_obj = session.get("course", {})
                course_key = (
                    clean_string(course_obj.get("course_id"))
                    or clean_string(course_obj.get("course_number"))
                    or f"session:{session.get('session_id')}"
                )
                rollup = mapped_course_rollup.setdefault(
                    course_key,
                    {
                        "course_id": clean_string(course_obj.get("course_id")),
                        "course_number": clean_string(course_obj.get("course_number")),
                        "clean_title": clean_string(session.get("mapped_clean_title")),
                        "official_title": clean_string(course_obj.get("course_name_primary_clean")),
                        "family": clean_string(session.get("mapped_family")),
                        "subtype": clean_string(session.get("mapped_subtype")),
                        "certifying_body": clean_string(session.get("mapped_certifying_body")),
                        "delivery_mode": clean_string(session.get("mapped_delivery_mode")),
                        "description_status": normalize_description_status(session.get("mapped_description_status")),
                        "fields": {
                            "short_description": clean_string(session.get("mapped_short_description")),
                            "long_description": clean_string(session.get("mapped_long_description")),
                            "who_class_for": clean_string(session.get("mapped_who_class_for")),
                            "prerequisites": clean_string(session.get("mapped_prerequisites")),
                            "what_to_expect": clean_string(session.get("mapped_what_to_expect")),
                            "certification_card": clean_string(session.get("mapped_certification_card")),
                            "renewal_info": clean_string(session.get("mapped_renewal_info")),
                            "official_description_source": clean_string(session.get("mapped_official_description_source")),
                        },
                        "example_session_ids": [],
                        "affected_session_count": 0,
                    },
                )
                rollup["affected_session_count"] += 1
                if len(rollup["example_session_ids"]) < 5:
                    rollup["example_session_ids"].append(session.get("session_id"))
            if session.get("mapping_status") != "mapped":
                unmapped_rows.append(
                    {
                        "session_id": session.get("session_id"),
                        "course_id": session.get("course", {}).get("course_id"),
                        "course_number": session.get("course", {}).get("course_number"),
                        "course_name_display": session.get("course", {}).get("course_name_primary_clean"),
                        "mapping_status": session.get("mapping_status"),
                        "mapping_notes": session.get("mapping_notes", []),
                        "registration_url": session.get("commerce", {}).get("registration_url"),
                    }
                )
            if session.get("mapping_status") == "mapped_conflict":
                conflict_rows.append(
                    {
                        "session_id": session.get("session_id"),
                        "course_id": session.get("course", {}).get("course_id"),
                        "course_number": session.get("course", {}).get("course_number"),
                        "mapping_notes": session.get("mapping_notes", []),
                    }
                )
        else:
            skipped_no_session_id += 1
        reporter.update(current=index, total=len(class_report_rows))

    sessions.sort(
        key=lambda s: (
            s.get("timing", {}).get("start_at") or "",
            s.get("session_id") or "",
        )
    )

    output = {
        "build": {
            "generated_at": now_iso,
            "repo_root": str(repo_root),
            "course_map_path": course_map.get("_path"),
            "inputs": {
                "class_report": str(class_report_path),
                "classes_csv": str(classes_csv_path),
                "students_csv": str(students_csv_path),
            },
            "counts": {
                "class_report_rows": len(class_report_rows),
                "classes_csv_rows": len(classes_rows),
                "classes_deduped_sessions": len(classes_index),
                "students_csv_rows": len(students_rows),
                "sessions_written": len(sessions),
                "skipped_no_session_id": skipped_no_session_id,
                "unmapped_sessions": len(unmapped_rows),
                "mapping_conflicts": len(conflict_rows),
            },
        },
        "sessions": sessions,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8")

    audit_dir.mkdir(parents=True, exist_ok=True)
    unmapped_payload = {
        "generated_at": now_iso,
        "count": len(unmapped_rows),
        "items": unmapped_rows,
    }
    conflicts_payload = {
        "generated_at": now_iso,
        "count": len(conflict_rows),
        "items": conflict_rows,
    }
    missing_descriptions_items: list[dict[str, Any]] = []
    missing_field_counts: defaultdict[str, int] = defaultdict(int)
    complete_descriptions_count = 0
    for course_key in sorted(mapped_course_rollup.keys()):
        item = mapped_course_rollup[course_key]
        fields = item.get("fields", {})
        missing_fields: list[str] = []
        placeholder_fields: list[str] = []

        for field_name in DESCRIPTION_REQUIRED_FIELDS:
            value = clean_string(fields.get(field_name))
            if not value:
                missing_fields.append(field_name)
        for field_name in DESCRIPTION_FIELDS:
            if field_name in {"description_status"}:
                continue
            value = clean_string(fields.get(field_name))
            if value and is_placeholder_text(value):
                placeholder_fields.append(field_name)

        description_status = normalize_description_status(item.get("description_status"))
        if description_status in {"missing", "needs_review"}:
            missing_fields.append("description_status")

        if missing_fields or placeholder_fields:
            for field_name in set(missing_fields + placeholder_fields):
                missing_field_counts[field_name] += 1
            missing_descriptions_items.append(
                {
                    "course_id": item.get("course_id"),
                    "course_number": item.get("course_number"),
                    "clean_title": item.get("clean_title"),
                    "official_title": item.get("official_title"),
                    "family": item.get("family"),
                    "subtype": item.get("subtype"),
                    "certifying_body": item.get("certifying_body"),
                    "delivery_mode": item.get("delivery_mode"),
                    "description_status": description_status,
                    "missing_fields": sorted(set(missing_fields)),
                    "placeholder_fields": sorted(set(placeholder_fields)),
                    "example_session_ids": item.get("example_session_ids", []),
                    "count_of_affected_sessions": item.get("affected_session_count", 0),
                }
            )
        else:
            complete_descriptions_count += 1

    missing_descriptions_payload = {
        "generated_at": now_iso,
        "counts": {
            "mapped_courses": len(mapped_course_rollup),
            "courses_with_complete_descriptions": complete_descriptions_count,
            "courses_needing_review_or_missing": len(missing_descriptions_items),
            "sessions_affected": sum(int(i.get("count_of_affected_sessions", 0)) for i in missing_descriptions_items),
        },
        "missing_field_counts": dict(sorted(missing_field_counts.items(), key=lambda kv: (-kv[1], kv[0]))),
        "items": missing_descriptions_items,
    }
    (audit_dir / "unmapped_courses.json").write_text(
        json.dumps(unmapped_payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    (audit_dir / "course_map_conflicts.json").write_text(
        json.dumps(conflicts_payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    (audit_dir / "missing_descriptions.json").write_text(
        json.dumps(missing_descriptions_payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    missing_report_lines = [
        "# Missing Descriptions Report",
        "",
        f"- Generated at: `{now_iso}`",
        f"- Total mapped courses: `{len(mapped_course_rollup)}`",
        f"- Courses with complete descriptions: `{complete_descriptions_count}`",
        f"- Courses needing review: `{len(missing_descriptions_items)}`",
        f"- Sessions affected: `{missing_descriptions_payload['counts']['sessions_affected']}`",
        "",
        "## Top Missing Fields",
        "",
    ]
    if missing_field_counts:
        for field_name, count in sorted(missing_field_counts.items(), key=lambda kv: (-kv[1], kv[0])):
            missing_report_lines.append(f"- `{field_name}`: `{count}`")
    else:
        missing_report_lines.append("- None")
    missing_report_lines.extend(
        [
            "",
            "## Next Cleanup Recommendations",
            "",
            "- Fill required fields first: short_description, long_description, who_class_for.",
            "- Mark reviewed courses with description_status=official only after QA.",
            "- Keep unmapped courses blocked from public routing until mapped.",
            "",
        ]
    )
    (audit_dir / "missing_descriptions_report.md").write_text("\n".join(missing_report_lines), encoding="utf-8")
    name_dependency_report = (
        "# Name Dependency Report\n\n"
        "This build keeps legacy Course HTML parsing for backward compatibility, but mapping is authoritative for operational metadata.\n\n"
        f"- Course map: `{course_map.get('_path')}`\n"
        f"- Sessions written: `{len(sessions)}`\n"
        f"- Unmapped sessions: `{len(unmapped_rows)}`\n"
        f"- Mapping conflicts: `{len(conflict_rows)}`\n\n"
        "## Phase 1 Rules\n\n"
        "- Do not infer certifying body/family/delivery from display name when mapping is missing.\n"
        "- Display text may come from cleaned course name.\n"
        "- Operational fields must come from course map when available.\n"
    )
    (audit_dir / "name_dependency_report.md").write_text(name_dependency_report, encoding="utf-8")

    print(f"Wrote {output_path}")
    print(f"Sessions written: {len(sessions)}")
    print(f"Skipped no session id: {skipped_no_session_id}")
    print(f"Unmapped sessions: {len(unmapped_rows)}")
    print(f"Course map conflicts: {len(conflict_rows)}")
    print(f"Wrote {audit_dir / 'unmapped_courses.json'}")
    print(f"Wrote {audit_dir / 'course_map_conflicts.json'}")
    print(f"Wrote {audit_dir / 'missing_descriptions.json'}")
    print(f"Wrote {audit_dir / 'missing_descriptions_report.md'}")
    print(f"Wrote {audit_dir / 'name_dependency_report.md'}")

    if unmapped_rows:
        print("UNMAPPED COURSE DETECTED: one or more sessions are missing structured course mapping.")
        if args.fail_on_unmapped:
            reporter.error(
                current=len(class_report_rows),
                total=len(class_report_rows),
                last_output_file=output_path,
                message="Unmapped sessions found and --fail-on-unmapped was set.",
            )
            return 2

    warnings = []
    files_needing_review = []
    if unmapped_rows:
        warnings.append(f"{len(unmapped_rows)} unmapped sessions need course mapping.")
        files_needing_review.append(audit_dir / "unmapped_courses.json")
    if conflict_rows:
        warnings.append(f"{len(conflict_rows)} course map conflicts need review.")
        files_needing_review.append(audit_dir / "course_map_conflicts.json")
    if missing_descriptions_items:
        warnings.append(f"{len(missing_descriptions_items)} mapped courses need description review.")
        files_needing_review.append(audit_dir / "missing_descriptions.json")
    reporter.done(
        current=len(class_report_rows),
        total=len(class_report_rows),
        last_output_file=output_path,
        counts={
            "sessions_written": len(sessions),
            "class_report_rows": len(class_report_rows),
            "unmapped_sessions": len(unmapped_rows),
            "mapping_conflicts": len(conflict_rows),
            "courses_needing_description_review": len(missing_descriptions_items),
        },
        warnings=warnings,
        files_needing_review=files_needing_review,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
